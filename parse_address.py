# -*- coding: utf-8 -*-
"""
parse_address.py — Tách đơn vị hành chính + tên đường + thôn/xóm từ địa chỉ "bẩn".

Cách dùng:
    python3 parse_address.py input.xlsx [-c "Tên cột địa chỉ"] [-o output.xlsx]

Kết quả: file Excel với các cột
  - Tên đường (tách)          : chỉ tên đường, bỏ số nhà/hẻm/ngõ/ngách
  - Thôn/Xóm/Ấp (tách)        : thôn, xóm, ấp, bản, buôn, khu phố, tổ, đội...
  - Phường/Xã, Quận/Huyện, Tỉnh/TP (hệ cũ 63 tỉnh - snapshot cuối trước 7/2025)
  - Phường/Xã MỚI, Tỉnh/TP MỚI (hệ 34 tỉnh sau sáp nhập)

Cần file dữ liệu vn_units_data.json đặt cùng thư mục với script.
"""
import argparse
import json
import os
import re
import sys
import unicodedata

# ---------------------------------------------------------------- normalize

def strip_diacritics(s: str) -> str:
    s = s.replace("đ", "d").replace("Đ", "D")
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def norm_text(s: str) -> str:
    s = strip_diacritics(str(s).lower())
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return " " + s.strip() + " "


UNIT_PREFIX = re.compile(
    r"^(thanh pho|tinh|quan|huyen|thi xa|thi tran|phuong|xa)\s+", re.I
)

TOKEN_RE = re.compile(r"[^\W_]+", re.UNICODE)


def split_prefix(norm_name: str):
    m = UNIT_PREFIX.match(norm_name.strip())
    if m:
        return m.group(1), norm_name.strip()[m.end():].strip()
    return "", norm_name.strip()


def num_variants(base: str):
    if base.isdigit():
        return list(dict.fromkeys([base, str(int(base)), str(int(base)).zfill(2)]))
    return [base]


def levenshtein(a: str, b: str, cutoff: int) -> int:
    if abs(len(a) - len(b)) > cutoff:
        return cutoff + 1
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(min(prev[j] + 1, cur[j - 1] + 1,
                           prev[j - 1] + (ca != cb)))
        if min(cur) > cutoff:
            return cutoff + 1
        prev = cur
    return prev[-1]

# ---------------------------------------------------------------- data model


class Unit:
    __slots__ = ("uid", "name", "parent", "aliases", "strong")

    def __init__(self, uid, name, parent=None):
        self.uid = uid
        self.name = name
        self.parent = parent
        self.aliases = []
        self.strong = []


def build_aliases(unit: Unit, kind: str):
    n = norm_text(unit.name).strip()
    prefix, base = split_prefix(n)
    weak, strong = set(), {n}
    if kind == "province":
        for p in ("tinh", "thanh pho", "tp", "t p"):
            strong.add(f"{p} {base}")
        weak.add(base)
        if " " in base and len(base) >= 6:
            weak.add(base.replace(" ", ""))     # "gialai", "binhduong"...
    else:
        table = {
            "quan": ["quan", "q"], "huyen": ["huyen", "h"],
            "thi xa": ["thi xa", "tx"], "thanh pho": ["thanh pho", "tp"],
            "phuong": ["phuong", "p", "f"], "xa": ["xa"],
            "thi tran": ["thi tran", "tt"],
        }.get(prefix, [prefix] if prefix else [])
        for b in num_variants(base):
            for p in table:
                strong.add(f"{p} {b}")
                if len(p) == 1 and b.isdigit():
                    strong.add(f"{p}{b}")          # q5, p5, f5
        if not base.isdigit():
            weak.add(base)
    unit.strong = sorted(strong, key=len, reverse=True)
    unit.aliases = sorted(weak, key=len, reverse=True)


SPECIAL_PROVINCE_ALIASES = {
    "79": ["tp hcm", "tphcm", "hcm", "hcmc", "sai gon", "sg",
           "ho chi minh", "hochiminh"],
    "01": ["ha noi", "hn"],
    "48": ["da nang"],
    "77": ["ba ria vung tau", "brvt", "vung tau"],
    "46": ["thua thien hue", "tp hue", "thanh pho hue"],
}

BEFORE_BLACKLIST = {
    "thon", "ap", "xom", "ban", "buon", "khu", "kp", "tdp", "to",
    "duong", "pho", "ngo", "ngach", "hem", "so", "sn", "cau", "cho", "doi",
    "phuong", "xa", "tt", "tx",
}

UNIT_WORDS = {"quan", "huyen", "phuong", "xa", "tinh", "thanh", "pho",
              "thi", "tran", "tt", "tx", "tp",
              "thon", "xom", "ap", "ban", "kp", "to", "tdp", "doi", "khu"}

# với phường/xã, người dân hay ghi sai loại đơn vị ("xã" thay vì "thị trấn")
# nên không chặn các từ này khi match cấp xã
WARD_ALLOWED_PREV = {"phuong", "xa", "tt", "tx"}


def blacklisted_before(t: str, p: int, ward_mode=False) -> bool:
    toks = t[:p].split()
    if not toks:
        return False
    prev = toks[-1]
    if prev == "pho" and len(toks) >= 2 and toks[-2].endswith("thanh"):
        return False
    if ward_mode and prev in WARD_ALLOWED_PREV:
        return False
    return prev in BEFORE_BLACKLIST or prev.isdigit()


PAREN_OLD = re.compile(r"\([^)]*\bc[ũu]\b[^)]*\)", re.I)


def clean_text(raw) -> str:
    s = str(raw)
    s = PAREN_OLD.sub(" ", s)
    s = re.sub(r"(?<!\d)0\d{8,10}(?!\d)", " ", s)
    return s


def make_key(raw) -> str:
    """Khóa khử trùng lặp: cùng nội dung địa chỉ -> chỉ xử lý 1 lần."""
    return norm_text(clean_text(raw))


def find_alias(t: str, alias: str) -> int:
    return t.rfind(" " + alias + " ")


def overlaps(pos, length, spans) -> bool:
    s0, e0 = pos + 1, pos + 1 + length
    return any(s0 < p2 + 1 + l2 and p2 + 1 < e0 for p2, l2 in spans)

# ------------------------------------------------- trích tên đường/thôn xóm

NUMISH = re.compile(r"^\d+[a-z]{0,2}$")
# mã lô/block/căn hộ đứng đầu địa chỉ kiểu "B3", "A12", "C4A" (chữ cái đứng
# trước số, khác với NUMISH là số đứng trước) — vd "B3/24S Nguyễn Văn Linh..."
BLOCKISH = re.compile(r"^\d+[a-z]{0,2}$|^[a-z]\d{1,3}[a-z]?$")
ROUTE_TOKEN = re.compile(r"^(ql|tl|dt|hl)\d+\w*$")
BREAK_CHARS = set(",.;:()|–—-\n\r\"'")

HARD_STOP = {
    "phuong", "xa", "quan", "huyen", "tp", "tt", "tx", "tinh",
    "thon", "xom", "ap", "buon", "khom", "kp", "tdp", "tieu", "khu",
    "duong", "dg", "ngach", "hem", "kiet", "sn", "sdt", "dt",
    "kcn", "kdt", "kdc", "dc", "gan",
}
STREET_JUNK_FIRST = {"cc", "cty", "cong", "chung", "toa", "can", "lo",
                     "nha", "khach", "benh", "truong", "khong", "e",
                     "kien", "thung", "goi", "chai", "hop", "bich", "kg",
                     "bo", "cai", "vnd", "trieu", "combo", "set", "don",
                     "sl", "ky", "thu", "ten",
                     # ghi chú giao hàng hay đứng ngay sau số nhà khi địa chỉ
                     # viết theo kiểu "Số 10, giao hàng giờ HC, Quận 1..."
                     "giao", "ship", "gui", "gio", "hen"}

# Đơn vị thôn/xóm viết LIỀN với số, không có khoảng trắng (vd "Kp22", "To5",
# "Tdp03", "Ap3A"). Tokenizer gộp chữ+số liền nhau thành 1 token duy nhất
# ("kp22"), nên phải nhận diện riêng bằng regex — nếu không, các hàm capture
# tên đường/tên khác sẽ nuốt luôn cụm này vì không thấy nó khớp với danh sách
# từ dừng (HARD_STOP) vốn chỉ chứa dạng tách rời ("kp").
MERGED_HAMLET_RE = re.compile(
    r"^(thon|xom|ap|ban|buon|khom|kp|tdp|to|doi)(\d[\da-z]{0,3})$")
MERGED_HAMLET_UNIT = {
    "thon": "Thôn", "xom": "Xóm", "ap": "Ấp", "ban": "Bản", "buon": "Buôn",
    "khom": "Khóm", "kp": "Khu phố", "tdp": "TDP", "to": "Tổ", "doi": "Đội",
}

# token kiểu giá tiền / cân nặng: 390k, 10kg, 1tr800k, 750k...
MONEY_WEIGHT = re.compile(r"^\d+[.,]?\d*(k|kg|g|gr|tr|ky|d|vnd)\w*$")


def is_junky_detail(v):
    """Kiểm tra street/hamlet TRƯỚC KHI học vào danh bạ SĐT (--hoc) — chặn
    giá trị rõ ràng là rác đơn hàng (giá tiền kiểu "390k") lọt vào, để về
    sau không bị đối chiếu SĐT khách cũ tái sử dụng nhầm giá trị rác này
    đè lên kết quả tách đúng của các đơn hàng khác cùng SĐT."""
    if not v:
        return False
    words = [strip_diacritics(w.lower()) for w in str(v).split()]
    return any(MONEY_WEIGHT.match(w) for w in words)


class DetailExtractor:
    """Bóc tên đường và thôn/xóm/ấp từ phần địa chỉ chi tiết
    (sau khi đã che các cụm đơn vị hành chính nhận diện được)."""

    def __init__(self, s, toks, norms, admin_mask):
        self.s = s
        self.toks = toks          # (word, start, end)
        self.norms = norms
        self.lows = [unicodedata.normalize("NFC", w.lower()) for w, _, _ in toks]
        self.admin = admin_mask
        self.n = len(toks)
        # vị trí cụm hành chính cuối cùng: số nhà + đường luôn đứng TRƯỚC nó
        self.last_admin = max((i for i, m in enumerate(admin_mask) if m),
                              default=-1)

    def sep(self, k):
        if k + 1 >= self.n:
            return ""
        return self.s[self.toks[k][2]:self.toks[k + 1][1]]

    def brk(self, k):
        return any(ch in BREAK_CHARS for ch in self.sep(k))

    def is_stop(self, k):
        nk, lo = self.norms[k], self.lows[k]
        nxt_digit = k + 1 < self.n and self.norms[k + 1].isdigit()
        if nk in HARD_STOP:
            return True
        if MERGED_HAMLET_RE.match(nk):
            return True
        if lo in ("bản", "làng", "ngõ", "phố", "số", "ấp"):
            return True
        if nk in ("to", "doi", "ban", "ngo", "so") and nxt_digit:
            return True
        if lo in ("tổ", "đội") and nxt_digit:
            return True
        if nk == "thanh" and k + 1 < self.n and self.norms[k + 1] == "pho":
            return True
        if nk == "thi" and k + 1 < self.n and self.norms[k + 1] in ("tran", "xa"):
            return True
        return False

    def pretty(self, i, j):
        """Chuỗi gốc từ token i..j, chuẩn hóa hoa/thường."""
        raw = self.s[self.toks[i][1]:self.toks[j][2]]
        words = raw.split()
        out = []
        for w in words:
            if w.isdigit() or w in ("/",):
                out.append(w)
            elif any(c.isdigit() for c in w) and len(w) <= 6:
                out.append(w.upper())
            else:
                out.append(w[:1].upper() + w[1:].lower())
        return " ".join(out)

    def capture(self, k, maxn=4, allow_digits="any"):
        """Thu thập token tên từ vị trí k. allow_digits:
        'any' = số ở đâu cũng được (thôn/xóm: Xóm Máy 2, Ấp 7A),
        'lead' = chỉ chấp nhận số nếu tên bắt đầu bằng số (đường 30 tháng 4),
        'no' = dừng khi gặp số (tên đường sau số nhà)."""
        got = []
        while k < self.n and len(got) < maxn:
            if self.admin[k] or self.is_stop(k):
                break
            nk = self.norms[k]
            if MONEY_WEIGHT.match(nk):
                break               # 390k, 10kg... là giá/cân nặng, không phải tên
            if nk.isdigit():
                if allow_digits == "no":
                    break
                if allow_digits == "lead" and got \
                        and self.norms[k - 1] != "thang" \
                        and not self.norms[got[0]].isdigit():
                    break
            got.append(k)
            if self.brk(k):
                break
            k += 1
        return got

    def run(self):
        streets_kw, streets_num, hamlets = [], [], []
        seen_h = set()
        k = 0
        while k < self.n:
            if self.admin[k]:
                k += 1
                continue
            nk, lo = self.norms[k], self.lows[k]
            nxt = self.norms[k + 1] if k + 1 < self.n else ""
            nxt_digit = nxt.isdigit()

            # ---------- thôn/xóm/ấp/... viết LIỀN với số (Kp22, To5, Tdp03)
            mm = MERGED_HAMLET_RE.match(nk)
            if mm:
                pfx, num = mm.groups()
                unit_m = MERGED_HAMLET_UNIT[pfx]
                # Đã viết liền số (kp22, to5...) nên đây chắc chắn là đơn vị
                # đánh số -> chỉ lấy đúng token này, không "vợt" thêm chữ phía
                # sau (tránh nuốt nhầm tên đường, giống lý do ở nhánh dưới).
                val = f"{unit_m} {num.upper()}"
                key = strip_diacritics(val.lower())
                if key not in seen_h:
                    seen_h.add(key)
                    hamlets.append((self.toks[k][1], val))
                k += 1
                continue

            # ---------- "Khu Phố<số>" — "Khu" viết rời nhưng "Phố5" bị dính
            # liền số (khác case Kp22 ở trên vì đây là 2 từ "Khu"+"Phố<số>")
            if nk == "khu" and nxt[:3] == "pho" and nxt[3:].isdigit():
                val = f"Khu phố {nxt[3:]}"
                key = strip_diacritics(val.lower())
                if key not in seen_h:
                    seen_h.add(key)
                    hamlets.append((self.toks[k][1], val))
                k += 2
                continue

            # ---------- thôn/xóm/ấp/bản/buôn/khu phố/tổ/đội
            unit, start = None, k + 1
            if lo == "thôn" or nk == "thon":
                unit = "Thôn"
            elif lo == "xóm" or nk == "xom":
                unit = "Xóm"
            elif lo == "ấp" or nk == "ap":
                unit = "Ấp"
            elif lo == "bản" or (nk == "ban" and nxt_digit):
                unit = "Bản"
            elif lo == "buôn" or nk == "buon":
                unit = "Buôn"
            elif lo == "khóm" or nk == "khom":
                unit = "Khóm"
            elif lo == "làng":
                unit = "Làng"
            elif nk == "kp":
                unit = "Khu phố"
            elif nk == "khu" and nxt == "pho":
                unit, start = "Khu phố", k + 2
            elif nk == "tieu" and nxt == "khu":
                unit, start = "Tiểu khu", k + 2
            elif nk == "to" and nxt == "dan" and k + 2 < self.n \
                    and self.norms[k + 2] == "pho":
                unit, start = "TDP", k + 3
            elif nk == "tdp":
                unit = "TDP"
            elif nk == "khu" and nxt not in ("do", "vuc", "cong", "che"):
                unit = "Khu"
            elif lo == "tổ" or (nk == "to" and nxt_digit):
                unit = "Tổ"
            elif lo == "đội" or (nk == "doi" and nxt_digit):
                unit = "Đội"
            if unit:
                if start < self.n and NUMISH.match(self.norms[start]):
                    # Đơn vị đánh SỐ (Ấp 6, Tổ 5, Ấp 2A...) hầu như luôn
                    # KHÔNG có tên riêng kèm theo -> chỉ lấy đúng con số, tránh
                    # nuốt nhầm chữ phía sau (thường là tên đường, vd "Ấp 6
                    # Trần Văn Giàu" không được để lẫn "Trần Văn" vào hamlet).
                    got = [start]
                else:
                    got = self.capture(start, maxn=3, allow_digits="any")
                if got:
                    val = f"{unit} {self.pretty(got[0], got[-1])}"
                    key = strip_diacritics(val.lower())
                    if key not in seen_h:
                        seen_h.add(key)
                        hamlets.append((self.toks[k][1], val))
                    k = got[-1] + 1
                    continue
                k += 1
                continue

            # ---------- đường/phố/quốc lộ...
            sk = None       # vị trí bắt đầu tên đường
            if lo in ("đường", "duong", "đg", "dg", "duòng", "durong"):
                sk = k + 1
            elif (lo in ("phố", "pho")) and \
                    (k == 0 or self.norms[k - 1] not in ("thanh", "khu")):
                sk = k + 1
            elif (nk, nxt) in (("quoc", "lo"), ("tinh", "lo"), ("huong", "lo"),
                               ("dai", "lo"), ("xa", "lo"), ("cao", "toc")):
                got = self.capture(k + 2, maxn=3, allow_digits="lead")
                if got:
                    streets_kw.append(
                        (self.toks[k][1],
                         f"{self.pretty(k, k + 1)} {self.pretty(got[0], got[-1])}"))
                    k = got[-1] + 1
                    continue
            elif ROUTE_TOKEN.match(nk):
                streets_kw.append((self.toks[k][1], nk.upper()))
                k += 1
                continue
            elif nk in ("ql", "tl", "hl", "dt") and nxt.isdigit() \
                    and len(nxt) <= 4:
                streets_kw.append((self.toks[k][1], f"{nk.upper()} {nxt}"))
                k += 2
                continue
            if sk is not None:
                if sk < self.n and self.norms[sk] == "so" \
                        and sk + 1 < self.n and self.norms[sk + 1].isdigit():
                    streets_kw.append(
                        (self.toks[k][1], f"Số {self.norms[sk + 1]}"))
                    k = sk + 2
                    continue
                got = self.capture(sk, maxn=4, allow_digits="lead")
                if got and self.norms[got[0]] not in STREET_JUNK_FIRST:
                    streets_kw.append(
                        (self.toks[k][1], self.pretty(got[0], got[-1])))
                    k = got[-1] + 1
                    continue
                k += 1
                continue

            # ---------- ngõ/hẻm/ngách/kiệt + số -> tên đường phía sau
            if nk in ("ngo", "ngach", "nghach", "hem", "kiet") and nxt_digit:
                j = k + 1
                while j < self.n and NUMISH.match(self.norms[j]):
                    j += 1
                if j < self.n and not self.admin[j] and not self.is_stop(j) \
                        and self.norms[j].isalpha() and not self.brk(j - 1):
                    got = self.capture(j, maxn=4, allow_digits="no")
                    if got and self.norms[got[0]] not in STREET_JUNK_FIRST:
                        streets_num.append(
                            (self.toks[k][1], self.pretty(got[0], got[-1])))
                        k = got[-1] + 1
                        continue
                k = j
                continue

            # ---------- số nhà ở đầu đoạn -> tên đường phía sau
            seg_start = (k == 0 or self.brk(k - 1) or self.admin[k - 1]
                         or self.norms[k - 1] in ("dc", "tai", "diachi")
                         or self.lows[k - 1] == "ở"
                         or (self.norms[k - 1] == "chi" and k >= 2
                             and self.norms[k - 2] == "dia"))
            if seg_start and self.last_admin >= 0 and k > self.last_admin:
                seg_start = False   # sau cụm hành chính chỉ còn ghi chú đơn hàng
            if seg_start:
                j = k
                if self.norms[j] in ("so", "sn", "lo"):
                    j += 1
                    if j < self.n and self.norms[j] == "nha":
                        j += 1
                cnt = 0
                while j < self.n and BLOCKISH.match(self.norms[j]) \
                        and not (cnt and self.brk(j - 1)):
                    j += 1
                    cnt += 1
                if cnt and j < self.n and len(self.norms[j]) == 1 \
                        and self.norms[j].isalpha():
                    j += 1      # "312 E Võ Văn Hát" -> bỏ chữ cái lẻ của số nhà
                if cnt and j < self.n and not self.admin[j] \
                        and not self.is_stop(j) and self.norms[j].isalpha():
                    # Cho phép có dấu phẩy/ngắt câu giữa số nhà và tên đường
                    # (vd "Số 25, Đồng Bông, Thôn..." — mẫu rất phổ biến khi
                    # địa chỉ được ghi theo từng trường cách nhau bằng dấu
                    # phẩy). capture() vẫn tự dừng đúng ở dấu ngắt câu TIẾP
                    # THEO nên không lo nuốt lan sang trường sau tên đường.
                    got = self.capture(j, maxn=4, allow_digits="no")
                    if got and self.norms[got[0]] not in STREET_JUNK_FIRST \
                            and len(self.pretty(got[0], got[-1])) >= 4:
                        streets_num.append(
                            (self.toks[k][1], self.pretty(got[0], got[-1])))
                        k = got[-1] + 1
                        continue
                if cnt:
                    k = j
                    continue
            k += 1

        streets = streets_kw or streets_num
        street = streets[0][1] if streets else None
        hamlets.sort(key=lambda x: x[0])
        hamlet = ", ".join(v for _, v in hamlets[:3]) if hamlets else None
        return street, hamlet

# ---------------------------------------------------------------- parser


class Parser:
    def __init__(self, data_path):
        with open(data_path, encoding="utf-8") as f:
            d = json.load(f)

        self.provinces = {}
        for p in d["oldP"]:
            u = Unit(p["idProvince"], p["name"])
            build_aliases(u, "province")
            for extra in SPECIAL_PROVINCE_ALIASES.get(p["idProvince"], []):
                if extra not in u.strong:
                    u.strong.append(extra)
            self.provinces[u.uid] = u

        self.districts = {}
        self.districts_by_prov = {}
        for dd in d["oldD"]:
            u = Unit(dd["idDistrict"], dd["name"], parent=dd["idProvince"])
            build_aliases(u, "district")
            self.districts[u.uid] = u
            self.districts_by_prov.setdefault(dd["idProvince"], []).append(u)

        self.wards = {}
        self.wards_by_dist = {}
        for w in d["oldC"]:
            u = Unit(w["idCommune"], w["name"], parent=w["idDistrict"])
            build_aliases(u, "ward")
            self.wards[u.uid] = u
            self.wards_by_dist.setdefault(w["idDistrict"], []).append(u)

        self.ward_mig = {}
        for m in d["mig"]:
            if m.get("status") == "mapped" and m.get("mappings"):
                mp = m["mappings"][0]
                self.ward_mig[m["oldWardCode"]] = (
                    mp["newWardName"], mp["newProvinceName"])
        self.prov_mig = {p["idProvince"]: p["name"] for p in d["newP"]}
        for rec in d["mergedP"]:
            for old in rec["mergedFrom"]:
                self.prov_mig[old["oldProvinceId"]] = rec["newProvinceName"]

        self.new_prov_by_name = {p["name"]: p["idProvince"] for p in d["newP"]}
        self.new_wards_by_prov = {}
        for w in d["newC"]:
            u = Unit(w["idCommune"], w["name"], parent=w["idProvince"])
            build_aliases(u, "ward")
            self.new_wards_by_prov.setdefault(w["idProvince"], []).append(u)

    # ---------------- exact matching

    @staticmethod
    def _collect(t, units, exclude=(), ward_mode=False):
        cands = []
        for u in units:
            found = None
            for a in u.strong:
                p = find_alias(t, a)
                if p >= 0 and not overlaps(p, len(a), exclude):
                    if found is None or (p, len(a)) > (found[0], found[1]):
                        found = (p, len(a), 2)
            if found is None:
                for a in u.aliases:
                    p = find_alias(t, a)
                    if p < 0 or overlaps(p, len(a), exclude):
                        continue
                    if blacklisted_before(t, p, ward_mode):
                        continue
                    if found is None or (p, len(a)) > (found[0], found[1]):
                        found = (p, len(a), 1)
            if found is None:
                continue
            p, ln, s = found
            after = t[p + ln + 2:].lstrip()
            if after.startswith("cu "):
                s = 3
            elif after.startswith("moi "):
                s = 0
            cands.append((u, p, ln, s))
        return cands

    @classmethod
    def _best(cls, t, units, before=None, by_position=False, exclude=(),
              ward_mode=False):
        cands = cls._collect(t, units, exclude, ward_mode)
        if not cands:
            return None, -1, 0
        if before is not None and before >= 0:
            pre = [c for c in cands if c[1] < before]
            if pre:
                cands = pre
        if by_position:
            cands.sort(key=lambda c: (c[3], c[1], c[2]), reverse=True)
        else:
            cands.sort(key=lambda c: (c[3], c[2], c[1]), reverse=True)
        u, p, ln, s = cands[0]
        return u, p, ln

    # ---------------- fuzzy matching

    @staticmethod
    def _fuzzy(t, units, before=None, exclude=(), min_len=6, max_cutoff=1,
               tail_only=False):
        toks = t.split()
        offs, o = [], 0
        for tk in toks:
            offs.append(o)
            o += len(tk) + 1
        best = []
        for u in units:
            for alias in u.aliases:
                if len(alias) < min_len:
                    continue
                wc = alias.count(" ") + 1
                cutoff = 1 if len(alias) < 8 else max_cutoff
                for i in range(len(toks) - wc + 1):
                    prev = toks[i - 1] if i > 0 else ""
                    prev2 = toks[i - 2] if i > 1 else ""
                    if prev == "pho" and prev2.endswith("thanh"):
                        pass                    # "thành phố X" viết sai chính tả
                    elif prev in BEFORE_BLACKLIST or prev.isdigit():
                        continue
                    if toks[i] in UNIT_WORDS or toks[i + wc - 1] in UNIT_WORDS:
                        continue
                    cand = " ".join(toks[i:i + wc])
                    if abs(len(cand) - len(alias)) > cutoff:
                        continue
                    if cand[0] != alias[0]:
                        continue
                    pos = offs[i]
                    if tail_only and pos < len(t) * 0.55:
                        continue
                    if overlaps(pos, len(cand), exclude):
                        continue
                    d = levenshtein(alias, cand, cutoff)
                    if d <= cutoff:
                        best.append((d, u, pos, len(cand)))
        if not best:
            return None
        best.sort(key=lambda x: (x[0], -x[2]))
        d0 = best[0][0]
        top = [b for b in best if b[0] == d0]
        if before is not None and before >= 0:
            pre = [b for b in top if b[2] < before]
            if pre:
                top = pre
        names = {norm_text(b[1].name) for b in top}
        if len(names) > 1:
            return None
        b = max(top, key=lambda x: x[2])
        return b[1], b[2], b[3]

    # ---------------- cleaning

    def clean(self, raw: str) -> str:
        return clean_text(raw)

    # ---------------- main parse

    def parse(self, raw: str):
        res = {"street": None, "hamlet": None,
               "ward": None, "district": None, "province": None,
               "ward_new": None, "province_new": None, "note": []}
        if raw is None or not str(raw).strip():
            res["note"].append("địa chỉ trống")
            return self._finish(res, None, None, None)

        s = self.clean(raw)
        toks = [(m.group(), m.start(), m.end()) for m in TOKEN_RE.finditer(s)]
        norms = [strip_diacritics(w.lower()) for w, _, _ in toks]
        t = " "
        offs = []
        for nn in norms:
            offs.append(len(t) - 1)
            t += nn + " "
        adm_spans = []      # (pos, len) trong t — để che khi bóc tên đường

        spans = []

        # 1) tỉnh/thành
        prov, ppos, pln = self._best(
            t, self.provinces.values(), by_position=True)
        if prov is None:
            f = self._fuzzy(t, self.provinces.values(), tail_only=True)
            if f:
                cand_prov, cand_pos, cand_ln = f
                # Xác thực chéo trước khi tin: chỉ chấp nhận tỉnh khớp GẦN
                # ĐÚNG nếu tìm được ít nhất 1 quận/huyện thuộc tỉnh đó cũng
                # xuất hiện trong văn bản. Nếu không, rất dễ bị 1 từ tình cờ
                # gần giống tên tỉnh kéo cả địa chỉ sang tỉnh sai hoàn toàn
                # (vd "An Thới, Bình Thủy" (Cần Thơ) bị fuzzy match nhầm
                # thành "Thái Bình" vì "thới" và "thái" chỉ lệch 1 ký tự).
                dist_pool = self.districts_by_prov.get(cand_prov.uid, [])
                if self._collect(t, dist_pool):
                    prov, ppos, pln = cand_prov, cand_pos, cand_ln
                    res["note"].append("tỉnh/TP khớp gần đúng")
        if prov and ppos >= 0:
            spans.append((ppos, pln))
            adm_spans.append((ppos, pln))

        # 2) quận/huyện
        dist, dpos, dln = None, -1, 0
        if prov:
            dist, dpos, dln = self._best(
                t, self.districts_by_prov.get(prov.uid, []),
                before=ppos, exclude=spans)
            if dist is None:
                f = self._fuzzy(t, self.districts_by_prov.get(prov.uid, []),
                                before=ppos, exclude=spans)
                if f:
                    dist, dpos, dln = f
                    res["note"].append("quận/huyện khớp gần đúng")
        else:
            hits = [c for c in self._collect(t, self.districts.values())
                    if c[3] >= 2 or (c[3] == 1 and c[2] >= 6)]
            if hits:
                hits.sort(key=lambda c: (c[3], c[1], c[2]), reverse=True)
                u0, p0, l0, s0 = hits[0]
                base0 = split_prefix(norm_text(u0.name).strip())[1]
                same = [h for h in hits
                        if split_prefix(norm_text(h[0].name).strip())[1] == base0]
                provs = {h[0].parent for h in same}
                if len(provs) == 1:
                    dist, dpos, dln = u0, p0, l0
                    prov, ppos = self.provinces[dist.parent], -1
                    res["note"].append("suy ra tỉnh/TP từ quận/huyện")
        if dist and dpos >= 0:
            spans.append((dpos, dln))
            adm_spans.append((dpos, dln))

        # 3) phường/xã
        ward, wpos, wln = None, -1, 0
        anchor = dpos if dpos >= 0 else ppos
        if dist:
            pool = self.wards_by_dist.get(dist.uid, [])
            ward, wpos, wln = self._best(t, pool, before=anchor, exclude=spans,
                                         ward_mode=True)
            if ward is None:
                f = self._fuzzy(t, pool, before=anchor, exclude=spans,
                                max_cutoff=2)
                if f:
                    ward, wpos, wln = f
                    res["note"].append("phường/xã khớp gần đúng")
        if ward is None and prov:
            pool = [w for du in self.districts_by_prov.get(prov.uid, [])
                    for w in self.wards_by_dist.get(du.uid, [])]
            cands = self._collect(t, pool, exclude=spans, ward_mode=True)
            pre = [c for c in cands if anchor < 0 or c[1] < anchor]
            if pre:
                cands = pre
            if cands:
                cands.sort(key=lambda c: (c[3], c[2], c[1]), reverse=True)
                u0, p0, l0, s0 = cands[0]
                same = [c for c in cands
                        if norm_text(c[0].name) == norm_text(u0.name)]
                dists = {c[0].parent for c in same}
                if len(dists) == 1 and (s0 >= 2 or dist is None):
                    if dist is None or u0.parent == dist.uid:
                        ward, wpos, wln = u0, p0, l0
                        if dist is None:
                            dist = self.districts[u0.parent]
                            res["note"].append("suy ra quận/huyện từ phường/xã")
        if prov is None:
            hits = [c for c in self._collect(t, self.wards.values())
                    if c[3] >= 2 and c[2] >= 8]
            if hits:
                hits.sort(key=lambda c: (c[3], c[2], c[1]), reverse=True)
                u0, p0, l0, _ = hits[0]
                same = [h for h in hits
                        if norm_text(h[0].name) == norm_text(u0.name)]
                dup = [h for h in hits if h not in same and h[2] == l0]
                dists = {h[0].parent for h in same}
                if not dup and len(dists) == 1:
                    ward, wpos, wln = u0, p0, l0
                    dist = self.districts[ward.parent]
                    prov = self.provinces[dist.parent]
                    res["note"].append("suy ra tỉnh/TP từ phường/xã")
        if ward and wpos >= 0:
            adm_spans.append((wpos, wln))

        res = self._finish(res, ward, dist, prov)

        # 4) pass hệ MỚI
        if res["province_new"] and not res["ward_new"]:
            npid = self.new_prov_by_name.get(res["province_new"])
            pool = self.new_wards_by_prov.get(npid, [])
            cands = self._collect(t, pool, exclude=spans, ward_mode=True)
            if cands:
                cands.sort(key=lambda c: (c[3], c[2], c[1]), reverse=True)
                u0 = cands[0][0]
                same = {norm_text(c[0].name) for c in cands
                        if c[2] == cands[0][2] and c[3] == cands[0][3]}
                if len(same) == 1:
                    res["ward_new"] = u0.name
                    adm_spans.append((cands[0][1], cands[0][2]))
                    res["note"] = (res["note"] + "; " if res["note"] else "") + \
                        "phường/xã nhận diện trực tiếp theo hệ ĐVHC mới"

        # 5) bóc tên đường + thôn/xóm từ phần còn lại
        admin_mask = [False] * len(toks)
        for pos, ln in adm_spans:
            for i, off in enumerate(offs):
                if off >= pos and off + 1 + len(norms[i]) <= pos + 1 + ln:
                    admin_mask[i] = True
        street, hamlet = DetailExtractor(s, toks, norms, admin_mask).run()
        res["street"], res["hamlet"] = street, hamlet
        return res

    def _finish(self, res, ward, dist, prov):
        if prov:
            res["province"] = prov.name
        else:
            res["note"].append("không tìm thấy tỉnh/TP trong địa chỉ")
        if dist:
            res["district"] = dist.name
        elif prov:
            res["note"].append("không tìm thấy quận/huyện")
        if ward:
            res["ward"] = ward.name
        elif prov:
            res["note"].append("không tìm thấy phường/xã")

        if ward and ward.uid in self.ward_mig:
            res["ward_new"], res["province_new"] = self.ward_mig[ward.uid]
        elif prov:
            res["province_new"] = self.prov_mig.get(prov.uid)
        res["note"] = "; ".join(res["note"])
        return res

# ------------------------------------------------- SĐT + danh bạ khách cũ

PHONE_RE = re.compile(r"(?:\+?84|0)[\d .\-]{8,13}")


def extract_phone(raw):
    for m in PHONE_RE.finditer(str(raw)):
        digits = re.sub(r"\D", "", m.group())
        if digits.startswith("84"):
            digits = "0" + digits[2:]
        if 10 <= len(digits) <= 11 and digits.startswith("0"):
            return digits
    return None


def mask_phone(raw) -> str:
    """Che số điện thoại trong văn bản TRƯỚC KHI gửi cho AI qua API,
    tránh lộ thông tin cá nhân khách hàng ra bên ngoài.
    Giữ lại chữ số đầu để AI vẫn nhận biết đây là một dãy số/SĐT,
    còn lại thay bằng 'x' (vd 0912345678 -> 0xxxxxxxxx)."""
    def _repl(m):
        raw_match = m.group()
        digits = re.sub(r"\D", "", raw_match)
        if digits.startswith("84"):
            digits = "0" + digits[2:]
        if not digits:
            return raw_match
        return digits[0] + "x" * (len(digits) - 1)
    return PHONE_RE.sub(_repl, str(raw))


ENTRY_FIELDS = ("street", "hamlet", "ward", "district", "province",
                "ward_new", "province_new")


def base_name(s):
    if not s:
        return ""
    b = split_prefix(norm_text(s).strip())[1]
    return re.sub(r"\b0(\d)\b", r"\1", b)


def lookup_labels(parser, wname, dname, pname):
    """Từ 3 tên nhãn có sẵn -> tên chuẩn trong danh mục + mapping hệ mới."""
    prov = next((p for p in parser.provinces.values()
                 if base_name(p.name) == base_name(pname)), None) if pname else None
    dist = None
    if prov and dname:
        dist = next((d for d in parser.districts_by_prov.get(prov.uid, [])
                     if base_name(d.name) == base_name(dname)), None)
    ward = None
    if dist and wname:
        ward = next((w for w in parser.wards_by_dist.get(dist.uid, [])
                     if base_name(w.name) == base_name(wname)), None)
    out = {}
    if prov:
        out["province"] = prov.name
        out["province_new"] = parser.prov_mig.get(prov.uid)
    if dist:
        out["district"] = dist.name
    if ward:
        out["ward"] = ward.name
        if ward.uid in parser.ward_mig:
            out["ward_new"], out["province_new"] = parser.ward_mig[ward.uid]
    return out


# ------------------------------------------------- tầng AI (Claude API)

AI_SYSTEM = """Bạn là công cụ tách địa chỉ Việt Nam từ văn bản tự do (có thể lẫn tên người, SĐT, thông tin đơn hàng, sai chính tả, viết tắt, thiếu dấu).
Với mỗi địa chỉ, trả về các trường:
- ten_duong: CHỈ tên đường/phố/quốc lộ (không kèm số nhà, hẻm, ngõ, ngách). null nếu không có.
- thon_xom: thôn/xóm/ấp/bản/buôn/khu phố/tổ dân phố, kèm loại (vd "Thôn Quảng Hòa", "Ấp 3"). null nếu không có.
- phuong_xa: tên phường/xã/thị trấn kèm loại. null nếu không có.
- quan_huyen: tên quận/huyện/thị xã/TP thuộc tỉnh kèm loại. null nếu không có.
- tinh_tp: tên tỉnh/thành phố. null nếu không có.
Chỉ lấy thông tin CÓ trong văn bản, sửa lỗi chính tả hiển nhiên, không suy diễn thêm.
Trả về DUY NHẤT một mảng JSON, mỗi phần tử: {"idx": <số thứ tự>, "ten_duong": ..., "thon_xom": ..., "phuong_xa": ..., "quan_huyen": ..., "tinh_tp": ...}. Không giải thích gì thêm."""

# System prompt riêng cho model LOCAL (Ollama, vd qwen2.5:3b) — model nhỏ nên
# cần quy tắc rõ ràng hơn + ví dụ mẫu (few-shot) để giảm bịa đặt và bám sát
# định dạng JSON, bù lại phần năng lực yếu hơn Claude.
AI_SYSTEM_LOCAL = """Bạn là công cụ tách địa chỉ Việt Nam từ văn bản tự do (có thể lẫn tên người, SĐT, thông tin đơn hàng, sai chính tả, viết tắt, thiếu dấu).

QUY TẮC BẮT BUỘC:
1. CHỈ lấy thông tin CÓ THẬT trong văn bản đầu vào. TUYỆT ĐỐI KHÔNG suy diễn, không tự bịa thêm phường/xã/quận/huyện/tỉnh nếu văn bản không nói rõ.
2. Nếu không chắc chắn hoặc không tìm thấy trường nào, trả về null cho trường đó — KHÔNG đoán mò, KHÔNG điền đại một địa danh khác.
3. Chỉ sửa lỗi chính tả/viết tắt hiển nhiên của CHÍNH địa danh xuất hiện trong văn bản (vd "tp hcm"->"TP Hồ Chí Minh", "q1"->"Quận 1"), không đổi sang địa danh khác.
4. ten_duong: CHỈ tên đường/phố/quốc lộ, KHÔNG kèm số nhà/hẻm/ngõ/ngách.
5. thon_xom: kèm loại đơn vị (Thôn/Xóm/Ấp/Bản/Buôn/Khu phố/Tổ dân phố) + tên.
6. phuong_xa, quan_huyen: LUÔN kèm loại đơn vị hành chính (Phường/Xã/Thị trấn; Quận/Huyện/Thị xã/Thành phố).
7. tinh_tp: tên đầy đủ tỉnh/thành phố.
8. Mỗi "idx" trong output phải khớp đúng "idx" tương ứng trong input, không bỏ sót, không thêm idx lạ.

VÍ DỤ:
Input mục 0: "Nguyễn Văn A, 0912345678, 123 Nguyễn Trãi, P5, Q5, TPHCM"
-> {"idx":0,"ten_duong":"Nguyễn Trãi","thon_xom":null,"phuong_xa":"Phường 5","quan_huyen":"Quận 5","tinh_tp":"TP Hồ Chí Minh"}

Input mục 1: "giao hang 500k, Thôn Đông, xã Yên Sở, Hoài Đức, Hà Nội"
-> {"idx":1,"ten_duong":null,"thon_xom":"Thôn Đông","phuong_xa":"Xã Yên Sở","quan_huyen":"Huyện Hoài Đức","tinh_tp":"Hà Nội"}

Input mục 2: "ship cho chị Lan, đã đặt cọc, sdt 0987xxxxxx"
-> {"idx":2,"ten_duong":null,"thon_xom":null,"phuong_xa":null,"quan_huyen":null,"tinh_tp":null}

Trả về DUY NHẤT một object JSON dạng:
{"items": [{"idx": <số>, "ten_duong": ..., "thon_xom": ..., "phuong_xa": ..., "quan_huyen": ..., "tinh_tp": ...}, ...]}
Không markdown, không giải thích, không thêm chữ nào khác ngoài object JSON đó."""


def ai_call(batch, api_key, model):
    """batch: list (idx, raw_text). Trả dict idx -> item.
    SĐT trong văn bản được che (mask_phone) trước khi đưa vào request
    gửi cho API, tránh lộ thông tin cá nhân khách hàng."""
    import urllib.request
    user = "\n\n".join(f"### {i}\n{mask_phone(raw)[:500]}" for i, raw in batch)
    body = json.dumps({
        "model": model, "max_tokens": 4096, "temperature": 0,
        "system": AI_SYSTEM,
        "messages": [{"role": "user", "content": user}],
    }).encode()
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages", data=body,
        headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                 "content-type": "application/json"})
    with urllib.request.urlopen(req, timeout=120) as r:
        resp = json.load(r)
    text = "".join(b.get("text", "") for b in resp.get("content", []))
    m = re.search(r"\[.*\]", text, re.S)
    out = {}
    if m:
        for item in json.loads(m.group()):
            if isinstance(item, dict) and "idx" in item:
                out[item["idx"]] = item
    return out


def ai_call_ollama(batch, model, host, keep_alive="30m"):
    """batch: list (idx, raw_text). Trả dict idx -> item.
    Gọi model AI chạy LOCAL qua Ollama (http://localhost:11434 mặc định)
    thay cho Claude API -> MIỄN PHÍ, không tốn tiền theo token, dữ liệu
    không rời khỏi máy. Cần cài Ollama và `ollama pull <model>` trước.
    SĐT vẫn được che bằng mask_phone dù chạy local.

    Để chạy NHANH và CHÍNH XÁC hơn với model nhỏ (vd qwen2.5:3b):
    - format="json": ép Ollama chỉ sinh JSON hợp lệ (đỡ phải sinh thừa
      markdown/giải thích -> nhanh hơn, ít lỗi parse hơn).
    - keep_alive: giữ model luôn nằm sẵn trong RAM giữa các lô, tránh
      phải nạp lại (rất chậm) mỗi lần gọi.
    - num_predict giới hạn theo số địa chỉ trong lô, tránh model sinh
      lan man làm chậm vô ích.
    - system prompt AI_SYSTEM_LOCAL có quy tắc chặt + ví dụ mẫu (few-shot)
      để bù độ chính xác cho model nhỏ."""
    import urllib.request
    import urllib.error
    user = "\n\n".join(f"### {i}\n{mask_phone(raw)[:500]}" for i, raw in batch)
    body = json.dumps({
        "model": model,
        "messages": [
            {"role": "system", "content": AI_SYSTEM_LOCAL},
            {"role": "user", "content": user},
        ],
        "stream": False,
        "format": "json",
        "keep_alive": keep_alive,
        "options": {
            "temperature": 0,
            "top_p": 0.1,
            "num_predict": 80 * len(batch) + 200,
        },
    }).encode()
    req = urllib.request.Request(
        host.rstrip("/") + "/api/chat", data=body,
        headers={"content-type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=600) as r:
            resp = json.load(r)
    except urllib.error.URLError as exc:
        raise RuntimeError(
            f"Không kết nối được Ollama tại {host} ({exc}). "
            f"Cài Ollama (https://ollama.com), chạy `ollama pull {model}` "
            f"rồi đảm bảo Ollama đang chạy (`ollama serve`).") from exc
    text = resp.get("message", {}).get("content", "")
    out = {}

    def _collect(items):
        for item in items:
            if isinstance(item, dict) and "idx" in item:
                out[item["idx"]] = item

    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            _collect(parsed.get("items") or parsed.get("results") or [])
        elif isinstance(parsed, list):
            _collect(parsed)
    except json.JSONDecodeError:
        pass
    if not out:
        # dự phòng: model không tuân thủ format json tuyệt đối -> vẫn thử
        # mò mảng JSON trong text bằng regex như cách cũ
        m = re.search(r"\[.*\]", text, re.S)
        if m:
            try:
                _collect(json.loads(m.group()))
            except json.JSONDecodeError:
                pass
    return out


def snap(name, units, extra_pool=None):
    """Ép tên AI trả về vào danh mục chuẩn: khớp tên gốc, cho lệch 1 ký tự."""
    if not name:
        return None
    b = base_name(name)
    if not b:
        return None
    pool = list(units) + list(extra_pool or [])
    exact = [u for u in pool if base_name(u.name) == b]
    if exact:
        return exact[0]
    near = [u for u in pool if levenshtein(base_name(u.name), b, 1) <= 1]
    names = {base_name(u.name) for u in near}
    return near[0] if len(names) == 1 else None


def vtitle(s):
    return " ".join(w[:1].upper() + w[1:] for w in str(s).split()) if s else None


def ai_merge(parser, res, item):
    """Gộp kết quả AI vào res, mọi tên ĐVHC đều ép về danh mục chuẩn."""
    ai_prov = snap(item.get("tinh_tp"), parser.provinces.values())
    prov = ai_prov
    if prov is None and res["province"]:
        prov = next((p for p in parser.provinces.values()
                     if p.name == res["province"]), None)
    dist = snap(item.get("quan_huyen"),
                parser.districts_by_prov.get(prov.uid, [])) if prov else None
    ward = None
    if prov:
        if dist:
            ward = snap(item.get("phuong_xa"),
                        parser.wards_by_dist.get(dist.uid, []))
        else:
            pool = [w for du in parser.districts_by_prov.get(prov.uid, [])
                    for w in parser.wards_by_dist.get(du.uid, [])]
            ward = snap(item.get("phuong_xa"), pool)
            if ward:
                dist = parser.districts[ward.parent]
    changed = False
    if ward:
        res["ward"], res["district"] = ward.name, dist.name
        res["province"] = prov.name
        res["ward_new"] = res["province_new"] = None
        if ward.uid in parser.ward_mig:
            res["ward_new"], res["province_new"] = parser.ward_mig[ward.uid]
        changed = True
    elif ai_prov and (not res["province"]
                      or base_name(res["province"]) != base_name(ai_prov.name)):
        # AI đọc ra tỉnh khác với phỏng đoán của tầng luật -> tin AI
        # (chỉ chạy trên dòng độ tin cậy thấp), xoá xã/huyện cũ cho nhất quán
        res["province"] = ai_prov.name
        res["province_new"] = parser.prov_mig.get(ai_prov.uid)
        res["district"] = dist.name if dist else None
        res["ward"] = res["ward_new"] = None
        changed = True
    elif prov and dist and not res["district"]:
        res["district"] = dist.name
        changed = True
    # Chỉ lấy street/hamlet từ AI khi tầng LUẬT chưa tìm ra — tầng luật khi
    # đã tách được thì thường chính xác hơn (không suy diễn, không bịa), nhất
    # là với model AI local nhỏ (dễ tự ý gộp thêm "Ngõ/Ngách" vào tên đường,
    # hoặc trả về chuỗi bị cắt/lộn xộn). AI chỉ bổ sung khi rule bó tay.
    if item.get("ten_duong") and not res.get("street"):
        res["street"] = vtitle(item["ten_duong"])
        changed = True
    if item.get("thon_xom") and not res.get("hamlet"):
        res["hamlet"] = vtitle(item["thon_xom"])
        changed = True
    if changed:
        res["note"] = (res["note"] + "; " if res["note"] else "") + \
            "AI đọc, đã đối chiếu từ điển"
    return changed


def confidence(res, phone_hit, conflict):
    if conflict:
        return "Cần kiểm tra"
    if phone_hit:
        return "Cao (SĐT khách cũ)"
    fuzzy = "gần đúng" in (res["note"] or "")
    if res["ward"] and not fuzzy:
        return "Cao"
    if res["ward"] or res["province"]:
        return "Trung bình"
    return "Thấp"

# ---------------------------------------------------------------- I/O lớn

CSV_EXTS = (".csv", ".tsv")
XLSX_ROW_LIMIT = 1_048_575


def read_rows(path):
    """Generator: yield từng dòng (list). Dòng đầu là header. Hỗ trợ xlsx/csv."""
    if path.lower().endswith(CSV_EXTS):
        import csv
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
            sample = f.readline()
            delim = "\t" if path.lower().endswith(".tsv") else \
                (";" if sample.count(";") > sample.count(",") else ",")
            f.seek(0)
            for row in csv.reader(f, delimiter=delim):
                yield row
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        for row in wb.active.iter_rows(values_only=True):
            yield list(row)
        wb.close()


class Writer:
    """Ghi kết quả dạng stream: CSV (không giới hạn dòng) hoặc XLSX."""

    def __init__(self, path, header, addr_idx):
        self.path = path
        self.is_csv = path.lower().endswith(CSV_EXTS)
        if self.is_csv:
            import csv
            self._f = open(path, "w", newline="", encoding="utf-8-sig")
            self._w = csv.writer(self._f)
            self._w.writerow(header)
        else:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            from openpyxl.cell import WriteOnlyCell
            self._wb = openpyxl.Workbook(write_only=True)
            self._ws = self._wb.create_sheet("Kết quả tách")
            widths = {1: 6, addr_idx + 1: 55}
            for i in range(1, len(header) + 1):
                self._ws.column_dimensions[get_column_letter(i)].width = \
                    widths.get(i, 18)
            self._ws.freeze_panes = "A2"
            fill = PatternFill("solid", fgColor="1F4E79")
            font = Font(name="Arial", bold=True, color="FFFFFF")
            hdr = []
            for h in header:
                c = WriteOnlyCell(self._ws, value=h)
                c.fill, c.font = fill, font
                c.alignment = Alignment(vertical="center", wrap_text=True)
                hdr.append(c)
            self._ws.append(hdr)
            self._ncols = len(header)
            self._nrows = 1

    def writerow(self, row):
        if self.is_csv:
            self._w.writerow(["" if v is None else v for v in row])
        else:
            self._ws.append(row)
            self._nrows += 1

    def close(self):
        if self.is_csv:
            self._f.close()
        else:
            from openpyxl.utils import get_column_letter
            self._ws.auto_filter.ref = \
                f"A1:{get_column_letter(self._ncols)}{self._nrows}"
            self._wb.save(self.path)

# ------------------------------------------------- cache + đa nhân


def cache_open(path):
    import sqlite3
    con = sqlite3.connect(path)
    con.execute("CREATE TABLE IF NOT EXISTS c (k TEXT PRIMARY KEY, v TEXT)")
    return con


def cache_get_many(con, keys):
    out = {}
    ks = list(keys)
    for i in range(0, len(ks), 900):
        chunk = ks[i:i + 900]
        q = f"SELECT k, v FROM c WHERE k IN ({','.join('?' * len(chunk))})"
        for k, v in con.execute(q, chunk):
            out[k] = json.loads(v)
    return out


def cache_put_many(con, d):
    con.executemany(
        "INSERT OR REPLACE INTO c VALUES (?, ?)",
        [(k, json.dumps(v, ensure_ascii=False)) for k, v in d.items()])
    con.commit()


_WORKER_PARSER = None


def _init_worker(data_path):
    global _WORKER_PARSER
    _WORKER_PARSER = Parser(data_path)


def _parse_one(item):
    key, raw = item
    return key, _WORKER_PARSER.parse(raw)

# ---------------------------------------------------------------- main


OUT_COLS = ["SĐT (tách)", "Tên đường (tách)", "Thôn/Xóm/Ấp (tách)",
            "Phường/Xã (tách)", "Quận/Huyện (tách)", "Tỉnh/TP (tách)",
            "Phường/Xã MỚI", "Tỉnh/TP MỚI", "Độ tin cậy", "Ghi chú"]


def main():
    ap = argparse.ArgumentParser(description="Tách ĐVHC + tên đường/thôn xóm")
    ap.add_argument("input", help="File đầu vào (.xlsx / .csv / .tsv)")
    ap.add_argument("-c", "--column", default=None,
                    help="Tên cột chứa địa chỉ (mặc định: tự tìm cột có chữ 'địa chỉ')")
    ap.add_argument("-o", "--output", default=None,
                    help="File kết quả (.xlsx hoặc .csv; mặc định <input>_parsed)")
    ap.add_argument("--hoc", action="store_true",
                    help="Học danh bạ SĐT→địa chỉ từ các cột nhãn có sẵn trong file")
    ap.add_argument("--ai", action="store_true",
                    help="Ép bật tầng AI (mặc định: tự bật nếu có ANTHROPIC_API_KEY)")
    ap.add_argument("--khong-ai", action="store_true",
                    help="Tắt tầng AI dù có API key")
    ap.add_argument("--ai-model", default="claude-haiku-4-5-20251001")
    ap.add_argument("--ai-backend", choices=["anthropic", "ollama"],
                    default="anthropic",
                    help="anthropic = Claude API (trả phí theo token); "
                         "ollama = model AI chạy LOCAL trên máy (MIỄN PHÍ, "
                         "cần cài Ollama + `ollama pull <model>` trước)")
    ap.add_argument("--ollama-model", default="qwen2.5:3b",
                    help="Tên model khi dùng --ai-backend ollama "
                         "(mặc định qwen2.5:3b - nhẹ, phù hợp máy RAM hạn chế; "
                         "có thể đổi qwen2.5:7b/14b nếu máy khỏe hơn)")
    ap.add_argument("--ollama-keep-alive", default="30m",
                    help="Thời gian giữ model trong RAM giữa các lần gọi "
                         "(tránh nạp lại model mỗi lô -> nhanh hơn nhiều)")
    ap.add_argument("--ollama-host", default="http://localhost:11434",
                    help="Địa chỉ Ollama server (mặc định chạy trên máy)")
    ap.add_argument("--nhan", type=int, default=0,
                    help="Số tiến trình song song (mặc định: tự chọn theo CPU)")
    ap.add_argument("--no-cache", action="store_true",
                    help="Không dùng/ghi cache địa chỉ đã xử lý")
    args = ap.parse_args()

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    use_ollama = args.ai_backend == "ollama"
    if use_ollama:
        # backend local: không cần API key, không tốn tiền -> chỉ cần --ai
        args.ai = args.ai and not args.khong_ai
    else:
        if args.ai and not api_key:
            sys.exit("--ai cần API key: chạy  export ANTHROPIC_API_KEY=sk-ant-...  "
                      "(hoặc dùng --ai-backend ollama để chạy model local miễn phí)")
        # tự bật AI khi có key (trừ khi --khong-ai); không key -> chạy thuần luật
        args.ai = (args.ai or bool(api_key)) and not args.khong_ai
    if use_ollama and args.ai:
        print(f"Tầng AI: BẬT — backend LOCAL (Ollama, model {args.ollama_model}, "
              f"miễn phí, dữ liệu không rời máy)")
    elif args.ai:
        print("Tầng AI: BẬT (tự phát hiện API key)" if not sys.argv.count("--ai")
              else "Tầng AI: BẬT")
    else:
        print("Tầng AI: TẮT" + ("" if api_key else " (chưa có ANTHROPIC_API_KEY)"))

    import time
    t0 = time.time()
    here = os.path.dirname(os.path.abspath(__file__))
    data_path = os.path.join(here, "vn_units_data.json")
    if not os.path.exists(data_path):
        sys.exit("Thiếu file vn_units_data.json cạnh script.")
    parser = Parser(data_path)

    # ---------- lượt 1: quét file, gom địa chỉ duy nhất
    it = read_rows(args.input)
    header = [str(h) if h is not None else "" for h in next(it)]
    if args.column:
        try:
            col_idx = header.index(args.column)
        except ValueError:
            sys.exit(f"Không thấy cột '{args.column}'. Các cột: {header}")
    else:
        cands = [i for i, h in enumerate(header) if "dia chi" in norm_text(h)]
        if not cands:
            sys.exit(f"Không tự tìm được cột địa chỉ. Các cột: {header}. Dùng -c.")
        col_idx = cands[0]

    def find_col(*keys):
        for i, h in enumerate(header):
            if i != col_idx and all(k in norm_text(h) for k in keys):
                return i
        return None

    wc_ = find_col("phuong", "xa")
    dc_ = find_col("quan")
    if dc_ is None:
        dc_ = find_col("huyen")
    pc_ = find_col("tinh")
    if args.hoc and pc_ is None:
        sys.exit("--hoc: không thấy cột Tỉnh/Thành phố trong file.")

    unique = {}          # key -> raw đại diện
    learn = []           # (phone, ward, dist, prov, key)
    total = 0
    for row in it:
        total += 1
        row = list(row) + [None] * (len(header) - len(row))
        raw = row[col_idx]
        if raw is not None and str(raw).strip():
            unique.setdefault(make_key(raw), raw)
        if args.hoc and raw and row[pc_]:
            ph = extract_phone(raw)
            if ph:
                learn.append((ph, row[wc_] if wc_ is not None else None,
                              row[dc_] if dc_ is not None else None,
                              row[pc_], make_key(raw)))
    print(f"Đọc {total} dòng, {len(unique)} địa chỉ duy nhất "
          f"({time.time()-t0:.0f}s)")

    # ---------- cache: bỏ qua địa chỉ đã xử lý lần trước
    con = None
    resmap = {}
    if not args.no_cache:
        try:
            con = cache_open(os.path.join(here, "cache_diachi.sqlite"))
            resmap = cache_get_many(con, unique.keys())
            if resmap:
                print(f"Cache: dùng lại {len(resmap)} địa chỉ đã xử lý")
        except Exception as exc:
            print(f"Không dùng được cache ({exc}) - chạy không cache")
            con = None
            resmap = {}
    todo = [(k, unique[k]) for k in unique if k not in resmap]

    # ---------- parse phần còn lại (song song nếu nhiều)
    if todo:
        nproc = args.nhan or min(os.cpu_count() or 1, 8)
        if len(todo) >= 5000 and nproc > 1:
            from multiprocessing import Pool
            print(f"Parse {len(todo)} địa chỉ trên {nproc} nhân...")
            new = {}
            with Pool(nproc, initializer=_init_worker,
                      initargs=(data_path,)) as pool:
                for k, res in pool.imap_unordered(_parse_one, todo,
                                                  chunksize=500):
                    new[k] = res
        else:
            new = {k: parser.parse(raw) for k, raw in todo}
        resmap.update(new)
        if con is not None and new:
            try:
                cache_put_many(con, new)
            except Exception:
                pass
        print(f"Đã parse {len(todo)} địa chỉ mới ({time.time()-t0:.0f}s)")

    # ---------- danh bạ SĐT
    dir_path = os.path.join(here, "danh_ba_khach.json")
    directory = {}
    if os.path.exists(dir_path):
        with open(dir_path, encoding="utf-8") as f:
            directory = json.load(f)
    if args.hoc:
        added = 0
        for ph, wn, dn, pn, key in learn:
            entry = lookup_labels(parser, wn, dn, pn)
            if not entry:
                continue
            pres = resmap.get(key) or {}
            for f_ in ("street", "hamlet"):
                if pres.get(f_) and not is_junky_detail(pres[f_]):
                    entry[f_] = pres[f_]
            old = directory.get(ph, {})
            old.update(entry)
            old["n"] = old.get("n", 0) + 1
            directory[ph] = old
            added += 1
        with open(dir_path, "w", encoding="utf-8") as f:
            json.dump(directory, f, ensure_ascii=False)
        print(f"Đã học {added} dòng -> danh bạ có {len(directory)} SĐT")

    # ---------- tầng AI: chỉ trên địa chỉ DUY NHẤT còn khó
    if args.ai:
        hard = [k for k, res in resmap.items()
                if confidence(res, False, False) in ("Thấp", "Trung bình")]
        model_name = args.ollama_model if use_ollama else args.ai_model
        print(f"--ai: gửi {len(hard)} địa chỉ khó cho {model_name} "
              f"({'local/Ollama' if use_ollama else 'Claude API'})...")
        BATCH = 20
        done = 0
        ai_new = {}
        for b0 in range(0, len(hard), BATCH):
            chunk = hard[b0:b0 + BATCH]
            batch = [(i, unique[k]) for i, k in enumerate(chunk)]
            try:
                if use_ollama:
                    items = ai_call_ollama(batch, args.ollama_model,
                                           args.ollama_host,
                                           args.ollama_keep_alive)
                else:
                    items = ai_call(batch, api_key, args.ai_model)
            except Exception as exc:
                print(f"  lỗi ({'Ollama' if use_ollama else 'API'}, bỏ qua lô "
                      f"{b0 // BATCH + 1}): {exc}")
                continue
            for i, k in enumerate(chunk):
                if i in items and ai_merge(parser, resmap[k], items[i]):
                    resmap[k]["_ai"] = True
                    ai_new[k] = resmap[k]
                    done += 1
        if con is not None and ai_new:
            try:
                cache_put_many(con, ai_new)
            except Exception:
                pass
        print(f"--ai: cải thiện {done}/{len(hard)} địa chỉ")

    # ---------- lượt 2: ghi kết quả từng dòng
    out_path = args.output
    if not out_path:
        base, ext = os.path.splitext(args.input)
        ext = ext if ext.lower() in CSV_EXTS else ".xlsx"
        out_path = base + "_parsed" + ext
    if not out_path.lower().endswith(CSV_EXTS) and total > XLSX_ROW_LIMIT:
        out_path = os.path.splitext(out_path)[0] + ".csv"
        print(f"File > {XLSX_ROW_LIMIT} dòng, Excel không chứa nổi -> "
              f"chuyển sang CSV: {out_path}")
    writer = Writer(out_path, header + OUT_COLS, col_idx)

    EMPTY = {"street": None, "hamlet": None, "ward": None, "district": None,
             "province": None, "ward_new": None, "province_new": None,
             "note": "địa chỉ trống"}
    n_detail = n_full = n_partial = n_none = 0
    conf_count = {}
    it = read_rows(args.input)
    next(it)
    for row in it:
        row = list(row) + [None] * (len(header) - len(row))
        raw = row[col_idx]
        key = make_key(raw) if raw is not None and str(raw).strip() else None
        res = dict(resmap.get(key, EMPTY))
        phone = extract_phone(raw) if raw else None
        phone_hit = conflict = False
        if phone and phone in directory:
            e = directory[phone]
            if res["province"] and e.get("province") and \
                    base_name(res["province"]) != base_name(e["province"]):
                conflict = True
                res["note"] = (res["note"] + "; " if res["note"] else "") + \
                    "SĐT trùng khách cũ nhưng khác tỉnh - cần kiểm tra"
            else:
                phone_hit = True
                filled = [f_ for f_ in ENTRY_FIELDS
                          if not res.get(f_) and e.get(f_)]
                for f_ in filled:
                    res[f_] = e[f_]
                res["note"] = (res["note"] + "; " if res["note"] else "") + \
                    "đối chiếu SĐT khách cũ" + \
                    (" (bổ sung từ danh bạ)" if filled else "")
        if conflict:
            conf = "Cần kiểm tra"
        elif phone_hit:
            conf = "Cao (SĐT khách cũ)"
        elif res.get("_ai"):
            conf = "Cao (AI)" if res["ward"] else \
                ("Trung bình (AI)" if res["province"] else "Thấp")
        else:
            conf = confidence(res, False, False)
        conf_count[conf] = conf_count.get(conf, 0) + 1
        writer.writerow(list(row[:len(header)]) + [
            phone, res["street"], res["hamlet"],
            res["ward"], res["district"], res["province"],
            res["ward_new"], res["province_new"], conf, res["note"]])
        if res["street"] or res["hamlet"]:
            n_detail += 1
        if res["ward"]:
            n_full += 1
        elif res["province"]:
            n_partial += 1
        else:
            n_none += 1
    writer.close()
    if con is not None:
        con.close()

    print(f"Đã xử lý {total} dòng -> {out_path} ({time.time()-t0:.0f}s)")
    print(f"  Có tên đường/thôn xóm   : {n_detail} ({n_detail/total:.1%})")
    print(f"  Tách được tới phường/xã : {n_full} ({n_full/total:.1%})")
    print(f"  Chỉ tới tỉnh/quận-huyện : {n_partial} ({n_partial/total:.1%})")
    print(f"  Không tách được         : {n_none} ({n_none/total:.1%})")
    print("  Độ tin cậy:", ", ".join(
        f"{k}: {v}" for k, v in sorted(conf_count.items(), key=lambda x: -x[1])))


if __name__ == "__main__":
    main()
