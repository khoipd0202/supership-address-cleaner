# -*- coding: utf-8 -*-
"""
address_ui.py — UI làm sạch địa chỉ: upload file Excel → xử lý batch → tải kết quả.

Chạy:
    python3 address_ui.py            (mặc định cổng 8899)
    python3 address_ui.py 9000       (chọn cổng)

Rồi mở trình duyệt: http://localhost:8899
"""
import io
import json
import os
import re
import sys
import tempfile
import time
import traceback
import unicodedata
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

try:
    import openpyxl
except ImportError:
    print("Cần cài openpyxl: pip3 install openpyxl")
    sys.exit(1)

from parse_address import Parser, lookup_labels, strip_diacritics
from clean_address import (
    extract_level4_and_poi,
    parse_address_components,
    _titlecase_vn as titlecase_vn,
)

DATA = os.path.join(HERE, "vn_units_data.json")
PARSER = Parser(DATA)

# Store last result for download
_last_result = {"wb_bytes": None, "filename": ""}


def process_excel(file_bytes, filename):
    """Process uploaded Excel file and return (rows_for_ui, stats, output_bytes)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], {}, None

    hdr = list(rows[0])

    def find(*names):
        for i, h in enumerate(hdr):
            hn = strip_diacritics(str(h or "").lower())
            for nm in names:
                if nm in hn:
                    return i
        return None

    ci_addr = find("dia chi", "dia chi goc")
    ci_w = find("phuong/xa", "phuong", "xa/phuong")
    ci_d = find("quan/huyen", "quan", "huyen")
    ci_p = find("tinh")

    # Build output workbook
    out = openpyxl.Workbook()
    o = out.active
    o.title = "Địa chỉ sạch"
    o.append([
        "STT", "Địa chỉ gốc", "Điểm định vị (POI)", "Tên đường", "Cấp 4",
        "Phường/Xã", "Quận/Huyện", "Tỉnh/TP", "ĐỊA CHỈ SẠCH",
    ])

    ui_rows = []
    stats = {"full": 0, "detail": 0, "new": 0, "n": 0, "input_n": 0, "removed": 0}

    for idx, r in enumerate(rows[1:], 1):
        stats["input_n"] += 1
        raw = r[ci_addr] if ci_addr is not None else ""
        wname = r[ci_w] if ci_w is not None else None
        dname = r[ci_d] if ci_d is not None else None
        pname = r[ci_p] if ci_p is not None else None

        lab = lookup_labels(PARSER, wname, dname, pname)
        ward = lab.get("ward") or (str(wname).strip() if wname else None)
        dist = lab.get("district") or (str(dname).strip() if dname else None)
        prov = lab.get("province") or (str(pname).strip() if pname else None)
        ward_new = lab.get("ward_new")
        prov_new = lab.get("province_new")

        parsed = parse_address_components(
            raw, [wname, dname, pname, ward, dist, prov, ward_new, prov_new]
        )
        poi = parsed["poi"]
        street = parsed["street"]
        level4 = parsed["level4"]
        if not (poi or street or level4):
            stats["removed"] += 1
            continue
        parts = [p for p in [poi, street, level4, ward, dist, prov] if p]
        full = ", ".join(parts)
        o.append([idx, raw, poi, street, level4, ward, dist, prov, full])

        stats["n"] += 1
        if ward and dist and prov:
            stats["full"] += 1
        if poi or street or level4:
            stats["detail"] += 1
        if ward_new and prov_new:
            stats["new"] += 1

        ui_rows.append({
            "idx": idx,
            "raw": str(raw or ""),
            "poi": poi or "",
            "street": street or "",
            "l4": level4 or "",
            "ward": ward or "",
            "district": dist or "",
            "province": prov or "",
            "full": full,
        })

    # Save to bytes
    buf = io.BytesIO()
    out.save(buf)
    out_bytes = buf.getvalue()

    return ui_rows, stats, out_bytes


# ─────────────────────────── HTML PAGE ───────────────────────────

PAGE = r"""<!doctype html><html lang="vi"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Làm sạch địa chỉ</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
:root{
  --bg:#0a0e1a;--card:#141825;--card2:#1a1f30;--line:#252a3a;
  --txt:#e2e8f0;--mut:#7a8396;--acc:#6366f1;--acc2:#818cf8;
  --ok:#22c55e;--mid:#eab308;--low:#ef4444;
  --grad:linear-gradient(135deg,#6366f1 0%,#8b5cf6 50%,#a78bfa 100%);
}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Inter',system-ui,sans-serif;background:var(--bg);color:var(--txt);
  min-height:100vh}

.wrap{max-width:1400px;margin:0 auto;padding:24px 28px}

/* Header */
.header{display:flex;align-items:center;gap:14px;margin-bottom:28px}
.header h1{font-size:22px;font-weight:700;
  background:var(--grad);-webkit-background-clip:text;-webkit-text-fill-color:transparent}
.header .sub{color:var(--mut);font-size:13px}

/* Upload Zone */
.upload-zone{
  border:2px dashed var(--line);border-radius:16px;padding:48px 24px;
  text-align:center;cursor:pointer;transition:all .3s ease;
  background:var(--card);position:relative;overflow:hidden;
}
.upload-zone:hover,.upload-zone.drag{
  border-color:var(--acc);background:rgba(99,102,241,.06);
  box-shadow:0 0 30px rgba(99,102,241,.1);
}
.upload-zone .icon{font-size:48px;margin-bottom:12px;display:block}
.upload-zone .title{font-size:16px;font-weight:600;margin-bottom:6px}
.upload-zone .desc{color:var(--mut);font-size:13px}
.upload-zone input{position:absolute;inset:0;opacity:0;cursor:pointer}

/* File Info */
.file-info{
  display:none;align-items:center;gap:14px;padding:16px 20px;
  background:var(--card);border:1px solid var(--line);border-radius:12px;
  margin-bottom:20px;
}
.file-info.show{display:flex}
.file-info .fname{font-weight:600;font-size:14px;flex:1}
.file-info .fsize{color:var(--mut);font-size:12px}
.file-info .remove{background:none;border:none;color:var(--low);cursor:pointer;
  font-size:18px;padding:4px 8px;border-radius:6px}
.file-info .remove:hover{background:rgba(239,68,68,.15)}

/* Buttons */
.actions{display:flex;gap:12px;align-items:center;margin-bottom:24px;flex-wrap:wrap}
.btn{border:none;border-radius:10px;padding:12px 24px;font-size:14px;
  font-weight:600;cursor:pointer;transition:all .2s;font-family:inherit;
  display:inline-flex;align-items:center;gap:8px}
.btn-primary{background:var(--grad);color:#fff;box-shadow:0 4px 14px rgba(99,102,241,.3)}
.btn-primary:hover{transform:translateY(-1px);box-shadow:0 6px 20px rgba(99,102,241,.4)}
.btn-primary:disabled{opacity:.5;cursor:default;transform:none}
.btn-secondary{background:var(--card2);color:var(--acc2);border:1px solid var(--line)}
.btn-secondary:hover{background:rgba(99,102,241,.1);border-color:var(--acc)}
.btn-download{background:linear-gradient(135deg,#059669,#10b981);color:#fff;
  box-shadow:0 4px 14px rgba(16,185,129,.3)}
.btn-download:hover{transform:translateY(-1px)}

/* Stats Cards */
.stats{display:none;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:24px}
.stats.show{display:grid}
.stat-card{background:var(--card);border:1px solid var(--line);border-radius:12px;
  padding:16px 18px;text-align:center}
.stat-card .val{font-size:28px;font-weight:700;margin-bottom:2px}
.stat-card .lbl{font-size:12px;color:var(--mut);font-weight:500}
.stat-card.s1 .val{color:var(--acc2)}
.stat-card.s2 .val{color:var(--ok)}
.stat-card.s3 .val{color:#38bdf8}
.stat-card.s4 .val{color:var(--mid)}

/* Progress */
.progress{display:none;margin-bottom:20px}
.progress.show{display:block}
.progress-bar{height:6px;background:var(--card2);border-radius:3px;overflow:hidden}
.progress-fill{height:100%;background:var(--grad);border-radius:3px;
  transition:width .3s;width:0%}
.progress-text{color:var(--mut);font-size:12px;margin-top:6px;text-align:center}

/* Table */
.table-wrap{overflow-x:auto;border-radius:12px;border:1px solid var(--line);
  background:var(--card);display:none}
.table-wrap.show{display:block}
table{width:100%;border-collapse:collapse;font-size:12.5px}
th{background:#0d1120;color:var(--mut);font-weight:600;padding:10px 12px;
  text-align:left;position:sticky;top:0;z-index:1;white-space:nowrap;
  border-bottom:1px solid var(--line)}
td{padding:8px 12px;border-bottom:1px solid rgba(37,42,58,.5);vertical-align:top}
tr:hover td{background:rgba(99,102,241,.04)}
td.raw{color:var(--mut);max-width:260px;overflow:hidden;text-overflow:ellipsis}
td.street{color:#bfdbfe;font-weight:500}
td.l4{color:#a5f3fc;font-weight:500}
td.poi{color:#fbbf24}
td.full{font-weight:600;color:var(--txt)}
.idx{color:var(--mut);font-size:11px;text-align:center;width:40px}

/* Error */
.errbar{display:none;background:rgba(239,68,68,.1);border:1px solid rgba(239,68,68,.3);
  color:#fca5a5;border-radius:10px;padding:12px 16px;font-size:13px;margin-bottom:16px}
.errbar.show{display:block}

/* Responsive */
@media(max-width:768px){
  .stats{grid-template-columns:repeat(2,1fr)}
  .wrap{padding:16px}
}
</style></head><body>
<div class="wrap">
  <div class="header">
    <div>
      <h1>🧹 Làm sạch địa chỉ Việt Nam</h1>
      <div class="sub">Upload file Excel → Xử lý batch → Tải kết quả sạch</div>
    </div>
  </div>

  <div class="upload-zone" id="dropzone">
    <span class="icon">📁</span>
    <div class="title">Kéo thả file Excel vào đây</div>
    <div class="desc">hoặc click để chọn file (.xlsx)</div>
    <input type="file" id="fileInput" accept=".xlsx,.xls">
  </div>

  <div class="file-info" id="fileInfo">
    <span>📄</span>
    <span class="fname" id="fileName"></span>
    <span class="fsize" id="fileSize"></span>
    <button class="remove" id="removeFile" title="Xoá file">✕</button>
  </div>

  <div class="actions">
    <button class="btn btn-primary" id="goClean" disabled>
      <span>⚡</span> Làm sạch
    </button>
    <button class="btn btn-download" id="download" style="display:none">
      <span>📥</span> Tải Excel đã làm sạch
    </button>
  </div>

  <div class="progress" id="progress">
    <div class="progress-bar"><div class="progress-fill" id="progressFill"></div></div>
    <div class="progress-text" id="progressText">Đang xử lý...</div>
  </div>

  <div class="errbar" id="errbar"></div>

  <div class="stats" id="stats">
    <div class="stat-card s1"><div class="val" id="stTotal">0</div><div class="lbl">Dòng giữ lại</div></div>
    <div class="stat-card s2"><div class="val" id="stFull">0</div><div class="lbl">Đủ 3 cấp HC</div></div>
    <div class="stat-card s3"><div class="val" id="stDetail">0</div><div class="lbl">Có POI/Đường/Cấp 4</div></div>
    <div class="stat-card s4"><div class="val" id="stRemoved">0</div><div class="lbl">Đã loại</div></div>
  </div>

  <div class="table-wrap" id="tableWrap">
    <table>
      <thead><tr>
        <th class="idx">#</th><th>Địa chỉ gốc</th><th>POI</th>
        <th>Tên đường</th><th>Cấp 4</th><th>Phường/Xã</th><th>Quận/Huyện</th>
        <th>Tỉnh/TP</th><th>ĐỊA CHỈ SẠCH</th>
      </tr></thead>
      <tbody id="tbody"></tbody>
    </table>
  </div>
</div>

<script>
const $=s=>document.querySelector(s);
let selectedFile=null;

// Drag & drop
const dz=$('#dropzone');
['dragenter','dragover'].forEach(e=>dz.addEventListener(e,ev=>{ev.preventDefault();dz.classList.add('drag')}));
['dragleave','drop'].forEach(e=>dz.addEventListener(e,ev=>{ev.preventDefault();dz.classList.remove('drag')}));
dz.addEventListener('drop',ev=>{
  const f=ev.dataTransfer.files[0];
  if(f&&(f.name.endsWith('.xlsx')||f.name.endsWith('.xls')))setFile(f);
});
$('#fileInput').addEventListener('change',ev=>{
  if(ev.target.files[0])setFile(ev.target.files[0]);
});

function setFile(f){
  selectedFile=f;
  $('#fileName').textContent=f.name;
  $('#fileSize').textContent=(f.size/1024).toFixed(1)+' KB';
  $('#fileInfo').classList.add('show');
  $('#dropzone').style.display='none';
  $('#goClean').disabled=false;
  $('#download').style.display='none';
  $('#stats').classList.remove('show');
  $('#tableWrap').classList.remove('show');
  $('#errbar').classList.remove('show');
}

$('#removeFile').onclick=()=>{
  selectedFile=null;
  $('#fileInfo').classList.remove('show');
  $('#dropzone').style.display='';
  $('#goClean').disabled=true;
  $('#fileInput').value='';
  $('#download').style.display='none';
  $('#stats').classList.remove('show');
  $('#tableWrap').classList.remove('show');
};

function esc(s){return(s||'').replace(/[&<>]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;'}[c]))}

$('#goClean').onclick=async()=>{
  if(!selectedFile)return;
  const btn=$('#goClean');
  btn.disabled=true;btn.innerHTML='<span>⏳</span> Đang xử lý...';
  $('#errbar').classList.remove('show');
  $('#progress').classList.add('show');
  $('#progressFill').style.width='30%';
  $('#progressText').textContent='Đang upload và xử lý...';

  try{
    const fd=new FormData();
    fd.append('file',selectedFile);
    $('#progressFill').style.width='50%';

    const r=await fetch('/api/upload',{method:'POST',body:fd});
    $('#progressFill').style.width='90%';

    if(!r.ok){
      const err=await r.text();
      throw new Error(err);
    }
    const data=await r.json();
    $('#progressFill').style.width='100%';

    if(data.error){
      $('#errbar').textContent='⚠️ '+data.error;
      $('#errbar').classList.add('show');
    }

    // Stats
    const st=data.stats||{};
    const n=st.n||0;
    $('#stTotal').textContent=n;
    $('#stFull').textContent=(st.full||0)+' ('+((st.full||0)*100/Math.max(n,1)).toFixed(1)+'%)';
    $('#stDetail').textContent=(st.detail||0)+' ('+((st.detail||0)*100/Math.max(n,1)).toFixed(1)+'%)';
    $('#stRemoved').textContent=(st.removed||0)+' / '+(st.input_n||n);
    $('#stats').classList.add('show');

    // Table
    const rows=data.rows||[];
    let h='';
    rows.forEach(r=>{
      h+=`<tr><td class="idx">${r.idx}</td><td class="raw">${esc(r.raw)}</td>`
        +`<td class="poi">${esc(r.poi)}</td><td class="street">${esc(r.street)}</td><td class="l4">${esc(r.l4)}</td>`
        +`<td>${esc(r.ward)}</td><td>${esc(r.district)}</td><td>${esc(r.province)}</td>`
        +`<td class="full">${esc(r.full)}</td></tr>`;
    });
    $('#tbody').innerHTML=h;
    $('#tableWrap').classList.add('show');
    $('#download').style.display='';

  }catch(e){
    $('#errbar').textContent='⚠️ Lỗi: '+e.message;
    $('#errbar').classList.add('show');
  }finally{
    btn.disabled=false;btn.innerHTML='<span>⚡</span> Làm sạch';
    setTimeout(()=>$('#progress').classList.remove('show'),500);
  }
};

$('#download').onclick=()=>{
  window.location='/api/download';
};
</script></body></html>"""


# ─────────────────────────── HTTP Handler ───────────────────────────

class H(BaseHTTPRequestHandler):
    def _send(self, code, body, ctype="text/html; charset=utf-8"):
        b = body.encode("utf-8") if isinstance(body, str) else body
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(b)))
        self.end_headers()
        self.wfile.write(b)

    def do_GET(self):
        if self.path in ("/", "/index.html"):
            self._send(200, PAGE)
        elif self.path == "/api/download":
            if _last_result["wb_bytes"]:
                fname = _last_result["filename"]
                b = _last_result["wb_bytes"]
                self.send_response(200)
                self.send_header("Content-Type",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition",
                    f'attachment; filename="{fname}"')
                self.send_header("Content-Length", str(len(b)))
                self.end_headers()
                self.wfile.write(b)
            else:
                self._send(404, "No file to download", "text/plain")
        else:
            self._send(404, "not found", "text/plain")

    def do_POST(self):
        if self.path != "/api/upload":
            self._send(404, "not found", "text/plain")
            return

        try:
            content_type = self.headers.get("Content-Type", "")
            if "multipart/form-data" not in content_type:
                self._send(400, "Expected multipart/form-data", "text/plain")
                return

            # Parse multipart
            boundary = content_type.split("boundary=")[1].strip()
            content_length = int(self.headers.get("Content-Length", 0))
            body = self.rfile.read(content_length)

            # Find file data between boundaries
            boundary_bytes = ("--" + boundary).encode()
            parts = body.split(boundary_bytes)

            file_bytes = None
            orig_filename = "uploaded.xlsx"

            for part in parts:
                if b"filename=" in part:
                    # Extract filename
                    header_end = part.find(b"\r\n\r\n")
                    if header_end == -1:
                        continue
                    header_str = part[:header_end].decode("utf-8", "replace")
                    fn_match = re.search(r'filename="([^"]+)"', header_str)
                    if fn_match:
                        orig_filename = fn_match.group(1)

                    # Extract file content
                    file_bytes = part[header_end + 4:]
                    # Remove trailing \r\n-- if present
                    if file_bytes.endswith(b"\r\n"):
                        file_bytes = file_bytes[:-2]
                    break

            if not file_bytes:
                self._send(400, "No file found in upload", "text/plain")
                return

            # Process
            t0 = time.time()
            ui_rows, stats, out_bytes = process_excel(file_bytes, orig_filename)
            elapsed = time.time() - t0

            # Store for download
            base = os.path.splitext(orig_filename)[0]
            _last_result["wb_bytes"] = out_bytes
            _last_result["filename"] = f"{base}_cleaned.xlsx"

            n = stats.get("n", 0)
            stats["new_pct"] = (stats.get("new", 0) * 100 / max(n, 1))
            stats["elapsed"] = round(elapsed, 2)

            resp = json.dumps({
                "rows": ui_rows,
                "stats": stats,
                "error": None,
            }, ensure_ascii=False)

            self._send(200, resp, "application/json; charset=utf-8")

        except Exception as e:
            traceback.print_exc()
            self._send(500,
                json.dumps({"error": str(e)}, ensure_ascii=False),
                "application/json; charset=utf-8")

    def log_message(self, *a):
        pass


def main():
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 8899
    srv = ThreadingHTTPServer(("127.0.0.1", port), H)
    print(f"\n  UI làm sạch địa chỉ đang chạy:")
    print(f"     ->  http://localhost:{port}\n")
    print("  (Ctrl+C để dừng)\n")
    try:
        srv.serve_forever()
    except KeyboardInterrupt:
        print("\nĐã dừng.")


if __name__ == "__main__":
    main()
