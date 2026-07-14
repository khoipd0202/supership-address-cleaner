import io
import unittest

import openpyxl

from address_ui import _combine_output_workbooks
from vn_address_cleaner.cleaner import admin_sort_key, sort_worksheet_by_admin


class OutputSortingTests(unittest.TestCase):
    def test_admin_sort_key_groups_province_district_ward_naturally(self):
        rows = [
            ("Thành phố Hồ Chí Minh", "Quận 1", "Phường 10"),
            ("Thành phố Hà Nội", "Quận Nam Từ Liêm", "Phường Mễ Trì"),
            ("Thành phố Hà Nội", "Huyện Đông Anh", "Thị trấn Đông Anh"),
            ("Thành phố Hồ Chí Minh", "Quận 1", "Phường 2"),
            ("", "", ""),
        ]

        rows.sort(key=lambda row: admin_sort_key(row[0], row[1], row[2]))

        self.assertEqual(rows, [
            ("Thành phố Hà Nội", "Huyện Đông Anh", "Thị trấn Đông Anh"),
            ("Thành phố Hà Nội", "Quận Nam Từ Liêm", "Phường Mễ Trì"),
            ("Thành phố Hồ Chí Minh", "Quận 1", "Phường 2"),
            ("Thành phố Hồ Chí Minh", "Quận 1", "Phường 10"),
            ("", "", ""),
        ])

    def test_combined_download_is_sorted_across_source_files(self):
        headers = [
            "STT", "Địa chỉ gốc", "Điểm định vị (POI)", "Tên đường", "Cấp 4",
            "Phường/Xã", "Quận/Huyện", "Tỉnh/TP", "ĐỊA CHỈ SẠCH",
        ]

        def workbook_bytes(row):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Địa chỉ sạch"
            sheet.append(headers)
            sheet.append(row)
            buffer = io.BytesIO()
            workbook.save(buffer)
            return buffer.getvalue()

        hanoi = [1, "HN", "", "", "", "Phường Mễ Trì", "Quận Nam Từ Liêm", "Thành phố Hà Nội", "HN"]
        an_giang = [1, "AG", "", "", "", "Xã Bình Mỹ", "Huyện Châu Phú", "Tỉnh An Giang", "AG"]
        combined = _combine_output_workbooks([
            ("hanoi.xlsx", workbook_bytes(hanoi)),
            ("an_giang.xlsx", workbook_bytes(an_giang)),
        ])

        result = openpyxl.load_workbook(io.BytesIO(combined), data_only=True)
        sheet = result["Địa chỉ sạch"]
        self.assertEqual(sheet.cell(2, 9).value, "Tỉnh An Giang")
        self.assertEqual(sheet.cell(3, 9).value, "Thành phố Hà Nội")

    def test_sort_worksheet_leaves_missing_admin_at_end(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["POI", "Phường/Xã", "Quận/Huyện", "Tỉnh/TP"])
        sheet.append(["Thiếu", "", "", ""])
        sheet.append(["Có", "Xã Bình Mỹ", "Huyện Châu Phú", "Tỉnh An Giang"])

        self.assertTrue(sort_worksheet_by_admin(sheet))
        self.assertEqual(sheet.cell(2, 1).value, "Có")
        self.assertEqual(sheet.cell(3, 1).value, "Thiếu")


if __name__ == "__main__":
    unittest.main()
