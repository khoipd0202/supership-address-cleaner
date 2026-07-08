# -*- coding: utf-8 -*-
"""
clean_address.py — Ghép ĐỊA CHỈ CẤP 4 SẠCH NHẤT.

Nguyên tắc:
  * Đơn vị hành chính (Xã/Phường, Huyện/Quận, Tỉnh) LẤY TỪ 3 CỘT SẠCH có sẵn
    -> chuẩn hoá tên + map hệ 34 tỉnh mới bằng vn_units_data.json (chính xác ~100%).
  * Chi tiết cấp 4 (số nhà / đường / thôn / xóm / ấp / khu phố / tổ) BÓC TỪ ô địa chỉ
    gốc bằng cách TRỪ ĐI: tên người, nhãn ("Khách hàng:", "Địa chỉ:"...), SĐT, và
    chính các tên đơn vị hành chính đã biết -> phần còn lại là chi tiết, rồi chuẩn hoá
    hoa/thường + dấu câu.
  * Ghép: <chi tiết>, <Xã>, <Huyện>, <Tỉnh>  -> 1 cột địa chỉ sạch hoàn chỉnh.
"""
import os
import re
import sys
import unicodedata
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from parse_address import (  # noqa: E402
    Parser, lookup_labels, extract_phone, strip_diacritics, clean_text,
    levenshtein,
)

DATA = os.path.join(HERE, "vn_units_data.json")

# ---------- chuẩn hoá chữ ----------

def _norm(s):
    s = strip_diacritics(str(s or "").lower())
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return " " + s.strip() + " "

# từ chỉ loại đơn vị — CHỈ dùng để nuốt phần tiền tố ĐỨNG NGAY TRƯỚC một tên
# đơn vị đã khớp (vd "thành phố" trong "thành phố Hà Nội"), KHÔNG xoá đứng lẻ
# để tránh phá tên thật ("Lê Thị", "Long Thành").
# tiền tố AN TOÀN để nuốt đứng lẻ (hầu như không bao giờ là tên riêng)
SAFE_PREFIX = {"tinh", "tp", "tphcm", "xa", "quan", "q", "huyen", "phuong", "p",
               "tt", "tx"}
# tiền tố loại đơn vị 2 từ — chỉ nuốt khi khớp trọn cụm (tránh phá "Long Thành")
PREFIX_PAIRS = [("thanh", "pho"), ("thi", "xa"), ("thi", "tran")]


def _base(nm):
    n = _norm(nm).strip()
    n = re.sub(r"^(thanh pho|thi xa|thi tran|tinh|tp|quan|huyen|phuong|xa|q|p|h|x)\s+",
               "", n)
    return n.strip()
# viết tắt tỉnh/thành hay gặp trong text bẩn -> để xoá khỏi phần chi tiết
PROV_ABBR = {
    "tphcm": "ho chi minh", "hcm": "ho chi minh", "tp hcm": "ho chi minh",
    "hn": "ha noi", "hp": "hai phong", "dn": "da nang", "dnai": "dong nai",
    "bd": "binh duong", "bp": "binh phuoc", "brvt": "ba ria vung tau",
    "vt": "ba ria vung tau",
    "qn": "quang ninh", "daklak": "dak lak", "dak lak": "dak lak",
    "daclak": "dak lak", "dac lac": "dak lak",
}

LABELS_RE = re.compile(
    r"(tên\s*kh\b|tên\s*khách\s*hàng|khách\s*hàng|kh\b|người\s*nhận|ngày\s*nhận|nhận\s*hàng|"
    r"số\s*(điện\s*thoại|đt)|s[đd]t|đt\b|điện\s*thoại|"
    r"địa\s*chỉ|đ/c|đc\b|d/c|dc\b|ghi\s*chú|note)\s*[:：\-]*",
    re.IGNORECASE,
)


def _titlecase_vn(s):
    def fix(w):
        if not w:
            return w
        if w.isdigit():
            return w
        # giữ nguyên token có số + chữ (12a, 200/5, kp3) nhưng viết hoa chữ cái đầu cụm chữ
        if any(ch.isdigit() for ch in w):
            return w.lower()
        return w[0].upper() + w[1:].lower()
    parts = re.split(r"(\s+|/)", s)
    return "".join(fix(p) if p.strip() and p != "/" else p for p in parts)


def remove_order_garbage(s):
    if not s:
        return ""
        
    s = unicodedata.normalize("NFC", str(s))
    
    # 1. Clean emojis and special symbols, keeping alphanumeric, spaces, commas, periods, hyphens, slashes, colons
    s = re.sub(r'[^\w\s,.\-\(\)\/：:]', ' ', s)
    
    # 2. Split by comma/newline and clean segment by segment
    segments = re.split(r'[,;\n]+', s)
    cleaned_segments = []
    
    address_kws = {
        "đường", "phố", "ngõ", "ngách", "hẻm", "kiệt", "thôn", "xóm", "ấp", "tổ", "khu", "bản",
        "phường", "quận", "huyện", "tỉnh", "thành", "ql", "tl", "hl", "dt", "km", "lộ", "sn", "khóm",
        "duong", "pho", "ngo", "ngach", "hem", "kiet", "ap", "thon", "xom", "to", "khu", "kp", "khom",
        "quan", "huyen", "tinh", "thanh", "xa", "ban"
    }
    
    garbage_keywords = {
        "ship", "giao", "gọi", "goi", "hàng", "hang", "cod", "tiền", "tien", "cước", "cuoc",
        "lưu ý", "luu y", "size", "sz", "màu", "mau", "quần", "áo", "giày", "dép", "váy", "đầm",
        "kiện", "cái", "chiếc", "bộ", "kg", "cước", "cửa hàng", "shop", "kho", "xuong", "xưởng",
        "tên hàng", "đồng kiểm", "cho xem", "hàng dễ vỡ", "giao hành chính", "hành chính", "hc", "giao gấp", "alo", "khách", "khach",
        "trả tiền", "trả ship", "nhận nhé", "gửi chị", "giao chiều", "giao tối", "hành chánh", "trừ thứ"
    }

    poi_kws = {
        "truong", "mn", "mam non", "tieu hoc", "trung hoc", "thpt", "thcs", "cao dang",
        "dai hoc", "cty", "cong ty", "tnhh", "chua", "benh vien", "toa", "toa",
        "bv", "phong kham", "pk", "nha thuoc", "quay thuoc",
        "nha sach", "cua hang", "shop", "kho",
        "xuong", "ben xe", "cang", "ubnd", "uy ban",
        "ngan hang", "toa nha", "building", "chung cu",
        "khach san", "ks", "nha nghi", "sieu thi",
        "cho", "trung tam", "bhxh",
        "bao hiem", "nha van hoa", "nha ga",
        "nha tho", "nha tro", "nha khach",
        "gara", "garage", "dai ly", "co so", "hkd",
        "ho kinh doanh", "buu dien", "mieu",
        "dinh", "nghia trang", "ttyt", "tram y te"
    }
    
    shipping_prefixes = [
        r"^\s*(?:đc|đ/c|dc|địa\s*chỉ\s*giao\s*hàng|địa\s*chỉ\s*nhận\s*hàng|địa\s*chỉ|ship\s*mình\s*đến\s*địa\s*chỉ|ship\s*mình\s*đến|ship\s*đến|ship\s*về|giao\s*về|giao\s*đến\s*địa\s*chỉ|giao\s*đến|gởi\s*về|gửi\s*về|sđt|sdt|người\s*nhận|ng\s*nhận|ng\s*nhan|tên\s*người\s*nhận|khách\s*hàng|em\s*gởi\s*về\s*địa\s*chỉ\s*người\s*nhận|em\s*gửi\s*về\s*địa\s*chỉ\s*người\s*nhận|em\s*gởi\s*về|em\s*gửi\s*về|cho\s*mình\s*về|gửi\s*chị\s*đi\s*nhé\s*về|gửi\s*chị\s*đi\s*nhé|đổi\s*địa\s*chỉ\s*về)\s*[:：\-]*\s*"
    ]
    shipping_prefix_re = re.compile("|".join(shipping_prefixes), re.IGNORECASE)
    
    from parse_address import strip_diacritics
    
    # Strip diacritics from all address keywords to ensure perfect matching
    address_kws = {strip_diacritics(kw.lower()) for kw in address_kws}
    
    for seg in segments:
        seg_strip = seg.strip()
        if not seg_strip:
            continue
            
        # Strip shipping prefixes inside the segment first
        seg_strip = shipping_prefix_re.sub("", seg_strip).strip()
        if not seg_strip:
            continue
            
        # Check strong garbage phrases first
        strong_garbage_kws = {
            "ko dung hang", "tra lai", "mien ship", "cho xem hang", "dong kiem", 
            "khong dung hang", "ko dung hang nhu hinh", "alo sdt", "ship gio hanh chinh", 
            "khong cho xem hang", "cho xem", "khong dung", "ko dung", "bom hang", 
            "boom hang", "hoan hang", "tra hang", "tra lai nhe", "tra lai em",
            "hang tra lai", "ship nhe", "goi truoc", "goi sdt", "goi dien",
            "khong nhan", "ko nhan", "huy don", "ko lay", "khong lay"
        }
        strong_address_kws = {
            "duong", "pho", "ngo", "ngach", "hem", "kiet", "ap", "thon", "xom", "to", "khu", 
            "sn", "so nha"
        }
        
        seg_norm = strip_diacritics(seg_strip.lower())
        words = seg_norm.split()
        
        has_strong_garbage = False
        for skw in strong_garbage_kws:
            if skw in seg_norm:
                has_strong_garbage = True
                break
                
        has_strong_address = any(skw in words or skw in seg_norm for skw in strong_address_kws)
        if has_strong_garbage and not has_strong_address:
            continue

        # Check if the segment contains any garbage keywords safely
        has_garbage = False
        for kw in garbage_keywords:
            kw_norm = strip_diacritics(kw.lower())
            if " " in kw_norm:
                if re.search(r"\b" + re.escape(kw_norm) + r"\b", seg_norm):
                    has_garbage = True
                    break
            else:
                if kw_norm in words:
                    has_garbage = True
                    break
                    
        has_poi = False
        for pkw in poi_kws:
            pkw_norm = strip_diacritics(pkw.lower())
            if " " in pkw_norm:
                if re.search(r"\b" + re.escape(pkw_norm) + r"\b", seg_norm):
                    has_poi = True
                    break
            else:
                if pkw_norm in words:
                    has_poi = True
                    break
                    
        # Check if the segment contains any administrative name from the database
        has_admin_name = False
        all_admin = get_all_admin_names()
        common_single_words = {
            "lai", "chi", "em", "hang", "nhe", "dung", "cho", "an", "loc", "son",
            "hoa", "hai", "dong", "nam", "bac", "trung", "tay", "binh", "minh",
            "yen", "thanh", "long", "duc", "viet", "nam", "quang", "phu", "khanh",
            "gia", "lam", "hung", "phong", "son", "tan", "hiep", "phuoc",
            "my", "cat", "tien", "thuan", "nghia", "tri", "dao", "huong", "bac",
            "thu", "tu", "nam", "sau", "bay", "chu", "nhut"
        }
        for length in (1, 2, 3):
            for i in range(len(words) - length + 1):
                phrase = " ".join(words[i:i+length])
                if phrase in all_admin:
                    if length == 1 and phrase in common_single_words:
                        continue
                    has_admin_name = True
                    break
            if has_admin_name:
                break
                
        has_address = has_poi or has_admin_name or any(kw in words or kw in seg_norm for kw in address_kws)
        
        # If segment has garbage keywords and no address/POI keywords, we discard it!
        if has_garbage and not has_address:
            continue
            
        # Also discard segments that are pure delivery notes like "trừ thứ 4" or "t2 đến t6"
        if re.search(r"\bt[2-7]\b|\bchủ\s*nhật\b|\bcn\b", seg_norm) and not has_address:
            continue
            
        cleaned_segments.append(seg_strip)
        
    s_cleaned = ", ".join(cleaned_segments)
    
    # Also clean item count patterns like "3 cái", "1 bộ" from the final string
    s_cleaned = re.sub(r"\b\d+\s*(?:kiện|cái|chiếc|bộ|kg|g|k|đ|d|vnd|vnd|vnđ)\b", " ", s_cleaned, flags=re.I)
    return s_cleaned


_PARSER = None
_ALL_ADMIN_NAMES = None

def get_all_admin_names():
    global _PARSER, _ALL_ADMIN_NAMES
    if _ALL_ADMIN_NAMES is not None:
        return _ALL_ADMIN_NAMES
    if _PARSER is None:
        _PARSER = Parser(DATA)
    names = set()
    
    def add_name_and_variants(u_name, u_strong_aliases):
        for a in u_strong_aliases:
            names.add(a)
            if " " in a:
                names.add(a.replace(" ", ""))
        if "'" in u_name:
            clean_name = strip_diacritics(u_name.replace("'", "").lower())
            clean_name = re.sub(r"[^a-z0-9]+", " ", clean_name).strip()
            names.add(clean_name)
            if " " in clean_name:
                names.add(clean_name.replace(" ", ""))
            base = re.sub(r"^(tinh|tp|thanh pho|quan|huyen|phuong|xa|tt|tx)\s+", "", clean_name)
            names.add(base)
            if " " in base:
                names.add(base.replace(" ", ""))

    for u in _PARSER.provinces.values():
        add_name_and_variants(u.name, u.strong + u.aliases)
    for u in _PARSER.districts.values():
        add_name_and_variants(u.name, u.strong + u.aliases)
    for u in _PARSER.wards.values():
        add_name_and_variants(u.name, u.strong + u.aliases)
    for p_name in _PARSER.prov_mig.values():
        names.add(strip_diacritics(p_name.lower()))
    for p_id, w_list in _PARSER.new_wards_by_prov.items():
        for u in w_list:
            add_name_and_variants(u.name, u.strong + u.aliases)
            
    _ALL_ADMIN_NAMES = names
    return _ALL_ADMIN_NAMES




def clean_detail(raw, admin_names):
    """Bóc chi tiết cấp 4 từ ô địa chỉ gốc."""
    s = clean_text(raw) or ""
    s = remove_order_garbage(s)
    # CHUẨN HOÁ NFC NGAY TỪ ĐẦU: nhiều ô Excel lưu Unicode tổ hợp (NFD) khiến
    # dấu thanh tách rời ký tự -> bộ tách token cắt "Bình"->"Bi"+"nh", làm tên
    # đơn vị không khớp được với 3 cột -> lồng vào chi tiết. NFC gộp lại đúng.
    s = unicodedata.normalize("NFC", s)
    s = s.replace("\r", " ")
    # bỏ SĐT + mọi chuỗi số dài (>=9) coi như số điện thoại/đơn
    ph = extract_phone(s)
    if ph:
        s = s.replace(ph, " ")
    s = re.sub(r"\+?\d[\d\.\-\s]{8,}\d", " ", s)
    # bỏ nhãn
    s = LABELS_RE.sub(" ", s)
    # ------- xoá tên đơn vị hành chính đã biết (không phân biệt dấu) -------
    # build danh sách cụm normalized cần xoá: base tên (đã bỏ tiền tố) + full
    kill = set()
    for nm in admin_names:
        if not nm:
            continue
        b = _base(nm)
        if b:
            kill.add(b)
            # thêm TIỀN TỐ CỤT (>=2 từ) để cắt tên bị ghi thiếu đuôi
            # vd "Tăng Nhơn" (cụt) của "Tăng Nhơn Phú A"
            bw = b.split()
            for L in range(2, len(bw)):
                kill.add(" ".join(bw[:L]))
        full = _norm(nm).strip()
        if full:
            kill.add(full)
    for ab, full in PROV_ABBR.items():
        if full in kill:
            kill.add(ab)
    # token hoá giữ vị trí
    toks = [(m.group(), m.start(), m.end()) for m in re.finditer(r"[^\W_]+", s, re.UNICODE)]
    norms = [strip_diacritics(w.lower()) for w, _, _ in toks]
    remove = [False] * len(toks)
    # đánh dấu các cụm token khớp tên đơn vị (dài trước, ngắn sau); mỗi lần khớp
    # còn "nuốt" thêm các token loại-đơn-vị đứng NGAY TRƯỚC (vd "thành phố").
    # 1. Thu thập tất cả các match ứng cử viên cho khớp chính xác
    candidate_matches = []
    kills_sorted = sorted(kill, key=lambda x: -len(x.split()))
    for kph in kills_sorted:
        kw = kph.split()
        L = len(kw)
        if L == 0:
            continue
        for i in range(len(norms) - L + 1):
            if norms[i:i + L] == kw:
                candidate_matches.append((i, i + L, kph))

    # Chọn các match không chồng lấn, ưu tiên dài trước
    candidate_matches.sort(key=lambda x: -(x[1] - x[0]))
    final_matches = []
    covered = set()
    for start, end, kph in candidate_matches:
        if not any(idx in covered for idx in range(start, end)):
            final_matches.append((start, end, kph))
            for idx in range(start, end):
                covered.add(idx)

    # Đánh giá bảo vệ và đánh dấu xoá cho các match được chọn
    guards = {
        "duong", "pho", "ngo", "ngach", "hem", "kiet", "ap", "thon",
        "xom", "to", "khu", "kp", "tdp", "doi", "ban", "buon", "lang",
        "truong", "mn", "ubnd", "cho", "ngan", "benh", "bv", "uy", "chua",
        "nha", "khach", "shop", "kho", "xuong"
    }
    phrase_guards = {
        "mam non", "tieu hoc", "trung hoc", "benh vien", "ngan hang",
        "uy ban", "khu pho", "to dan", "tieu khu"
    }

    for start, end, kph in final_matches:
        is_guarded = False
        if start > 0:
            prev_tok = norms[start - 1]
            if prev_tok in guards and (start - 1) not in covered:
                is_guarded = True
            if start > 1:
                prev_phrase = norms[start - 2] + " " + norms[start - 1]
                if prev_phrase in phrase_guards and (start - 2) not in covered and (start - 1) not in covered:
                    is_guarded = True
        # Nếu tên tỉnh/huyện/xã là một phần của tên POI trong cùng cụm
        # ("Bệnh viện ... Đồng Nai", "Trường Mầm Non Đức Bác"), đừng xoá
        # trừ khi bản thân match có tiền tố hành chính rõ ràng ("Tỉnh ...").
        has_explicit_admin_prefix = (
            norms[start] in SAFE_PREFIX
            or (start + 1 < len(norms) and (norms[start], norms[start + 1]) in PREFIX_PAIRS)
        )
        if has_explicit_admin_prefix:
            is_guarded = False
        if not is_guarded and not has_explicit_admin_prefix:
            seg_start = start
            while seg_start > 0:
                sep = s[toks[seg_start - 1][2]:toks[seg_start][1]]
                if re.search(r"[,;\n]|[-–]", sep):
                    break
                seg_start -= 1
            left_raw = s[toks[seg_start][1]:toks[start][1]] if seg_start < start else ""
            if LEADING_HOUSE_RE.match(left_raw.strip()):
                is_guarded = True
            left = norms[seg_start:start]
            poi_left = {
                "truong", "mam", "non", "thcs", "thpt", "benh", "vien", "bv",
                "ngan", "hang", "ubnd", "uy", "ban", "so", "toa", "cty",
                "cong", "ty", "phong", "kham", "nha", "thuoc", "ben", "xe",
                "sieu", "thi", "cho", "chung", "cu"
            }
            if any(x in poi_left for x in left):
                is_guarded = True
        
        if not is_guarded:
            for j in range(start, end):
                remove[j] = True
            k = start - 1
            # nuốt cặp tiền tố 2 từ (thành phố / thị xã / thị trấn)
            if k >= 1 and (norms[k - 1], norms[k]) in PREFIX_PAIRS:
                sep1 = s[toks[k][2] : toks[k+1][1]]
                sep2 = s[toks[k-1][2] : toks[k][1]]
                if not re.search(r"[,;\-\n/]", sep1) and not re.search(r"[,;\-\n/]", sep2):
                    remove[k] = remove[k - 1] = True
                    k -= 2
            # nuốt các tiền tố an toàn đứng lẻ
            while k >= 0 and norms[k] in SAFE_PREFIX:
                sep = s[toks[k][2] : toks[k+1][1]]
                if re.search(r"[,;\-\n/]", sep):
                    break
                remove[k] = True
                k -= 1
    # vòng khớp GẦN ĐÚNG (sai 1 ký tự) cho tên >=6 ký tự -> bắt bản viết sai
    # chính tả như "Kiêng Giang" (đúng "Kiên Giang"). Ngưỡng chặt (=1) để an toàn.
    for kph in kills_sorted:
        kw = kph.split()
        L = len(kw)
        kjoin = " ".join(kw)
        if L == 0 or len(kjoin.replace(" ", "")) < 6:
            continue
        for i in range(len(norms) - L + 1):
            if any(remove[i:i + L]):
                continue
            wjoin = " ".join(norms[i:i + L])
            if wjoin == kjoin:
                continue
            if levenshtein(wjoin, kjoin, 1) <= 1:
                # Kiểm tra bảo vệ các địa danh/đường đi kèm: Đường Lê Lợi, Trường Đức Bác, Chợ Ba Đồn...
                is_guarded = False
                if i > 0 and not remove[i - 1]:
                    prev_tok = norms[i - 1]
                    guards = {
                        "duong", "pho", "ngo", "ngach", "hem", "kiet", "ap", "thon",
                        "xom", "to", "khu", "kp", "tdp", "doi", "ban", "buon", "lang",
                        "truong", "mn", "ubnd", "cho", "ngan", "benh", "bv", "uy", "chua",
                        "nha", "khach", "shop", "kho", "xuong"
                    }
                    if prev_tok in guards:
                        is_guarded = True
                    if i > 1 and not remove[i - 2]:
                        prev_phrase = norms[i - 2] + " " + norms[i - 1]
                        phrase_guards = {
                            "mam non", "tieu hoc", "trung hoc", "benh vien", "ngan hang",
                            "uy ban", "khu pho", "to dan", "tieu khu"
                        }
                        if prev_phrase in phrase_guards:
                            is_guarded = True
                
                if not is_guarded:
                    for j in range(i, i + L):
                        remove[j] = True
                    k = i - 1
                    if k >= 1 and (norms[k - 1], norms[k]) in PREFIX_PAIRS:
                        sep1 = s[toks[k][2] : toks[k+1][1]]
                        sep2 = s[toks[k-1][2] : toks[k][1]]
                        if not re.search(r"[,;\-\n/]", sep1) and not re.search(r"[,;\-\n/]", sep2):
                            remove[k] = remove[k - 1] = True
                            k -= 2
                    while k >= 0 and norms[k] in SAFE_PREFIX:
                        sep = s[toks[k][2] : toks[k+1][1]]
                        if re.search(r"[,;\-\n/]", sep):
                            break
                        remove[k] = True
                        k -= 1

    # 3. Loại bỏ thêm bất kỳ đơn vị hành chính (tỉnh, quận/huyện, phường/xã) nào khác của Việt Nam
    # có đi kèm tiền tố hành chính (như "Xã", "Phường", "Huyện",...) và không bị guard.
    all_admin_names = get_all_admin_names()
    admin_prefixes = {"tinh", "tp", "quan", "huyen", "phuong", "xa", "tt", "tx",
                      "t", "q", "h", "p", "f", "x"}
    guards = {
        "duong", "pho", "ngo", "ngach", "hem", "kiet", "ap", "thon",
        "xom", "to", "khu", "kp", "tdp", "doi", "ban", "buon", "lang",
        "truong", "mn", "ubnd", "cho", "ngan", "benh", "bv", "uy", "chua",
        "nha", "khach", "shop", "kho", "xuong", "chung", "cu", "cc",
        "building", "toa", "ks", "khachsan", "nhanghi", "resort", "amlet", "village", "street"
    }
    phrase_guards = {
        "mam non", "tieu hoc", "trung hoc", "benh vien", "ngan hang",
        "uy ban", "khu pho", "to dan", "tieu khu", "chung cu", "toa nha",
        "khach san", "nha nghi", "nha van hoa", "nha tho", "nha tro",
        "nha hang"
    }

    def check_guarded(idx):
        if idx > 0 and not remove[idx - 1]:
            prev_tok = norms[idx - 1]
            if prev_tok in guards:
                return True
            if idx > 1 and not remove[idx - 2]:
                prev_phrase = norms[idx - 2] + " " + norms[idx - 1]
                if prev_phrase in phrase_guards:
                    return True
        return False

    i = 0
    while i < len(norms):
        if remove[i]:
            i += 1
            continue
            
        pfx_len = 0
        if norms[i] in admin_prefixes:
            pfx_len = 1
        elif i + 1 < len(norms) and (norms[i] == "thanh" and norms[i+1] == "pho"):
            pfx_len = 2
        elif i + 1 < len(norms) and (norms[i] == "thi" and norms[i+1] in ("xa", "tran")):
            pfx_len = 2
            
        if pfx_len > 0:
            if not check_guarded(i):
                start_idx = i + pfx_len
                matched_L = 0
                for L in range(3, 0, -1):
                    if start_idx + L <= len(norms):
                        phrase = " ".join(norms[start_idx : start_idx + L])
                        if phrase in all_admin_names:
                            matched_L = L
                            break
                if matched_L > 0:
                    for j in range(i, start_idx + matched_L):
                        remove[j] = True
                    i = start_idx + matched_L
                    continue
        i += 1

    # dựng lại chuỗi: xoá span của token bị remove khỏi s
    spans = [(toks[i][1], toks[i][2]) for i in range(len(toks)) if remove[i]]
    chars = list(s)
    for a, b in spans:
        for k in range(a, b):
            chars[k] = "\x00"
    s = "".join(c for c in chars if c != "\x00")
    # dọn dấu câu / khoảng trắng
    s = unicodedata.normalize("NFC", s)
    s = re.sub(r"(?<=\s)[̀-ͯ]+", "", s)   # bỏ dấu thanh mồ côi
    s = re.sub(r"[\s,;./\-]*$", "", s)
    s = re.sub(r"^[\s,;.\-]*", "", s)
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"\s*,\s*", ", ", s)
    s = re.sub(r"(,\s*){2,}", ", ", s)
    s = re.sub(r",\s*$", "", s).strip(" ,-")
    return _titlecase_vn(s)


# ---------------- luật lọc rác dùng chung (file + UI) ----------------
# đầu cụm là tổ chức -> bỏ cả cụm (đến dấu phẩy, dấu chấm, dấu gạch ngang hoặc xuống dòng)
ORG_RE = re.compile(
    r"^\s*(cửa\s*hàng|ch\b|cty|c\.?ty|công\s*ty|nhà\s*thuốc|quầy\s*thuốc|"
    r"nhà\s*sách|phòng\s*khám|pk\b|bệnh\s*viện|bv\b|phòng\s*giao\s*dịch|"
    r"trường(\s*(tiểu\s*học|thcs|thpt|mầm\s*non|mn|cấp\s*\d))?|"
    r"khách\s*sạn|ks\b|shop|kho\b|xưởng|gara|garage|đại\s*lý|dl\b|"
    r"cơ\s*sở|cs\b|hkd|hộ\s*kinh\s*doanh)\b[^,.\-\n]*",
    re.IGNORECASE)
# xưng hô + tên người ở đầu -> bỏ tới dấu ngăn
REL_RE = re.compile(
    r"^\s*(mẹ|me|bố|bo|ba|chị|chi|anh|ông|ong|bà|cô|co|chú|chu|em|bé|be|"
    r"cụ|cu|thầy|thay|cô\s*giáo|a\.|c\.|e\.)\s+[^,.\-]+[,.\-]",
    re.IGNORECASE)
# mốc định vị -> chỉ cắt trong cụm phân tách bởi dấu phẩy/chấm/gạch ngang
LANDMARK_RE = re.compile(
    r"\b(đối\s*diện|gần|cạnh|kế\s*bên|bên\s*cạnh|sau\s*lưng|trước\s*cổng|"
    r"cổng\s*trường)\b[^,.\-\n]*",
    re.IGNORECASE)
# từ mô tả vị trí (bỏ chữ, GIỮ tên đường/phố phía sau)
POS_RE = re.compile(r"\b(mặt\s*đường|mặt\s*tiền|mặt\s*phố|mt\b)\b", re.IGNORECASE)


# họ phổ biến VN (để nhận tên khách dẫn đầu) — Rule 4
SURNAMES = {
    "nguyen", "tran", "le", "pham", "hoang", "huynh", "phan", "vu", "vo",
    "dang", "bui", "do", "ho", "ngo", "duong", "ly", "dinh", "lam", "phung",
    "mai", "truong", "cao", "chu", "ta", "tong", "kieu", "luong", "la", "ha",
    "quach", "thai", "chau", "trinh", "ninh", "nghiem", "ong",
}
# từ khoá chỉ đường/loại đơn vị dưới xã (để guard, KHÔNG coi là tên người)
STREET_KW = {
    "duong", "pho", "ngo", "ngach", "hem", "kiet", "ql", "quoc", "lo", "so",
    "to", "khu", "ap", "thon", "xom", "khom", "kp", "tdp", "doi", "ban",
    "buon", "lang", "khoi", "tieu",
}
# Rule 1: Phòng/Tầng/Lầu/Căn hộ/Block dẫn đầu (mã phải chứa CHỮ SỐ) — GIỮ tên
# toà nhà/chung cư (không đưa vào đây) để không mất POI.
ROOM_FLOOR_RE = re.compile(
    r"^\s*(?:(?:p|phòng|phong|tầng|tang|lầu|lau|căn\s*hộ|can\s*ho|căn|can|"
    r"ch|block|blk)\b\.?\s*(?=[0-9a-zà-ỹA-Z.\-/]*\d)[0-9a-zà-ỹA-Z.\-/]+"
    r"\s*[,.\-]?\s*)+", re.IGNORECASE)


def _dia(t):
    return strip_diacritics(str(t).lower()).strip(".")


def _strip_room_floor(s):
    """Xoá phòng/tầng/lầu/căn hộ/block + số."""
    s = re.sub(r"\b(p|phòng|phong|tầng|tang|lầu|lau|căn\s*hộ|can\s*ho|căn|can|ch|block|blk)\b\.?\s*(?=[0-9a-zà-ỹA-Z.\-/]*\d)[0-9a-zà-ỹA-Z.\-/]+", " ", s, flags=re.I)
    return s


def strip_customer_name(s):
    if not s:
        return ""
    s = re.sub(r"\s+", " ", s).strip()
    last_s = None
    for _ in range(5):
        if s == last_s:
            break
        last_s = s
        s = _strip_customer_name_once(s)
    return s


def _strip_customer_name_once(s):
    if not s:
        return ""
    s = re.sub(r"\s+", " ", s).strip()
    
    POI_KW = {
        "truong", "cty", "congty", "tnhh", "chua", "shop", "kho", "xuong", "ubnd",
        "ks", "sieuthi", "cho", "gara", "garage", "buudien", "dinh", "mieu"
    }

    # 1. Chị Hạnh, Anh Ánh... -> check title + name
    title_match = re.match(r"^\s*(anh|chị|chi|cô|co|chú|chu|em|a|c|kh|khách\s*hàng|tên\s*kh)\.?\s+", s, re.I)
    if title_match:
        rest_idx = title_match.end()
        words_after = s[rest_idx:].split()
        name_words = []
        for w in words_after[:3]:
            # must be capitalized
            if not re.match(r"^[A-ZÀ-Ỹ][a-zà-ỹ]*$", w):
                break
            # must not be a street/POI keyword or number-like
            w_norm = strip_diacritics(w.lower()).strip(".,-")
            if w_norm in STREET_KW or w_norm in POI_KW or w_norm in {"so", "sn", "nha", "duong", "pho", "ngo", "ngach", "hem", "kiet"}:
                break
            name_words.append(w)
        
        if name_words:
            matched_len = len(" ".join(name_words))
            pos = s.find(" ".join(name_words), rest_idx)
            if pos != -1:
                rest = s[pos + matched_len:].lstrip(" ,.-–\n")
                if rest:
                    first_word = rest.split()[0].lower() if rest.split() else ""
                    first_word_clean = strip_diacritics(first_word).strip(".,-")
                    if first_word_clean in STREET_KW or any(c.isdigit() for c in first_word) or first_word_clean in {"so", "sn"}:
                        return rest

    # 2. Tên + kí tự ngăn cách: Ngọc Yến - Ubnd..., Nguyễn Thuỳ Trang - Sở...
    sep_match = re.match(r"^([A-ZÀ-Ỹ][a-zà-ỹ]*\s*){2,4}\s*[-–:;\n/.]\s*", s)
    if sep_match:
        name_part = s[:sep_match.end()].strip(" -–:;\n/.")
        name_words = [strip_diacritics(w.lower()) for w in name_part.split()]
        if not any(w in STREET_KW or w in POI_KW for w in name_words):
            return s[sep_match.end():].lstrip(" ,.-–\n")

    # 3. Tên + trực tiếp từ khoá đường/số: Nga Nguyễn Xóm Phú Thành..., Bùi Thị Vận Trường...
    words = s.split()
    if len(words) >= 3:
        w1_norm = strip_diacritics(words[0].lower())
        w2_norm = strip_diacritics(words[1].lower())
        w3_norm = strip_diacritics(words[2].lower())
        if (w1_norm in SURNAMES or w2_norm in SURNAMES) and (w1_norm not in STREET_KW and w2_norm not in STREET_KW) and (w1_norm not in POI_KW and w2_norm not in POI_KW):
            if w3_norm in STREET_KW or any(c.isdigit() for c in words[2]) or w3_norm in {"so", "sn"}:
                return " ".join(words[2:])
            
    if len(words) >= 4:
        w1_norm = strip_diacritics(words[0].lower())
        w2_norm = strip_diacritics(words[1].lower())
        w3_norm = strip_diacritics(words[2].lower())
        w4_norm = strip_diacritics(words[3].lower())
        if (w1_norm in SURNAMES or w2_norm in SURNAMES or w3_norm in SURNAMES) and not any(w in STREET_KW or w in POI_KW for w in [w1_norm, w2_norm, w3_norm]):
            if w4_norm in STREET_KW or any(c.isdigit() for c in words[3]) or w4_norm in {"so", "sn"}:
                return " ".join(words[3:])
                
    # 4. Bất kỳ 1-4 từ viết hoa nào dẫn đầu mà ngay sau đó là từ khoá đường/số/POI
    name_len = 0
    for idx, w in enumerate(words[:4]):
        if not re.match(r"^[A-ZÀ-Ỹ][a-zà-ỹ]*$", w):
            break
        w_norm = strip_diacritics(w.lower()).strip(".,-")
        if w_norm == "so":
            if idx + 1 < len(words):
                next_w = words[idx + 1]
                if any(c.isdigit() for c in next_w):
                    break
            else:
                break
        elif w_norm in STREET_KW or w_norm in POI_KW or w_norm in {"sn", "nha"}:
            break
        name_len += 1
    if name_len > 0 and len(words) > name_len:
        next_word = strip_diacritics(words[name_len].lower()).strip(".,-")
        is_num_indic = False
        if next_word == "sn":
            is_num_indic = True
        elif next_word == "so":
            if len(words) > name_len + 1:
                next_next = words[name_len + 1]
                if any(c.isdigit() for c in next_next):
                    is_num_indic = True
        if next_word in STREET_KW or any(c.isdigit() for c in words[name_len]) or is_num_indic or next_word in POI_KW:
            return " ".join(words[name_len:])

    # 5. Tách bằng dấu phân cách (dấu phẩy, gạch ngang, hoặc cuối chuỗi)
    first_part = re.split(r"[,–\-:;\n]", s)[0].strip()
    if first_part:
        fp_words = first_part.split()
        if 2 <= len(fp_words) <= 4:
            all_upper = all(re.match(r"^[A-ZÀ-Ỹ]", w) for w in fp_words)
            if all_upper:
                has_surname = any(strip_diacritics(w.lower()).strip(".,-") in SURNAMES for w in fp_words)
                if has_surname and fp_words[0].lower() not in {"đường", "đg", "dg"}:
                    fp_norms = [strip_diacritics(w.lower()).strip(".,-–") for w in fp_words]
                    STRICT_JUNK_KW = {
                        "duong", "pho", "ngo", "ngach", "hem", "kiet", "nha", "cty",
                        "shop", "kho", "xuong", "chua", "bv", "ubnd", "so", "sn"
                    }
                    has_kw = any(w in STRICT_JUNK_KW for w in fp_norms)
                    has_poi_kw = any(w in POI_KW for w in fp_norms)
                    has_digit = any(any(c.isdigit() for c in w) for w in fp_words)
                    if not has_kw and not has_digit and not has_poi_kw:
                        pattern = re.escape(first_part) + r"\s*[,–\-:;\n]?\s*"
                        s_new = re.sub(pattern, "", s, count=1).strip()
                        return s_new

    return s


def apply_number_rules(detail, strip_names=False):
    """BỎ số nhà + ngõ/ngách/hẻm + phòng/tầng/căn hộ;
    GIỮ đường/phố/thôn xóm/POI. strip_names=True: lọc tên khách."""
    if not detail:
        return ""
    s = unicodedata.normalize("NFC", detail)
    # Dọn dẹp ngoặc đơn trống, số điện thoại hoặc ghi chú giao hàng
    s = re.sub(r"\(\s*[\d\s.,;:\-/]*\s*\)", " ", s)
    s = re.sub(r"\(\s*[^)]*(?:giờ\s*hc|giờ\s*hành\s*chính|gọi|ship|giao|hàng|vỡ|cod)[^)]*\)", " ", s, flags=re.I)
    s = s.replace("(", " ").replace(")", " ")
    # Rule 2: tách số dính từ khoá đường ("195ngo"->"195 ngo")
    s = re.sub(r"(\d)\s?(đường|phố|ngõ|ngách|hẻm|kiệt|quốc\s*lộ|ql|duong|pho|"
               r"ngo|ngach|hem|kiet)\b", r"\1 \2", s, flags=re.I)
    s = re.sub(r"(\d)([A-ZÀ-Ỹ][a-zà-ỹ])", r"\1 \2", s)
    
    # Pre-clean some prefix instructions specifically for shipping
    s = re.sub(r"^\s*(ship\s*mình\s*đến|ship\s*đến|giao\s*đến|địa\s*chỉ\s*giao|đc\s*giao|đ/c\s*giao)\b", "", s, flags=re.I)
    
    # Lọc tên người dẫn đầu
    if strip_names:
        s = strip_customer_name(s)
        
    # Xoá phòng/tầng/block
    s = _strip_room_floor(s)
    
    # Bảo vệ "Đường Số 10", "Phố Số 5"
    s = re.sub(r"\b(đường|phố|duong|pho)\s+số\b", r"\1 ⟪SO⟫", s, flags=re.I)
    # Bỏ hẻm, ngõ, ngách, kiệt,... + số
    s = re.sub(r"\b(ngõ|ngo|ngách|ngach|hẻm|hem|kiệt|kiet)\b\.?\s*\d[\dA-Za-z/]*", " ", s, flags=re.I)
    # bỏ "số nhà 12", "sn 7", "số 5" (tránh xóa nhầm chữ "Số" trong "Số lượng"/"Số kiện" khi đi kèm chữ thường)
    s = re.sub(r"\b(số\s*nhà|sn)\b\.?\s*\d*[a-zA-Z]?", " ", s, flags=re.I)
    s = re.sub(r"\b(số)\b\.?\s*\d+[a-zA-Z]?", " ", s, flags=re.I)
    # bỏ số nhà/lô dẫn đầu mỗi cụm: "391", "12a", "30/106", "j63-64", "34 /9" (chấp nhận dấu phẩy hoặc khoảng trắng/hết dòng ở cuối)
    s = re.sub(r"(^|,)\s*[a-zA-Z]{0,2}\d+[a-zA-Z]?(?:\s*[/\-]\s*\d+[a-zA-Z]?)*(?:\s+|,|$)",
               r"\1 ", s)
    # dọn vụn "/9" hoặc dấu xuyệt-số còn sót ở đầu cụm
    s = re.sub(r"(^|,)\s*[/\-]\s*\d+[a-zA-Z]?(?:\s+|,|$)", r"\1 ", s)
    # bỏ viết tắt quận/phường dính số: q1, p5 (chỉ 1-2 số; giữ kp11 & P.1203)
    s = re.sub(r"(?<!\w)[qpQP]\.?\s?\d{1,2}[a-zA-Z]?(?!\w)", " ", s)
    s = s.replace("⟪SO⟫", "Số")
    s = re.sub(r"\s+[.]+\s+", " ", s)          # bỏ dấu chấm mồ côi giữa cụm
    s = re.sub(r"\s*,\s*", ", ", s)
    s = re.sub(r"(,\s*){2,}", ", ", s)
    s = re.sub(r"\s*[-–]+\s*(?=[,;.]|$)", " ", s)
    s = re.sub(r"(^|[,;.])\s*[-–]+\s*", r"\1 ", s)
    s = re.sub(r"\s*[-–]+(?:\s*[-–]+)+\s*", " ", s)
    s = re.sub(r"\s+", " ", s).strip(" ,.-")
    if re.fullmatch(r"(đường|phố|ngõ|ngo|ngách|hẻm|số|khu|tổ|đội|mặt)\.?",
                    s.strip(), flags=re.I):
        s = ""
    return s


POI_KEYWORDS = {
    "truong", "mn", "mam non", "tieu hoc", "trung hoc", "thpt", "thcs", "cao dang",
    "dai hoc", "cty", "cong ty", "tnhh", "chua", "benh vien", "toa", "tòa",
    "bv", "phong kham", "pk", "nha thuoc", "quay thuoc",
    "nha sach", "cua hang", "shop", "kho",
    "xuong", "ben xe", "cang", "ubnd", "uy ban",
    "ngan hang", "toa nha", "building", "chung cu",
    "khach san", "ks", "nha nghi", "sieu thi",
    "cho", "trung tam", "bhxh",
    "bao hiem", "nha van hoa", "nha ga",
    "nha tho", "nha tro", "nha khach",
    "gara", "garage", "dai ly", "co so", "hkd",
    "ho kinh doanh", "buu dien", "mieu",
    "dinh", "nghia trang", "ttyt", "tram y te"
}

POI_KW_SET = {
    "ngan", "truong", "benh", "vien", "uy", "ban", "ty", "cty", "cong", "tnhh",
    "ch", "cua", "shop", "kho", "xuong", "chua", "nha", "ca", "doan", "tram",
    "toa", "building", "chung", "cu", "ks", "khach", "san", "sieu", "cho",
    "ubnd", "mam", "non", "tieu", "hoc", "trung", "hoc", "thpt", "thcs", "thuoc",
    "quay", "phong", "kham", "pk", "sach", "gara", "garage", "dai", "ly", "dl",
    "co", "so", "cs", "hkd", "ho", "kinh", "doanh", "y", "te", "buu", "dien",
    "dinh", "mieu"
}

JUNK_SEGMENT_RE = re.compile(
    r"\b(ship|giao hang|giao hc|hanh chinh|hanh chanh|dong kiem|cho xem|khong xem|"
    r"hang de vo|thu ho|cod|chuyen khoan|ck|cuoc|phi ship|tien ship|size|sz|mau|kien|"
    r"cai|chiec|bo|kg|g|k|d|vnd|vnđ|nhe|nhe\s*e|nha|gui chi|gio hanh chinh|"
    r"t2|t3|t4|t5|t6|t7|cn|thu\s*\d|chu\s*nhat)\b",
    re.IGNORECASE
)

PRONOUN_NAME_RE = re.compile(
    r"^\s*(?:anh|chị|chi|cô|co|chú|chu|em|a|c|bố|bo|bà|ba|ông|ong|cụ|cu|thầy|thay|mẹ|me|bé|be|khách\s*hàng|khach\s*hang|tên\s*kh|ten\s*kh|người\s*nhận|nguoi\s*nhan)\.?\s+([a-zà-ỹA-ZÀ-Ỹ]+(?:\s+[a-zà-ỹA-ZÀ-Ỹ]+){0,2})\b",
    re.IGNORECASE
)

INSTRUCTION_RE = re.compile(
    r"^\s*(?:ship\s+mình\s+đến|ship\s+minh\s+den|ship\s+đến|ship\s+den|ship\s+hộ\s+đến|ship\s+ho\s+den|ship\s+cho|ship|giao\s+đến|giao\s+den|giao\s+hộ\s+đến|giao\s+ho\s+den|địa\s+chỉ\s+giao|dia\s+chi\s+giao|địa\s+chỉ|dia\s+chi|đc\s+giao|dc\s+giao|đ/c\s+giao|d/c\s+giao|đc|dc|d/c|gửi\s+chị\s+đi\s+nhé\s+về|gui\s+chi\s+di\s+nhe\s+ve|gửi\s+chị|gui\s+chi|gửi|gui|bán\s+cho|ban\s+cho|khách\s*hàng|khach\s*hang|tên\s*kh|ten\s*kh|người\s*nhận|nguoi\s*nhan)\b\.?\s*(?:[:：\-]+)?\s*",
    re.IGNORECASE
)

ADDR_MARK = "⟪ADDR⟫"

ADMIN_ABBR_WORDS = {
    "hcm", "tphcm", "tp hcm", "sg", "hn", "hp", "dn", "bd", "bp", "brvt",
    "vt", "daklak", "dak lak", "daclak", "dac lac", "dak nong", "dac nong",
}


def normalize_common_abbreviations(s):
    """Chuẩn hoá các viết tắt địa chỉ phổ biến trước khi token hoá.

    Chỉ mở rộng các mẫu có ngữ cảnh rõ để tránh biến nhầm tên riêng thành từ
    khoá địa chỉ.
    """
    s = unicodedata.normalize("NFC", str(s or ""))
    s = re.sub(r"\btthcs\b", "Trường THCS", s, flags=re.I)
    s = re.sub(r"\btrường\s+mn\b", "Trường Mầm Non", s, flags=re.I)
    s = re.sub(r"\bmn\s+(?=[A-ZÀ-Ỹ])", "Mầm Non ", s, flags=re.I)
    s = re.sub(r"\bbv\b", "Bệnh Viện", s, flags=re.I)
    s = re.sub(r"\bub\s*nd\b|\bubnd\b", "UBND", s, flags=re.I)
    s = re.sub(r"\bcty\b|\bc\.ty\b", "Công Ty", s, flags=re.I)
    s = re.sub(r"\bcn\b", "CN", s, flags=re.I)
    s = re.sub(r"\bkcn\b", "Khu Công Nghiệp", s, flags=re.I)
    s = re.sub(r"\bkdc\b", "Khu Dân Cư", s, flags=re.I)
    s = re.sub(r"\bkdt\b", "Khu Đô Thị", s, flags=re.I)
    s = re.sub(r"\bkcx\b", "Khu Chế Xuất", s, flags=re.I)
    s = re.sub(r"\bcc\b", "Chung Cư", s, flags=re.I)
    s = re.sub(r"\bkp\.?\s*(\d+[a-zA-Z]?)\b", r"Khu Phố \1", s, flags=re.I)
    s = re.sub(r"\btdp\.?\s*(\d+[a-zA-Z]?)\b", r"Tổ Dân Phố \1", s, flags=re.I)
    s = re.sub(r"\b(khu\s*phố|khu\s*pho)\s*(\d)", r"\1 \2", s, flags=re.I)
    s = re.sub(
        r"\b(ấp|ap|thôn|thon|xóm|xom|tổ|to|đội|doi|khóm|khom|bản|ban|buôn|buon)\s*(\d+[a-zA-Z]?)\b",
        r"\1 \2", s, flags=re.I)
    return s


def _plain_norm(s):
    return re.sub(r"\s+", " ", strip_diacritics(str(s or "").lower())).strip()


def _compact_norm(s):
    return re.sub(r"[^a-z0-9]+", "", strip_diacritics(str(s or "").lower()))


def _add_admin_abbrs(kill):
    """Nếu đã biết tỉnh từ cột chuẩn, thêm các viết tắt tương ứng để xoá."""
    joined = " ".join(kill)
    for abbr, full in PROV_ABBR.items():
        if full in kill or full in joined:
            kill.add(abbr)
            kill.add(abbr.replace(" ", ""))
    for phrase in list(kill):
        if " " in phrase:
            kill.add(phrase.replace(" ", ""))


def merge_level4_segments(segments):
    merged = []
    i = 0
    while i < len(segments):
        seg = (segments[i] or "").strip()
        if not seg:
            i += 1
            continue
        n = _plain_norm(seg)
        nxt = (segments[i + 1] or "").strip() if i + 1 < len(segments) else ""
        nn = _plain_norm(nxt)
        if (n == "khu" or n.endswith(" khu")) and (nn == "pho" or nn.startswith("pho")):
            merged.append(f"{seg} {nxt}".strip())
            i += 2
            continue
        if n == "tieu" and (nn == "khu" or nn.startswith("khu")):
            merged.append(f"{seg} {nxt}".strip())
            i += 2
            continue
        if n in {"to dan", "to"} and (nn == "pho" or nn.startswith("dan pho") or nn.startswith("pho")):
            merged.append(f"{seg} {nxt}".strip())
            i += 2
            continue
        merged.append(seg)
        i += 1
    return merged


LEADING_HOUSE_RE = re.compile(
    r"^\s*(?:số\s*nhà|so\s*nha|sn|số|so|nhà|nha)?\s*"
    r"(?:[A-Za-z]{0,3}\d+[A-Za-z]?(?:\s*[/\-]\s*\d+[A-Za-z]?)*"
    r"(?:\s+[A-Za-z](?=\s|[,.\-–/]|$))?\s*)+", re.I)


def strip_leading_house_number(seg):
    n = _plain_norm(seg)
    if re.match(r"^(ql|tl|hl|dt|km)\s*\d", n) or re.match(
            r"^(quoc lo|tinh lo|huong lo|duong tinh|dai lo|xa lo|cao toc)\b", n):
        return seg, False
    m = LEADING_HOUSE_RE.match(seg)
    if not m:
        return seg, False
    rest = seg[m.end():].lstrip(" ,.-–/:")
    return rest, True


def is_admin_or_noise_segment(seg):
    n = _plain_norm(seg).strip(" .,-")
    c = _compact_norm(seg)
    if not n:
        return True
    if n in {"cu", "moi", "cu moi", "nhe", "nha e", "a", "e", "vietnam", "viet nam"}:
        return True
    if c in {re.sub(r"[^a-z0-9]+", "", x) for x in ADMIN_ABBR_WORDS}:
        return True
    if re.fullmatch(r"[pq]\.?\d{1,2}[a-z]?", n) or re.fullmatch(r"(tp|tt|tx|q|p|h|x)\.?", n):
        return True
    if re.match(r"^(tu dau|roi|re trai|re phai|dang sau|dang truoc|doi dien|gan|canh|ke ben|ben canh)\b", n):
        return True
    if re.fullmatch(r"(size|sz|mau|cao|nang|kg|cod|thu ho).*", n):
        return True
    all_admin = get_all_admin_names()
    for suffix in (" cu", " moi"):
        if n.endswith(suffix):
            base = n[:-len(suffix)].strip()
            if base in all_admin or base.replace(" ", "") in all_admin:
                return True
    # We comment this out to protect valid street/hamlet names that share names with wards elsewhere in Vietnam
    # if n in all_admin or c in all_admin:
    #     return True
    for abbr, full in PROV_ABBR.items():
        if n == abbr or c == abbr.replace(" ", "") or n == full or c == full.replace(" ", ""):
            return True
    return False


def trim_segment_notes(seg):
    seg = re.sub(
        r"\b(?:lưu\s*ý\s*vận\s*chuyển|luu\s*y\s*van\s*chuyen|ship\s*giờ|ship\s*gio|"
        r"giao\s*giờ|giao\s*gio|giờ\s*hc|gio\s*hc|giờ\s*hành\s*chính|gio\s*hanh\s*chinh)\b.*$",
        " ", seg, flags=re.I)
    seg = re.sub(r"\b(?:ko|không|khong)\s+đúng\s+hàng\b.*$", " ", seg, flags=re.I)
    seg = re.sub(r"\b(?:nhé|nhe|nha\s*e|ạ)\b.*$", " ", seg, flags=re.I)
    seg = LANDMARK_RE.sub(" ", seg)
    seg = POS_RE.sub(" ", seg)
    return re.sub(r"\s+", " ", seg).strip(" ,.-–/:")


def strip_inline_admin_tokens(seg):
    seg = re.sub(r"(?<!\w)[pq]\.?\s?\d{1,2}[a-zA-Z]?(?!\w)", " ", seg, flags=re.I)
    seg = re.sub(
        r"\b(?:hcm|tphcm|hn|hp|bd|bp|brvt|vt|vietnam|viet\s*nam|daklak|daclak|"
        r"dak\s*lak|dac\s*lac)\b",
        " ", seg, flags=re.I)
    return re.sub(r"\s+", " ", seg).strip(" ,.-–/:")


def strip_known_admin_tail(seg, admin_names, skip_ward=False):
    tail_names = set()
    for idx, nm in enumerate(admin_names):
        if not nm:
            continue
        if skip_ward and idx in {0, 3, 6}:
            continue
        b = _base(nm)
        f = _norm(nm).strip()
        if b:
            tail_names.add(b)
            tail_names.add(b.replace(" ", ""))
        if f:
            tail_names.add(f)
            tail_names.add(f.replace(" ", ""))
    joined = " ".join(tail_names)
    for abbr, full in PROV_ABBR.items():
        if full in tail_names or full in joined:
            tail_names.add(abbr)
            tail_names.add(abbr.replace(" ", ""))

    out = seg
    for _ in range(4):
        toks = [(m.group(), m.start(), m.end()) for m in re.finditer(r"[^\W_]+", out, re.UNICODE)]
        if len(toks) < 2:
            break
        norms = [strip_diacritics(w.lower()) for w, _, _ in toks]
        removed = False
        for L in range(min(5, len(toks) - 1), 0, -1):
            phrase = " ".join(norms[-L:])
            compact = phrase.replace(" ", "")
            if phrase in tail_names or compact in tail_names:
                out = out[:toks[-L][1]].rstrip(" ,.-–/:")
                removed = True
                break
        if not removed:
            break
    return out


def protect_unit_numbers(s):
    s = re.sub(r"\b(tổ\s+dân\s+phố|to\s+dan\s+pho)\s+số\s+(\d+[a-zA-Z]?)\b",
               r"\1⟪\2⟫", s, flags=re.I)
    s = re.sub(r"\b(đường|duong)\s+số\s+(\d+)\b", r"\1⟪Số⟫⟪\2⟫", s, flags=re.I)
    s = re.sub(r"\b(đường|duong)\s+(\d+(?:/\d+)?)\b", r"\1⟪\2⟫", s, flags=re.I)
    s = re.sub(r"\b(phố|pho)\s+(\d+)\b", r"\1⟪\2⟫", s, flags=re.I)
    s = re.sub(r"\b(quốc\s+lộ|quoc\s+lo|ql|tl|hl|đt|dt)\s+(\d+[a-zA-Z]?)\b", r"\1⟪\2⟫", s, flags=re.I)
    s = re.sub(r"\b(xóm|xom|thôn|thon|ấp|ap|tổ|to|khu|khối|khoi|kp|tdp|đội|doi)\s+(\d+[a-zA-Z]?)\b", r"\1⟪\2⟫", s, flags=re.I)
    s = re.sub(r"\b(km)\s+(\d+(?:\+\d+)?)\b", r"\1⟪\2⟫", s, flags=re.I)
    return s

def restore_unit_numbers(s):
    s = s.replace("⟪Số⟫", " Số ")
    s = re.sub(r"⟪([^⟫]+)⟫", r" \1", s)
    return s

def is_customer_name(seg):
    seg_clean = PRONOUN_NAME_RE.sub("", seg).strip()
    words = seg_clean.split()
    if 2 <= len(words) <= 4:
        # A customer name cannot contain digits in any of its words
        if any(any(c.isdigit() for c in w) for w in words):
            return False
        all_cap = all(re.match(r"^[A-ZÀ-Ỹ]", w) for w in words)
        if all_cap:
            w1_norm = strip_diacritics(words[0].lower())
            w2_norm = strip_diacritics(words[1].lower()) if len(words) > 1 else ""
            if w1_norm in SURNAMES or w2_norm in SURNAMES:
                norms_seg = [strip_diacritics(w.lower()) for w in words]
                address_keywords = {
                    "duong", "pho", "ngo", "ngach", "hem", "kiet", "ap", "thon",
                    "xom", "to", "khu", "kp", "tdp", "doi", "ban", "buon", "lang",
                    "ql", "tl", "hl", "dt", "km", "quoc", "lo", "so", "sn", "khom"
                }
                if not any(w in address_keywords or w in POI_KEYWORDS for w in norms_seg):
                    return True
    return False

def classify_segment(seg):
    seg_norm = " " + strip_diacritics(seg.lower()) + " "
    raw_words = [w.lower().strip(".,-–/:") for w in seg.split()]
    
    level4_indicators = {
        "duong", "pho", "ngo", "ngach", "hem", "kiet", "ap", "thon", "xom", 
        "to", "khu", "kp", "tdp", "doi", "ban", "buon", "lang", "ql", "tl", 
        "hl", "dt", "km", "quoc lo", "tinh lo", "dai lo", "xa lo", "khoi", "khom"
    }
    
    strong_poi_keywords = {
        "truong", "mn", "mam non", "tieu hoc", "trung hoc", "thpt", "thcs", "cao dang",
        "dai hoc", "cty", "cong ty", "tnhh", "chua", "benh vien", "bv", "phong kham", 
        "pk", "nha thuoc", "quay thuoc", "nha sach", "cua hang", "shop", "kho",
        "xuong", "ben xe", "cang", "ubnd", "uy ban", "ngan hang", "toa nha", 
        "building", "chung cu", "khach san", "ks", "nha nghi", "sieu thi", 
        "cho", "trung tam", "bhxh", "bao hiem", "nha van hoa", "nha ga", 
        "nha tho", "nha tro", "nha khach", "gara", "garage", "dai ly", "co so", 
        "hkd", "ho kinh doanh", "buu dien", "mieu", "nghia trang", "ttyt", "tram y te",
        "ben xe", "so tai chinh", "toa an", "xe may"
    }
    if any(w == "sở" for w in raw_words):
        return "POI"
    
    first_l4_idx = len(seg_norm)
    for kw in level4_indicators:
        pos = seg_norm.find(f" {kw} ")
        if pos != -1 and pos < first_l4_idx:
            first_l4_idx = pos
            
    first_poi_idx = len(seg_norm)
    for kw in strong_poi_keywords:
        pos = seg_norm.find(f" {kw} ")
        if pos != -1 and pos < first_poi_idx:
            first_poi_idx = pos
            
    if first_poi_idx < first_l4_idx:
        return "POI"
    elif first_l4_idx < first_poi_idx:
        return "L4"
    else:
        return "L4"

def extract_level4_and_poi(raw, admin_names, debug=False):
    if not raw:
        return "", ""
    
    if debug:
        print(f"[DEBUG] Raw input: {raw}")

    # ==========================================
    # MODULE 1: Tiền xử lý & Phiên dịch (De-abbreviation)
    # ==========================================
    
    # 1. NFC normalization
    s = unicodedata.normalize("NFC", str(raw))
    
    # Save raw lines for standalone-name detection in Module 6
    _raw_lines = {line.strip().lower() for line in str(raw).split('\n') if line.strip()}
    
    # 2. Ward/District de-abbreviations (regex)
    s = re.sub(r"\b[Pp]\.?\s*(\d+)\b", r"Phường \1", s)
    s = re.sub(r"\b[Qq]\.?\s*(\d+)\b", r"Quận \1", s)
    
    # 3. City/Province dictionary mapping (case-insensitive)
    s = re.sub(r"\b(?:tp\.hcm|tp\s*hcm|hcm|sg)\b", "Hồ Chí Minh", s, flags=re.I)
    s = re.sub(r"\b(?:hn)\b", "Hà Nội", s, flags=re.I)
    s = re.sub(r"\b(?:đn)\b", "Đà Nẵng", s, flags=re.I)
    s = re.sub(r"\b(?:brvt|br-vt)\b", "Bà Rịa Vũng Tàu", s, flags=re.I)
    s = re.sub(r"\b(?:tp[\s\.\-]*vt|vt)\b", "Vũng Tàu", s, flags=re.I)
    s = re.sub(r"\b(?:dnai)\b", "Đồng Nai", s, flags=re.I)
    s = re.sub(r"\b(?:bd)\b", "Bình Dương", s, flags=re.I)
    s = re.sub(r"\b(?:la)\b", "Long An", s, flags=re.I)
    s = re.sub(r"\b(?:hp)\b", "Hải Phòng", s, flags=re.I)
    s = re.sub(r"\b(?:qn)\b", "Quảng Ninh", s, flags=re.I)
    s = re.sub(r"\b(?:py)\b", "Phú Yên", s, flags=re.I)
    s = re.sub(r"\b(?:kh)\b", "Khánh Hòa", s, flags=re.I)
    s = re.sub(r"\b(?:tb)\b", "Thái Bình", s, flags=re.I)
    s = re.sub(r"\b(?:nd)\b", "Nam Định", s, flags=re.I)
    s = re.sub(r"\b(?:hy)\b", "Hưng Yên", s, flags=re.I)
    s = re.sub(r"\b(?:hd)\b", "Hải Dương", s, flags=re.I)
    s = re.sub(r"\b(?:bg)\b", "Bắc Giang", s, flags=re.I)
    s = re.sub(r"\b(?:bn)\b", "Bắc Ninh", s, flags=re.I)
    s = re.sub(r"\b(?:cb)\b", "Cao Bằng", s, flags=re.I)
    s = re.sub(r"\b(?:ls)\b", "Lạng Sơn", s, flags=re.I)
    s = re.sub(r"\b(?:tq)\b", "Tuyên Quang", s, flags=re.I)
    s = re.sub(r"\b(?:yb)\b", "Yên Bái", s, flags=re.I)
    s = re.sub(r"\b(?:lc)\b", "Lai Châu", s, flags=re.I)
    s = re.sub(r"\b(?:db)\b", "Điện Biên", s, flags=re.I)
    s = re.sub(r"\b(?:sl)\b", "Sơn La", s, flags=re.I)
    s = re.sub(r"\b(?:hb)\b", "Hòa Bình", s, flags=re.I)
    s = re.sub(r"\b(?:qb)\b", "Quảng Bình", s, flags=re.I)
    s = re.sub(r"\b(?:qt)\b", "Quảng Trị", s, flags=re.I)
    s = re.sub(r"\b(?:tth)\b", "Thừa Thiên Huế", s, flags=re.I)
    s = re.sub(r"\b(?:qnam)\b", "Quảng Nam", s, flags=re.I)
    s = re.sub(r"\b(?:qng)\b", "Quảng Ngãi", s, flags=re.I)
    s = re.sub(r"\b(?:bdinh)\b", "Bình Định", s, flags=re.I)
    s = re.sub(r"\b(?:gl)\b", "Gia Lai", s, flags=re.I)
    s = re.sub(r"\b(?:kon\s*tum)\b", "Kon Tum", s, flags=re.I)
    s = re.sub(r"\b(?:dlak|dac\s*lac|dak\s*lak|dac\s*lak)\b", "Đắk Lắk", s, flags=re.I)
    s = re.sub(r"\b(?:dkno)\b", "Đắk Nông", s, flags=re.I)
    s = re.sub(r"\b(?:ldong)\b", "Lâm Đồng", s, flags=re.I)
    s = re.sub(r"\b(?:nt)\b", "Ninh Thuận", s, flags=re.I)
    s = re.sub(r"\b(?:bt)\b", "Bình Thuận", s, flags=re.I)
    s = re.sub(r"\b(?:tnin)\b", "Tây Ninh", s, flags=re.I)
    s = re.sub(r"\b(?:tgiang)\b", "Tiền Giang", s, flags=re.I)
    s = re.sub(r"\b(?:btre)\b", "Bến Tre", s, flags=re.I)
    s = re.sub(r"\b(?:vlong)\b", "Vĩnh Long", s, flags=re.I)
    s = re.sub(r"\b(?:dthap)\b", "Đồng Tháp", s, flags=re.I)
    s = re.sub(r"\b(?:ag)\b", "An Giang", s, flags=re.I)
    s = re.sub(r"\b(?:kgiang)\b", "Kiên Giang", s, flags=re.I)
    s = re.sub(r"\b(?:ctho)\b", "Cần Thơ", s, flags=re.I)
    s = re.sub(r"\b(?:hgiang)\b", "Hậu Giang", s, flags=re.I)
    s = re.sub(r"\b(?:st)\b", "Sóc Trăng", s, flags=re.I)
    s = re.sub(r"\b(?:bl)\b", "Bạc Liêu", s, flags=re.I)
    s = re.sub(r"\b(?:cm)\b", "Cà Mau", s, flags=re.I)
    s = re.sub(r"\b(?:tv)\b", "Trà Vinh", s, flags=re.I)
    
    # 4. Character normalization: replace '-', '\n', '(', ')' with ',' and add spaces around commas to prevent word merging
    for char in ['-', '\n', '(', ')', '\r']:
        s = s.replace(char, ',')
    s = s.replace(',', ', ')
    s = re.sub(r"\s+", " ", s)
        
    if debug:
        print(f"[DEBUG] After Module 1 (De-abbreviation): {s}")
        
    # ==========================================
    # MODULE 2: Gọt tỉa rác bề mặt & Rẽ nhánh (Pruning)
    # ==========================================
    
    # 1. Remove shipping prefixes at the beginning
    prefix_match = re.match(r"^\s*(?:đc|đ/c|dc|địa\s+chỉ|ship\s+đến|sđt|sdt|người\s+nhận|ng\s+nhận|ng\s+nhan|tên\s+người\s+nhận|khách\s+hàng)\s*[:：\-]*\s*", s, flags=re.I)
    if prefix_match:
        s = s[prefix_match.end():]
    else:
        # 2. Strip customer name using the defined strip_customer_name function
        s = strip_customer_name(s)
    
    # 3. Clean phone numbers & other garbage
    from parse_address import extract_phone
    ph = extract_phone(s)
    if ph:
        s = s.replace(ph, " ")
    s = re.sub(r"\+?\d[\d\.\-\s]{8,}\d", " ", s)
    s = remove_order_garbage(s)
    
    # 4. Remove sub-alleys (Ngõ, Ngách, Hẻm, Kiệt + digit sequence)
    s = re.sub(r"\b(?:ngõ|ngách|hẻm|kiệt|ngo|ngach|hem|kiet)\b\.?\s*\d+[\d\/\-]*[a-zA-Z]{0,2}\b", " ", s, flags=re.I)
    
    # 5. Remove house numbers with lookbehind protection for region identifiers
    protect_singles = [
        'to', 'ấp', 'ap', 'tổ', 'khu', 'thôn', 'thon', 'xóm', 'xom', 
        'đội', 'doi', 'khối', 'khoi', 'khóm', 'khom', 'đường', 'duong', 
        'phố', 'pho', 'ql', 'tl', 'hl', 'dt', 'km', 'lộ', 'lo'
    ]
    protect_multis = [
        'đường số', 'duong so', 'đường so', 'duong số',
        'phố số', 'pho so', 'phố so', 'pho số',
        'tổ số', 'to so', 'tổ so', 'to số',
        'khu số', 'khu so', 'ấp số', 'ap so',
        'thôn số', 'thon so', 'xóm số', 'xom so',
        'khu phố số', 'khu pho so', 'tổ dân phố số', 'to dan pho so'
    ]
    lookbehinds = []
    for w in protect_singles:
        lookbehinds.append(f'(?<!{w}\\s)')
    for w in protect_multis:
        ws = w.replace(' ', '\\s')
        lookbehinds.append(f'(?<!{ws}\\s)')
        
    s = re.sub(r"\s+", " ", s)
    # Improved house number regex to avoid swallowing full words like "Mayplaza" or names
    pat_str = r'\b(?:số\s*nhà|số|so\s*nha|so|sn)?\s*' + ''.join(lookbehinds) + r'\d+[a-zA-Z]?(?:\s*[/\-]\s*[a-zA-Z]?\d+[a-zA-Z]?)*\b'
    s = re.compile(pat_str, re.IGNORECASE).sub(" ", s)
    s = re.sub(r"\s+", " ", s).strip()
    
    if debug:
        print(f"[DEBUG] After Module 2 (Pruning): {s}")

    # ==========================================
    # MODULE 3: Khớp & Trừ lùi Hành chính (Subtractive Matching)
    # ==========================================
    
    # Tokenize current string s
    toks = [(m.group(), m.start(), m.end()) for m in re.finditer(r"[^\W_]+", s, re.UNICODE)]
    norms = [strip_diacritics(w.lower()) for w, _, _ in toks]
    
    kill_phrases = set()
    for nm in admin_names:
        if not nm:
            continue
        nm_str = str(nm).strip()
        if not nm_str:
            continue
        # Add original name
        kill_phrases.add(nm_str)
        # Add unaccented version
        kill_phrases.add(strip_diacritics(nm_str))
        # Add base name (without "Phường", "Quận" etc.)
        b = _base(nm_str)
        if b:
            kill_phrases.add(b)
            kill_phrases.add(strip_diacritics(b))
            
    # We want to match unaccented kill phrases
    kill_norms = set()
    for phrase in kill_phrases:
        p_norm = strip_diacritics(phrase.lower())
        kill_norms.add(p_norm)
        p_collapsed = p_norm.replace(" ", "")
        if p_collapsed:
            kill_norms.add(p_collapsed)
    
    _add_admin_abbrs(kill_norms)
    
    candidate_matches = []
    kills_sorted = sorted(list(kill_norms), key=lambda x: -len(x.split()))
    for kph in kills_sorted:
        kw = kph.split()
        L = len(kw)
        if L == 0:
            continue
        for i in range(len(norms) - L + 1):
            if norms[i:i + L] == kw:
                candidate_matches.append((i, i + L, kph))

    candidate_matches.sort(key=lambda x: -(x[1] - x[0]))
    final_matches = []
    covered = set()
    for start, end, kph in candidate_matches:
        if not any(idx in covered for idx in range(start, end)):
            final_matches.append((start, end, kph))
            for idx in range(start, end):
                covered.add(idx)
                
    # Now build remove spans with guards
    guards = {
        "duong", "pho", "ngo", "ngach", "hem", "kiet", "ap", "thon",
        "xom", "to", "khu", "kp", "tdp", "doi", "ban", "buon", "lang",
        "truong", "mn", "ubnd", "cho", "ngan", "benh", "bv", "uy", "chua",
        "nha", "khach", "shop", "kho", "xuong", "o", "so", "khom", "ql",
        "tl", "hl", "dt", "km", "lo", "sn"
    }
    phrase_guards = {
        "mam non", "tieu hoc", "trung hoc", "benh vien", "ngan hang",
        "uy ban", "khu pho", "to dan", "tieu khu"
    }
    
    # Map tokens to raw_norm indices to support preceding digit checks in raw address
    tok_raw_indices = []
    last_raw_idx = 0
    raw_norm = strip_diacritics(raw.lower())
    for w, _, _ in toks:
        w_norm = strip_diacritics(w.lower())
        idx = raw_norm.find(w_norm, last_raw_idx)
        if idx != -1:
            tok_raw_indices.append(idx)
            last_raw_idx = idx + len(w_norm)
        else:
            tok_raw_indices.append(-1)

    remove_spans = []
    for start, end, kph in final_matches:
        is_guarded = False
        if start > 0:
            prev_tok = norms[start - 1]
            if prev_tok in guards and (start - 1) not in covered:
                is_guarded = True
            if start > 1:
                prev_phrase = norms[start - 2] + " " + norms[start - 1]
                if prev_phrase in phrase_guards and (start - 2) not in covered and (start - 1) not in covered:
                    is_guarded = True
                    
        # Check if the matched token in the raw string was preceded by a digit or house number
        if not is_guarded and start < len(tok_raw_indices):
            raw_start_idx = tok_raw_indices[start]
            if raw_start_idx != -1:
                # Do not apply digit guard if the matched name starts with an administrative prefix keyword
                starts_with_admin = False
                for prefix in ["phuong", "quan", "thanh pho", "tinh", "huyen", "xa", "thi xa", "thi tran"]:
                    if kph.startswith(prefix):
                        starts_with_admin = True
                        break
                    before_text = raw_norm[:raw_start_idx]
                    before_clean = before_text.rstrip(" ,.-–/:")
                    if re.search(r"\d+$", before_clean) or re.search(r"\b(?:so|sn|nha|so nha)\b$", before_clean):
                        is_guarded = True
                    
        if is_guarded:
            if debug:
                print(f"[DEBUG] Guarded match from removal: {toks[start][0]} -> {toks[end-1][0]} (kph={kph})")
            continue
            
        remove_spans.append((toks[start][1], toks[end-1][2]))
        
        # Preceding admin prefix: e.g. "Xã Sông Lô" -> remove "Xã"
        k = start - 1
        # Check two-word prefixes
        if k >= 1 and (norms[k-1], norms[k]) in PREFIX_PAIRS:
            remove_spans.append((toks[k-1][1], toks[k][2]))
            k -= 2
        # Check single-word prefixes
        while k >= 0 and norms[k] in SAFE_PREFIX:
            remove_spans.append((toks[k][1], toks[k][2]))
            k -= 1
            
    # Remove these spans from s
    chars = list(s)
    for a, b in remove_spans:
        for idx in range(a, b):
            chars[idx] = "\x00"
    s = "".join(c for c in chars if c != "\x00")
    s = re.sub(r",\s*,", ",", s)
    s = re.sub(r"\s+", " ", s).strip(" ,")
    
    if debug:
        print(f"[DEBUG] After Module 3 (Subtractive Matching): {s}")

    # ==========================================
    # MODULE 4: Trích xuất Điểm định vị (POI Extraction)
    # ==========================================
    
    segments = [seg.strip() for seg in s.split(",") if seg.strip()]
    
    level4_indicators = {
        "duong", "pho", "ngo", "ngach", "hem", "kiet", "ap", "thon", "xom", 
        "to", "khu", "kp", "tdp", "doi", "ban", "buon", "lang", "ql", "tl", 
        "hl", "dt", "km", "quoc lo", "tinh lo", "dai lo", "xa lo", "khoi"
    }
    
    poi_kws = [
        "truong", "mn", "mam non", "tieu hoc", "trung hoc", "thpt", "thcs", "ubnd",
        "ngan hang", "benh vien", "bv", "tru so", "cho", "quan", "cong ty", "cty",
        "tnhh", "nha thuoc", "quay thuoc", "phong kham", "pk", "sieu thi", "cua hang",
        "shop", "kho", "xuong", "ben xe", "cang", "toa nha", "building", "chung cu",
        "khach san", "ks", "nha nghi", "trung tam", "buu dien", "chua", "nha tho",
        "kcn", "khu cong nghiep", "kcx", "khu che xuat", "kdc", "khu dan cu",
        "kdt", "khu do thi", "nha may", "cong vien", "nha hang"
    ]
    
    pois = []
    remaining_segments = []
    
    for seg in segments:
        seg_norm = " " + strip_diacritics(seg.lower()) + " "
        
        # Find earliest level 4 index
        first_l4_idx = len(seg_norm)
        for kw in level4_indicators:
            pos = seg_norm.find(f" {kw} ")
            if pos != -1 and pos < first_l4_idx:
                first_l4_idx = pos
                
        # Find earliest POI index (must start within the first 2 words of the segment)
        first_poi_idx = len(seg_norm)
        matched_kw_len = 0
        for kw in poi_kws:
            pos = seg_norm.find(f" {kw} ")
            if pos != -1 and pos < first_poi_idx:
                before_match = seg_norm[:pos]
                words_before = before_match.split()
                if len(words_before) < 2:
                    first_poi_idx = pos
                    matched_kw_len = len(kw)
                    
        # Compare indices
        if first_poi_idx < first_l4_idx:
            # We slice the original segment starting from the matched POI keyword
            # Adjust index for leading space in " {kw} "
            start_in_orig = first_poi_idx  # since we added a space at start
            poi_val = seg[start_in_orig:].strip()
            seg_left = seg[:start_in_orig].strip(" ,-–/:")
            if seg_left:
                remaining_segments.append(seg_left)
            pois.append(poi_val)
        else:
            remaining_segments.append(seg)
            
    if debug:
        print(f"[DEBUG] After Module 4 (POI): {pois}, Remaining: {remaining_segments}")

    # ==========================================
    # MODULE 5: Bóc tách Chi tiết Cấp 4 (Level 4 Parsing)
    # ==========================================
    
    l4_kws = ["duong", "pho", "thon", "xom", "khu", "to", "khom", "ap", "khoi"]
    l4_regex = re.compile(r"\b(" + "|".join(l4_kws) + r")\b", re.I)
    
    level4s = []
    leftover_segments = []
    
    for seg in remaining_segments:
        seg_norm = strip_diacritics(seg.lower())
        m = l4_regex.search(seg_norm)
        if m:
            l4_start = m.start()
            l4_val = seg[l4_start:].strip()
            level4s.append(l4_val)
            seg_left = seg[:l4_start].strip(" ,-–/:")
            if seg_left:
                leftover_segments.append(seg_left)
        else:
            leftover_segments.append(seg)
            
    # Bắt phần dư khuyết tiền tố:
    # Nếu leftover segment không phải tên người, ta đẩy vào Cấp 4
    for seg in leftover_segments:
        if not seg:
            continue
        seg_cleaned = PRONOUN_NAME_RE.sub("", seg).strip()
        seg_cleaned = seg_cleaned.strip(" ,.-–\n/:")
        if not seg_cleaned:
            continue
            
        # Check if preceded by a house number or digit sequence in the raw address
        is_protected = False
        raw_norm = strip_diacritics(raw.lower())
        seg_norm = strip_diacritics(seg_cleaned.lower())
        idx = raw_norm.find(seg_norm)
        if idx != -1:
            before_text = raw_norm[:idx]
            parts = [p.strip() for p in re.split(r"[,;()]+", before_text) if p.strip()]
            last_part = parts[-1] if parts else ""
            if (re.search(r"\d+", last_part) or 
                re.search(r"\b(?:so|sn|nha|so nha|dc|dia chi|ship|giao)\b$", last_part)):
                is_protected = True
                
        # Check if segment is a standalone name on its own line in the raw input
        is_standalone_name = False
        if not is_protected:
            seg_lower = seg_cleaned.strip().lower()
            if seg_lower in _raw_lines:
                seg_words = seg_cleaned.split()
                if 2 <= len(seg_words) <= 3:
                    if all(re.match(r'^[A-ZÀ-Ỹ][a-zà-ỹ]*$', w) for w in seg_words):
                        if not any(any(c.isdigit() for c in w) for w in seg_words):
                            norms_w = [strip_diacritics(w.lower()) for w in seg_words]
                            l4_kw = {
                                "duong", "pho", "ngo", "ngach", "hem", "kiet", "ap", "thon",
                                "xom", "to", "khu", "kp", "tdp", "doi", "ban", "buon", "lang",
                                "ql", "tl", "hl", "dt", "km", "khom", "khoi", "so", "sn",
                            }
                            if not any(w in l4_kw for w in norms_w):
                                is_standalone_name = True
        
        if not is_protected and (is_standalone_name or is_customer_name(seg_cleaned) or is_admin_or_noise_segment(seg_cleaned)):
            # Discard (Module 6)
            if debug:
                print(f"[DEBUG] Module 6 discarded: {seg_cleaned}")
            continue
        else:
            # Add to Cấp 4 details
            level4s.append(seg_cleaned)
            
    if debug:
        print(f"[DEBUG] After Module 5 (L4): {level4s}")

    # ==========================================
    # MODULE 6: Dọn rác (Garbage Collection)
    # ==========================================
    
    clean_pois = [_titlecase_vn(p) for p in pois]
    clean_level4s = [_titlecase_vn(l) for l in level4s]
    
    # Filter empty / duplicates in list
    clean_pois = list(dict.fromkeys([p for p in clean_pois if p]))
    clean_level4s = list(dict.fromkeys([l for l in clean_level4s if l]))
    
    return " | ".join(clean_pois), " | ".join(clean_level4s)


# ---------------------------------------------------------------------------
# Rule engine v2: chỉ bóc POI + đơn vị cấp 4 từ cột "Địa chỉ".
#
# Hàm legacy phía trên từng đẩy phần dư (đường/số nhà) vào "cấp 4", làm tỷ lệ
# có chi tiết cao giả tạo. Bộ luật mới từ file Markdown yêu cầu hẹp hơn: các
# cột Phường/Xã, Quận/Huyện, Tỉnh/TP là source of truth; raw address chỉ dùng
# để lấy POI và cấp 4 có keyword rõ ràng.

NOTE_KEYWORDS_V2 = [
    "khong dung hang", "ko dung hang", "tra lai", "hoan hang", "doi hang",
    "mien ship", "free ship", "freeship", "ship nhe", "ship nha",
    "goi truoc", "dung goi", "gio hanh chinh", "gio hanh chanh",
    "thu ho", "cod", "khach tra", "chi tra", "anh tra", "em tra",
    "sai mau", "sai size", "sai mau", "khong giong hinh", "ko giong hinh",
    "cho xem hang", "dong kiem", "hang de vo", "huy don", "khong lay",
    "ko lay", "trua", "chieu", "toi", "tranh thu", "khi chua nhap",
]

POI_PHRASES_V2 = [
    "truong", "mam non", "mn", "tieu hoc", "thcs", "thpt", "dai hoc",
    "cao dang", "cd", "ubnd", "uy ban", "uy ban nhan dan", "cong an",
    "toa an", "vien kiem sat", "chi cuc", "kho bac", "benh vien",
    "bv", "phong kham", "pk", "tram y te", "ttyt", "trung tam y te",
    "nha thuoc", "trung tam", "ngan hang", "bank", "atm", "vietcombank", "bidv",
    "agribank", "techcombank", "mbbank", "vpbank", "ban viet", "toa nha",
    "building", "chung cu", "apartment", "tower", "plaza", "center",
    "centre", "cong ty", "cong ty", "cty", "cong ty", "doanh nghiep",
    "xi nghiep", "nha may", "khu cong nghiep", "kcn", "cum cong nghiep",
    "cum cn", "khu che xuat", "kcx", "cho", "sieu thi", "buu dien",
    "nha van hoa", "san van dong", "ben xe", "ga", "cang", "chua",
    "nha tho", "den", "mieu", "tu vien", "khach san", "nha nghi",
    "nha sach", "phong giao dich", "bao hiem", "bhxh", "cong vien",
]

STREET_START_RE_V2 = re.compile(
    r"\b(?:"
    r"duong|pho|ngo|ngach|hem|kiet|quoc\s*lo|ql|tinh\s*lo|tl|"
    r"huong\s*lo|hl|duong\s*tinh|dt|cao\s*toc|km"
    r")\b",
    re.I,
)

HOUSE_TOKEN_V2 = r"(?:[A-ZĐ]{0,4}\d+[A-ZĐ0-9]*|\d+[A-ZĐ]+[0-9]*|[A-ZĐ]+\d+[A-ZĐ0-9]*)"

HOUSE_LIKE_RE_V2 = re.compile(
    r"(?<!\w)" + HOUSE_TOKEN_V2 +
    r"(?:\s*[\/\-\.]\s*" + HOUSE_TOKEN_V2 + r")*" +
    r"(?!\w)",
    re.I,
)

MONEY_WEIGHT_RE_V2 = re.compile(r"^\d+[.,]?\d*(k|kg|g|gr|tr|ky|d|vnd)\w*$")

LEVEL4_MAIN_TYPES_V2 = {
    "thon", "ap", "khu pho", "khoi", "ban", "buon", "soc", "lang",
    "khom", "khu", "tieu khu", "cum dan cu",
}

LEVEL4_UNIT_MAP_V2 = {
    "thon": "Thôn",
    "xom": "Xóm",
    "ap": "Ấp",
    "khu pho": "Khu phố",
    "khu": "Khu",
    "khoi": "Khối",
    "to dan pho": "Tổ dân phố",
    "tdp": "Tổ dân phố",
    "to": "Tổ",
    "buon": "Buôn",
    "ban": "Bản",
    "lang": "Làng",
    "soc": "Sóc",
    "doi": "Đội",
    "tieu khu": "Tiểu khu",
    "cum dan cu": "Cụm dân cư",
    "khom": "Khóm",
}


def _norm_match_v2(s):
    s = unicodedata.normalize("NFC", str(s or ""))
    s = strip_diacritics(s.lower())
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _compact_match_v2(s):
    return re.sub(r"[^a-z0-9]+", "", strip_diacritics(str(s or "").lower()))


def _clean_spaces_v2(s):
    s = unicodedata.normalize("NFC", str(s or ""))
    s = re.sub(r"[\r\n\t]+", ", ", s)
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"\s*,\s*", ", ", s)
    s = re.sub(r",{2,}", ",", s)
    return s.strip(" ,.-")


def _pretty_piece_v2(s):
    s = _clean_spaces_v2(s)
    s = re.sub(r"\bCông\s*Tỷ\b", "Công Ty", s, flags=re.I)
    s = _titlecase_vn(s)
    fixes = [
        (r"\bUbnd\b", "UBND"), (r"\bThcs\b", "THCS"),
        (r"\bThpt\b", "THPT"), (r"\bTnhh\b", "TNHH"),
        (r"\bBhxh\b", "BHXH"), (r"\bKcn\b", "KCN"),
        (r"\bKcx\b", "KCX"), (r"\bKdc\b", "KDC"),
        (r"\bKdt\b", "KĐT"), (r"\bTtyt\b", "TTYT"),
        (r"\bVg\b", "VG"), (r"\bCp\b", "CP"),
        (r"\bCđ\b", "CĐ"), (r"\bMn\b", "Mầm non"),
        (r"\bBv\b", "Bệnh viện"), (r"\bCty\b", "Công ty"),
        (r"\bToà\b", "Tòa"), (r"\bVpbank\b", "VPBank"),
        (r"\bCntt\b", "CNTT"), (r"\bTt\b", "TT"),
    ]
    for pat, repl in fixes:
        s = re.sub(pat, repl, s)
    phrase_fixes = [
        (r"\bMầm Non\b", "Mầm non"), (r"\bTiểu Học\b", "Tiểu học"),
        (r"\bTrung Học\b", "Trung học"), (r"\bCông Ty\b", "Công ty"),
        (r"\bBệnh Viện\b", "Bệnh viện"), (r"\bNgân Hàng\b", "Ngân hàng"),
        (r"\bKhu Phố\b", "Khu phố"), (r"\bTổ Dân Phố\b", "Tổ dân phố"),
        (r"\bTòa Nhà\b", "Tòa nhà"), (r"\bNhà Thuốc\b", "Nhà thuốc"),
        (r"\bPhòng Khám\b", "Phòng khám"), (r"\bTrạm Y Tế\b", "Trạm y tế"),
        (r"\bTrung Tâm\b", "Trung tâm"), (r"\bNhà Văn Hóa\b", "Nhà văn hóa"),
        (r"\bKhu Công Nghiệp\b", "Khu công nghiệp"),
        (r"\bCụm Công Nghiệp\b", "Cụm công nghiệp"),
    ]
    for pat, repl in phrase_fixes:
        s = re.sub(pat, repl, s)
    s = re.sub(
        r"\b(?=[A-Za-z0-9]*\d)([A-Za-z0-9]{1,6})\b",
        lambda m: m.group(1).upper(),
        s,
    )
    return s.strip(" ,.-")


def _normalize_abbrev_v2(s):
    s = unicodedata.normalize("NFC", str(s or ""))
    s = re.sub(r"\bkhu\s*dc\b", "Khu Dân Cư", s, flags=re.I)
    s = re.sub(r"\b[đd]\s*[\.,/]\s*c\b", "Địa chỉ", s, flags=re.I)
    s = re.sub(r"\bđ/c\b|\bd/c\b|\bđc\b|\bdc\b", "Địa chỉ", s, flags=re.I)
    s = re.sub(r"\bsnha\b|\bsn\b", "Số nhà", s, flags=re.I)
    s = re.sub(r"\btthcs\b", "Trường THCS", s, flags=re.I)
    s = re.sub(r"\btrường\s+mn\b", "Trường Mầm Non", s, flags=re.I)
    s = re.sub(r"\bmn\s+(?=[A-ZÀ-Ỹ])", "Mầm Non ", s, flags=re.I)
    s = re.sub(r"\bbv\b", "Bệnh Viện", s, flags=re.I)
    s = re.sub(r"\bub\s*nd\b|\bubnd\b", "UBND", s, flags=re.I)
    s = re.sub(r"\bcty\b|\bc\.ty\b|\bcông\s*tỷ\b", "Công Ty", s, flags=re.I)
    s = re.sub(r"\bkcn\b|\bkhu\s*cn\b", "Khu Công Nghiệp", s, flags=re.I)
    s = re.sub(r"\bcụm\s*cn\b|\bcum\s*cn\b", "Cụm Công Nghiệp", s, flags=re.I)
    s = re.sub(r"\bkcx\b", "Khu Chế Xuất", s, flags=re.I)
    s = re.sub(r"\bkdc\b", "Khu Dân Cư", s, flags=re.I)
    s = re.sub(r"\bkdt\b", "Khu Đô Thị", s, flags=re.I)
    s = re.sub(r"\bcc\b", "Chung Cư", s, flags=re.I)
    s = re.sub(r"\bkp\.?\s*(\d+[a-zA-Z]?)\b", r"Khu Phố \1", s, flags=re.I)
    s = re.sub(r"\bk\.p\.?\s*(\d+[a-zA-Z]?)\b", r"Khu Phố \1", s, flags=re.I)
    s = re.sub(r"\btdp\.?\s*(\d+[a-zA-Z]?)\b", r"Tổ Dân Phố \1", s, flags=re.I)
    s = re.sub(r"\b(khu\s*phố|khu\s*pho)\s*(\d)", r"\1 \2", s, flags=re.I)
    s = re.sub(
        r"\b(ấp|ap|thôn|thon|xóm|xom|tổ|to|đội|doi|khóm|khom|khối|khoi|"
        r"bản|ban|buôn|buon|sóc|soc)\s*(\d+[a-zA-Z]?)\b",
        r"\1 \2", s, flags=re.I)
    return s


def _strip_admin_prefix_v2(name):
    n = _norm_match_v2(name)
    for p in (
        "phuong", "xa", "thi tran", "quan", "huyen", "thanh pho",
        "thi xa", "tinh", "tp", "tx", "tt",
    ):
        if n.startswith(p + " "):
            return n[len(p):].strip()
    return n


def _build_admin_aliases_v2(admin_names):
    aliases = set()
    for nm in admin_names or []:
        if not nm:
            continue
        norm = _norm_match_v2(nm)
        core = _strip_admin_prefix_v2(nm)
        for val in (norm, core):
            if val:
                aliases.add(val)
                aliases.add(val.replace(" ", ""))
        initials = "".join(part[0] for part in core.split() if part)
        if len(initials) >= 2:
            aliases.add(initials)
        if norm.startswith("phuong ") and core:
            aliases.update({f"p {core}", f"p.{core}", f"p{core.replace(' ', '')}"})
        if norm.startswith("xa ") and core:
            aliases.update({f"x {core}", f"x.{core}", f"x{core.replace(' ', '')}"})
        if norm.startswith("quan ") and core:
            aliases.update({f"q {core}", f"q.{core}", f"q{core.replace(' ', '')}"})
        if norm.startswith("huyen ") and core:
            aliases.update({f"h {core}", f"h.{core}", f"h{core.replace(' ', '')}"})
        if norm.startswith("tinh ") and core:
            aliases.add(f"t {core}")
            aliases.add(f"t{core.replace(' ', '')}")

        m = re.search(r"phuong\s+(\d{1,2}[a-z]?)$", norm)
        if m:
            n = m.group(1)
            aliases.update({f"p{n}", f"p {n}", f"p.{n}", f"phuong {n}"})
        m = re.search(r"quan\s+(\d{1,2}[a-z]?)$", norm)
        if m:
            n = m.group(1)
            aliases.update({f"q{n}", f"q {n}", f"q.{n}", f"quan {n}"})

    joined = " ".join(aliases)
    province_abbrs = {
        "ho chi minh": ["hcm", "tphcm", "tp hcm", "sg", "sai gon"],
        "ha noi": ["hn"],
        "hai phong": ["hp"],
        "da nang": ["dn"],
        "ba ria vung tau": ["brvt", "vung tau", "vt", "tp vt", "tp-vt"],
        "binh duong": ["bd"],
        "dong nai": ["dnai"],
        "quang ninh": ["qn"],
        "dak lak": ["daklak", "dac lak", "daclak"],
    }
    for full, abbrs in province_abbrs.items():
        if full in aliases or full in joined:
            aliases.update(abbrs)
            aliases.update(a.replace(" ", "") for a in abbrs)
    return aliases


def _build_admin_tail_aliases_v2(admin_names):
    aliases = set()
    for nm in admin_names or []:
        if not nm:
            continue
        norm = _norm_match_v2(nm)
        core = _strip_admin_prefix_v2(nm)
        if not norm:
            continue
        is_tail_level = norm.startswith((
            "quan ", "huyen ", "thanh pho ", "thi xa ", "tinh ", "tp ", "tx ",
        ))
        if not is_tail_level:
            continue
        for val in (norm, core):
            if val:
                aliases.add(val)
                aliases.add(val.replace(" ", ""))
        initials = "".join(part[0] for part in core.split() if part)
        if len(initials) >= 2:
            aliases.add(initials)
        if norm.startswith("quan ") and core:
            aliases.update({f"q {core}", f"q.{core}", f"q{core.replace(' ', '')}"})
        if norm.startswith("huyen ") and core:
            aliases.update({f"h {core}", f"h.{core}", f"h{core.replace(' ', '')}"})

    joined = " ".join(aliases)
    province_abbrs = {
        "ho chi minh": ["hcm", "tphcm", "tp hcm", "sg", "sai gon"],
        "ha noi": ["hn"],
        "hai phong": ["hp"],
        "da nang": ["dn"],
        "ba ria vung tau": ["brvt", "vung tau", "vt", "tp vt", "tp-vt"],
        "binh duong": ["bd"],
        "dong nai": ["dnai"],
        "quang ninh": ["qn"],
        "dak lak": ["daklak", "dac lak", "daclak"],
        # Buôn Ma Thuột có nhiều cách viết phổ biến trong thực tế
        "buon ma thuot": ["buon me thuot", "buon me thuột", "bmt", "buonmathuot", "buonmethuot"],
    }
    for full, abbrs in province_abbrs.items():
        if full in aliases or full in joined:
            aliases.update(abbrs)
            aliases.update(a.replace(" ", "") for a in abbrs)
    return aliases


def _trim_admin_tail_aliases_v2(text, tail_aliases):
    out = _clean_spaces_v2(text)
    if not out or not tail_aliases:
        return out
    original = out
    removed_count = 0
    while True:
        toks, norms = _word_spans_v2(out)
        if not toks:
            return ""
        removed = False
        for L in range(min(5, len(toks)), 0, -1):
            phrase = " ".join(norms[-L:])
            if phrase in tail_aliases or phrase.replace(" ", "") in tail_aliases:
                out = out[:toks[-L][1]].rstrip(" ,.-–/:")
                removed = True
                removed_count += 1
                break
        if not removed:
            if removed_count < 2:
                return _clean_spaces_v2(original)
            return _clean_spaces_v2(out)


def _starts_with_admin_alias_v2(seg, aliases):
    n = _norm_match_v2(seg)
    c = n.replace(" ", "")
    if n in aliases or c in aliases:
        return True
    admin_prefixes = (
        "phuong", "xa", "thi tran", "quan", "huyen", "thanh pho", "tinh",
        "tp", "tt", "tx", "p", "q", "h", "x", "t",
    )
    return any(n == p or n.startswith(p + " ") for p in admin_prefixes)


def _strip_intro_v2(s):
    s = re.sub(r"^[^\wÀ-ỹĐđ]+", " ", str(s or "")).strip()
    return re.sub(
        r"^\s*(?:ok[!.]?\s*)?(?:em\s+g(?:ửi|ởi|ui|oi)\s+về\s+địa\s+chỉ\s+người\s+nhận|"
        r"em\s+g(?:ửi|ởi|ui|oi)\s+ve\s+dia\s+chi\s+nguoi\s+nhan|em\s+g(?:ửi|ởi|ui|oi)\s+về|"
        r"em\s+g(?:ửi|ởi|ui|oi)\s+ve|hàng\s+g(?:ửi|ởi)\s+về\s+địa\s+chỉ\s+này\s+giúp\s+mình\s+nhé|"
        r"hang\s+g(?:ui|oi)\s+ve\s+dia\s+chi\s+nay\s+giup\s+minh\s+nhe|"
        r"địa\s+chỉ\s+giao\s+hàng|dia\s+chi\s+giao\s+hang|"
        r"viet\s+ship\s+mình\s+địa\s+chỉ|"
        r"viet\s+ship\s+minh\s+dia\s+chi|ship\s+mình\s+đến\s+địa\s+chỉ|"
        r"ship\s+minh\s+den\s+dia\s+chi|ship\s+đến|ship\s+den|"
        r"giao\s+đến|giao\s+den|gửi\s+chị\s+đi\s+nhé\s+về|"
        r"gui\s+chi\s+di\s+nhe\s+ve|gửi\s+chị\s+đi\s+nhé|"
        r"gui\s+chi\s+di\s+nhe|g(?:ửi|ởi)\s+về|g(?:ui|oi)\s+ve|gửi\s+chị|gui\s+chi|"
        r"địa\s+chỉ|dia\s+chi|đc|dc|đ/c|d/c|người\s+nhận|nguoi\s+nhan|"
        r"khách\s*hàng|khach\s*hang|tên\s*kh|ten\s*kh)\s*[:：\-\.]*\s*",
        "",
        s,
        flags=re.I,
    ).strip()


def _is_note_segment_v2(seg):
    n = _norm_match_v2(seg)
    if not n:
        return True
    padded = f" {n} "
    for kw in NOTE_KEYWORDS_V2:
        if re.search(r"(?<!\w)" + re.escape(kw) + r"(?!\w)", padded):
            return True
    if re.search(r"\bt[2-7]\b|\bchu nhat\b|\bthu\s*[2-7]\b", n):
        return True
    if re.search(r"\b(?:cao|nang)\s*\d", n):
        return True
    if re.fullmatch(r"(ok|nhe|nha|a|e|em|chi|anh|vietnam|viet nam)", n):
        return True
    if re.fullmatch(r"(size|sz|mau|kg|cod|thu ho|phi ship|cuoc).*", n):
        return True
    return False


def _split_segments_v2(s):
    s = re.sub(r"[()]+", ",", s)
    s = re.sub(r"(?<=\d)\s*[-–—]\s*(?=\d)", " __ADDR_RANGE_DASH__ ", s)
    s = re.sub(r"\s*[-–—|;]\s*", ",", s)
    s = s.replace(" __ADDR_RANGE_DASH__ ", "-")
    return [p.strip(" ,.-:/") for p in re.split(r"[,;\n\r\t|]+", s) if p.strip(" ,.-:/")]


def _find_poi_keyword_v2(seg):
    norm = _norm_match_v2(seg)
    words = norm.split()
    best = None
    raw_words = [w.strip(" ,.-–/:").lower() for w in re.findall(r"[^\W_]+", seg, re.UNICODE)]
    for i, w in enumerate(raw_words):
        if w == "sở":
            best = (i, 1, "sở")
            break
    for phrase in POI_PHRASES_V2:
        parts = phrase.split()
        L = len(parts)
        for i in range(0, len(words) - L + 1):
            if words[i:i + L] == parts:
                if best is None or i < best[0] or (i == best[0] and L > best[1]):
                    best = (i, L, phrase)
    if best is None:
        for i, w in enumerate(words[:-1]):
            if w == "toa" and re.search(r"\d", words[i + 1]):
                return i, 1, "toa"
    if best is None:
        for i, w in enumerate(words):
            if w.endswith("plaza") and len(w) > 5:
                return i, 1, "plaza"
    return best


def _word_spans_v2(s):
    toks = [(m.group(), m.start(), m.end()) for m in re.finditer(r"[^\W_]+", s, re.UNICODE)]
    norms = [_norm_match_v2(w) for w, _, _ in toks]
    return toks, norms


def _strip_poi_tail_v2(seg, admin_tail_aliases=None, admin_aliases=None):
    toks, norms = _word_spans_v2(seg)
    if not toks:
        return ""
    admin_aliases = admin_aliases or set()
    raws = [unicodedata.normalize("NFC", w.lower()).strip(" ,.-–/:") for w, _, _ in toks]
    # Position-based protection: protect_until = max word index where admin/L4
    # keywords are considered part of the POI name (not location details).
    # Words at positions >= protect_until can be cut.
    protect_until = 0  # default: no protection
    if norms[0] in {"truong", "bv"} or " ".join(norms[:2]) in {"truong hoc", "benh vien", "mam non", "tieu hoc", "trung hoc"}:
        protect_until = 8  # schools/hospitals: protect up to 8 words for long names
    elif " ".join(norms[:3]) == "tram y te":
        protect_until = 6
    elif norms[0] in {"ubnd"} or " ".join(norms[:2]) in {"uy ban"}:
        protect_until = 3  # UBND: protect first 3 words (e.g. "Ủy Ban Xã" / "UBND Xã")
    elif norms[0] in {"cho", "chua", "den"} or toks[0][0].strip(" ,.-–/:").lower() == "sở":
        protect_until = 4  # Chợ/Chùa/Đền/Sở: protect first 4 words
    cut = None
    cut_by_street = False
    for i, n in enumerate(norms):
        phrase2 = " ".join(norms[i:i + 2])
        raw = raws[i]
        raw_phrase2 = " ".join(raws[i:i + 2])
        nxt = norms[i + 1] if i + 1 < len(norms) else ""
        is_known_admin_here = any(
            " ".join(norms[i:i + L]) in admin_aliases
            or " ".join(norms[i:i + L]).replace(" ", "") in admin_aliases
            for L in range(1, min(5, len(norms) - i) + 1)
        )
        is_street = (
            raw in {"đường", "duong", "phố", "pho"}
            or (raw in {"ngõ", "ngo", "ngách", "ngach", "hẻm", "hem", "kiệt", "kiet"} and bool(re.match(r"^\d", nxt)))
        )
        is_route = phrase2 in {"quoc lo", "tinh lo", "huong lo", "dai lo", "cao toc"}
        force_admin_tail = is_known_admin_here and raw in {"xã", "xa", "phường", "phuong", "p", "x"}
        protected_tt_abbrev = raw == "tt" and i > 0 and norms[i - 1] == "cntt"
        is_admin_tail = (
            i >= 2 and not protected_tt_abbrev and (i >= protect_until or force_admin_tail) and
            (raw in {"xã", "xa", "phường", "phuong", "quận", "quan", "huyện", "huyen", "tỉnh", "tinh", "tp", "tt", "tx", "p", "q", "h", "x", "t"}
             or raw_phrase2 in {"thành phố", "thanh pho", "thị xã", "thi xa", "thị trấn", "thi tran"})
        )
        is_address_code_tail = (
            i >= 2
            and n in {"ht", "lo", "lô", "lot", "block"}
            and bool(re.match(r"^\d", nxt))
        )
        is_level4 = (
            i >= 2 and i >= protect_until and
            raw in {"ấp", "ap", "thôn", "thon", "xóm", "xom", "tổ", "to", "khu", "khối", "khoi", "buôn", "buon", "bản", "ban", "làng", "lang", "đội", "doi", "tdp"}
            # KHÔNG cắt "bản/ban" nếu đây là tên thương hiệu đứng sau từ khóa POI
            # (vd "Ngân hàng Bản Việt", "VPBank Hoà Bình")
            and not (raw in {"bản", "ban"} and norms[0] in {
                "ngan", "truong", "benh", "buu", "uy", "cho", "cong", "vpbank",
                "bidv", "agribank", "techcombank", "mbbank",
            })
            # KHÔNG cắt "buôn/buon" nếu phần sau là tên thành phố trong admin alias
            # (vd "Bến Xe Phía Bắc Buôn Mê Thuột" — không được cắt "Buôn Mê Thuột")
            and not (raw in {"buôn", "buon"} and admin_tail_aliases and any(
                " ".join(norms[i:i + L]) in admin_tail_aliases
                or " ".join(norms[i:i + L]).replace(" ", "") in admin_tail_aliases
                for L in range(1, min(5, len(norms) - i) + 1)
            ))
        )
        is_recipient_tail = (
            i >= 2 and (
                phrase2 in {"nguoi nhan", "dia chi", "khach hang", "ten kh"}
                or raw in {"chị", "chi", "anh", "cô", "co", "em"}
            )
        )
        if is_street or is_route or is_admin_tail or is_level4 or is_recipient_tail or is_address_code_tail or n in {"ql", "tl", "hl", "dt"}:
            if i == 0:
                return ""
            cut = toks[i][1]
            cut_by_street = bool(is_street or is_route or n in {"ql", "tl", "hl", "dt"})
            break
    if cut is not None:
        seg = seg[:cut]
    seg = re.sub(r"\b(?:số|so|sn)\s*\d+\w*(?:[/\-]\w+)*\b.*$", " ", seg, flags=re.I)
    if cut_by_street:
        seg = re.sub(r"\s+" + HOUSE_TOKEN_V2 + r"\s*$", " ", seg, flags=re.I)
    seg = _trim_admin_tail_aliases_v2(seg, admin_tail_aliases or set())
    return _clean_spaces_v2(seg)


def _extract_pois_v2(segments, admin_aliases, admin_tail_aliases=None):
    pois = []
    admin_tail_aliases = admin_tail_aliases or set()
    for seg in segments:
        if _is_note_segment_v2(seg) or _starts_with_admin_alias_v2(seg, admin_aliases):
            continue
        seg = _strip_intro_v2(seg)
        if not seg:
            continue
        if re.match(r"^(?:sau|phia sau|dang sau|gan|doi dien|ben canh)\s+", _norm_match_v2(seg)):
            continue
        found = _find_poi_keyword_v2(seg)
        if not found:
            continue
        kw_idx, _, kw_phrase = found
        toks, norms = _word_spans_v2(seg)
        if not toks:
            continue
        start_tok = max(0, kw_idx)
        before = norms[:start_tok]
        before_has_signal = any(
            w in {
                "duong", "pho", "ngo", "ngach", "hem", "kiet", "thon",
                "xom", "ap", "to", "khu", "kp", "tdp", "doi", "ban",
                "buon", "khom", "khoi", "xa", "phuong", "quan", "huyen",
                "tinh", "tp", "tt", "tx", "so", "sn",
            }
            for w in before
        )
        before_has_note_marker = any(
            w in {"dia", "chi", "giao", "hang", "nguoi", "nhan", "goi", "gui", "ve", "lo"}
            for w in before
        )
        if before_has_signal:
            continue

        raw_kw = toks[start_tok][0].strip(" ,.-–/:").lower()
        local_context = norms[max(0, start_tok - 1):start_tok + 5]
        edu_context = any(
            w in {"mn", "mam", "non", "tieu", "hoc", "thcs", "thpt", "dai", "cao", "cd"}
            for w in local_context
        )
        if kw_phrase == "truong" and raw_kw != "trường" and not edu_context:
            continue
        if kw_phrase == "truong" and start_tok > 0 and not edu_context:
            continue
        if kw_phrase == "cho" and raw_kw != "chợ":
            continue
        if kw_phrase == "den" and raw_kw != "đền":
            continue

        if before and (len(before) <= 4 or before_has_note_marker):
            if not before_has_signal:
                seg = seg[toks[start_tok][1]:]
        if re.match(r"^(?:sau|phia sau|dang sau|gan|doi dien|ben canh)\s+", _norm_match_v2(seg)):
            continue
        poi = _strip_poi_tail_v2(seg, admin_tail_aliases, admin_aliases)
        if not poi:
            continue
        if re.fullmatch(r"\d+[a-z]?", _norm_match_v2(poi)):
            continue
        norm_poi = _norm_match_v2(poi)
        if _is_note_segment_v2(poi) or norm_poi in admin_aliases:
            continue
        word_count = len(norm_poi.split())
        if word_count < 1 or word_count > 18:
            continue
        pretty = _pretty_piece_v2(poi)
        key = _norm_match_v2(pretty)
        if pois and key.startswith("cum cong nghiep "):
            continue
        if key and key not in {_norm_match_v2(x) for x in pois}:
            pois.append(pretty)
    return pois


def _level4_unit_at_v2(norms, i):
    n = norms[i]
    nxt = norms[i + 1] if i + 1 < len(norms) else ""
    nxt2 = norms[i + 2] if i + 2 < len(norms) else ""
    if n in {"thon", "xom", "ap", "khoi", "buon", "khom"}:
        return n, LEVEL4_UNIT_MAP_V2[n], 1
    if n in {"soc"}:
        return "soc", "Sóc", 1
    if n == "lang":
        return "lang", "Làng", 1
    if n == "ban":
        return "ban", "Bản", 1
    if n == "doi":
        return "doi", "Đội", 1
    if n == "tdp":
        return "tdp", "Tổ dân phố", 1
    if n == "to" and nxt == "chuc":
        return None
    if n == "to" and nxt == "dan" and nxt2 == "pho":
        return "to dan pho", "Tổ dân phố", 3
    if n == "to":
        return "to", "Tổ", 1
    if n == "kp":
        return "khu pho", "Khu phố", 1
    if n == "khu" and nxt == "pho":
        return "khu pho", "Khu phố", 2
    if n == "khu" and nxt not in {"pho", "cong", "che", "dan", "do", "cn"}:
        return "khu", "Khu", 1
    if n == "tieu" and nxt == "khu":
        return "tieu khu", "Tiểu khu", 2
    if n == "cum" and nxt == "dan" and nxt2 == "cu":
        return "cum dan cu", "Cụm dân cư", 3
    return None


def _is_level4_stop_v2(norms, idx):
    if idx >= len(norms):
        return True
    n = norms[idx]
    phrase2 = " ".join(norms[idx:idx + 2])
    phrase3 = " ".join(norms[idx:idx + 3])
    if phrase2 in {"thi tran", "thi xa", "thanh pho", "quoc lo", "tinh lo", "duong so"}:
        return True
    if phrase3 in {"to dan pho", "cum dan cu"}:
        return True
    if n in {
        "xa", "phuong", "quan", "huyen", "tinh", "tp", "tx", "tt",
        "duong", "pho", "ngo", "ngach", "hem", "kiet", "sn",
        "nha", "ql", "tl", "hl", "dt", "km", "dia", "chi", "ship",
        "giao", "sdt", "dt", "cod",
    }:
        return True
    if _level4_unit_at_v2(norms, idx):
        return True
    return False


def _extract_level4s_v2(text, admin_aliases=None):
    admin_aliases = admin_aliases or set()
    toks, norms = _word_spans_v2(text)
    raws = [unicodedata.normalize("NFC", w.lower()).strip(" ,.-–/:") for w, _, _ in toks]
    results = []
    seen = set()
    i = 0
    while i < len(toks):
        unit_info = _level4_unit_at_v2(norms, i)
        if not unit_info:
            i += 1
            continue
        unit_key, unit_display, unit_len = unit_info
        if unit_key == "ban" and raws[i] != "bản":
            i += 1
            continue
        if unit_key == "ban" and any(w in {"ngan", "hang", "bank"} for w in norms[max(0, i - 4):i]):
            i += 1
            continue
        if unit_key == "soc" and raws[i] != "sóc":
            i += 1
            continue
        if unit_key == "soc" and i > 0 and norms[i - 1] in {"cham"}:
            i += 1
            continue
        if unit_key == "lang" and raws[i] != "làng":
            i += 1
            continue
        if unit_key == "buon" and " ".join(norms[i:i + 3]) in {"buon me thuot", "buon ma thuot"}:
            i += 1
            continue
        is_admin_phrase = False
        for L in range(1, min(5, len(norms) - i) + 1):
            phrase = " ".join(norms[i:i + L])
            if phrase in admin_aliases or phrase.replace(" ", "") in admin_aliases:
                is_admin_phrase = True
                break
        if is_admin_phrase:
            i += max(unit_len, 1)
            continue
        start = i + unit_len
        got = []
        j = start
        while j < len(toks) and len(got) < 5:
            if _is_level4_stop_v2(norms, j):
                break
            admin_tail_len = 0
            for L in range(1, min(5, len(norms) - j) + 1):
                phrase = " ".join(norms[j:j + L])
                if phrase in admin_aliases or phrase.replace(" ", "") in admin_aliases:
                    admin_tail_len = L
                    break
            if admin_tail_len:
                next_idx = j + admin_tail_len
                if next_idx < len(toks) and re.match(r"^\d+[a-z]?$", norms[next_idx], re.I):
                    got.extend(range(j, next_idx + 1))
                    break
                break
            if MONEY_WEIGHT_RE_V2.match(norms[j]):
                break
            if norms[j] in {"cu", "moi"}:
                break
            got.append(j)
            sep = text[toks[j][2]:toks[j + 1][1]] if j + 1 < len(toks) else ""
            if re.search(r"[,;.\-–—\n\r|()]", sep):
                break
            j += 1
        if got:
            # Với cấp 4 dạng số, chỉ lấy token số/mã ngay sau đơn vị để tránh
            # nuốt tên đường phía sau: "Ấp 6 Trần Văn Giàu" -> "Ấp 6".
            if re.match(r"^\d+[a-z]?$", norms[got[0]], re.I):
                got = [got[0]]
            raw_name = text[toks[got[0]][1]:toks[got[-1]][2]]
            val = f"{unit_display} {_pretty_piece_v2(raw_name)}".strip()
            val = re.sub(r"\s+", " ", val).strip(" ,.-")
            key = _norm_match_v2(val)
            if key and key not in seen:
                seen.add(key)
                results.append({"value": val, "type": unit_key, "pos": toks[i][1]})
            i = got[-1] + 1
            continue
        i += max(unit_len, 1)
    return results


def _street_removed_parts_v2(text):
    parts = []
    for seg in _split_segments_v2(text):
        n = _norm_match_v2(seg)
        if not n or _is_note_segment_v2(seg):
            continue
        if re.fullmatch(r"\d+[a-z]?", n):
            continue
        street_match = _first_street_keyword_match_v2(seg)
        if street_match:
            found_poi = _find_poi_keyword_v2(seg)
            toks, _norms = _word_spans_v2(seg)
            if found_poi and toks and toks[found_poi[0]][1] < street_match[0]:
                continue
            parts.append(_pretty_piece_v2(_street_noise_only_v2(seg)))
            continue
        if _find_poi_keyword_v2(seg):
            continue
        if HOUSE_LIKE_RE_V2.search(seg) and not _extract_level4s_v2(seg):
            parts.append(_pretty_piece_v2(_street_noise_only_v2(seg)))
    return list(dict.fromkeys([p for p in parts if p]))


def _street_keyword_matches_v2(seg):
    toks, norms = _word_spans_v2(seg)
    raws = [unicodedata.normalize("NFC", w.lower()).strip(" ,.-–/:") for w, _, _ in toks]
    for i, n in enumerate(norms):
        prev = norms[i - 1] if i > 0 else ""
        phrase2 = " ".join(norms[i:i + 2])
        raw = raws[i]
        raw_phrase2 = " ".join(raws[i:i + 2])
        if n == "duong" and raw in {"đường", "duong"}:
            yield toks[i][1], toks[i][2]
            continue
        if n in {"ngo", "ngach", "hem", "kiet"} and raw in {"ngõ", "ngo", "ngách", "ngach", "hẻm", "hem", "kiệt", "kiet"}:
            nxt = norms[i + 1] if i + 1 < len(norms) else ""
            if i == 0 or re.match(r"^\d", nxt):
                yield toks[i][1], toks[i][2]
            continue
        if n == "pho" and raw in {"phố", "pho"}:
            if prev in {"khu", "thanh"}:
                continue
            yield toks[i][1], toks[i][2]
            continue
        if (
            phrase2 in {"quoc lo", "tinh lo", "huong lo", "dai lo", "cao toc"}
            and raw_phrase2 in {"quốc lộ", "quoc lo", "tỉnh lộ", "tinh lo", "hương lộ", "huong lo", "đại lộ", "dai lo", "cao tốc", "cao toc"}
        ):
            yield toks[i][1], toks[min(i + 1, len(toks) - 1)][2]
            continue
        if n in {"dt", "ql", "tl", "hl"} or re.fullmatch(r"(?:dt|ql|tl|hl)\d+[a-z]?", n):
            yield toks[i][1], toks[i][2]


def _first_street_keyword_match_v2(seg):
    return next(_street_keyword_matches_v2(seg), None)


def _has_street_keyword_v2(seg):
    return _first_street_keyword_match_v2(seg) is not None


def _street_noise_only_v2(seg):
    toks, norms = _word_spans_v2(seg)
    if not toks:
        return seg
    raws = [unicodedata.normalize("NFC", w.lower()).strip(" ,.-–/:") for w, _, _ in toks]
    street_kw_positions = {
        i for i, r in enumerate(raws)
        if r in {"đường", "duong", "phố", "pho"}
    }
    cut = None
    for i, n in enumerate(norms):
        phrase2 = " ".join(norms[i:i + 2])
        raw = raws[i]
        raw_phrase2 = " ".join(raws[i:i + 2])
        protected_by_street = bool(
            street_kw_positions and max(street_kw_positions) >= i - 3
        )
        if i > 0 and (
            _level4_unit_at_v2(norms, i)
            or (
                raw in {"xã", "xa", "phường", "phuong", "quận", "quan", "huyện", "huyen", "tỉnh", "tinh", "tp", "tt", "tx"}
                and not protected_by_street
            )
            or raw_phrase2 in {"thành phố", "thanh pho", "thị xã", "thi xa", "thị trấn", "thi tran"}
        ):
            cut = toks[i][1]
            break
    return seg[:cut].strip(" ,.-–/:") if cut else seg


def _strip_leading_address_junk_v2(seg):
    seg = _strip_intro_v2(seg or "")
    seg = re.sub(r"^\s*(?:[đd]\s*,\s*)?c\s+", " ", seg, flags=re.I)
    seg = re.sub(r"^\s*(?:ok[!.]?\s*)+", " ", seg, flags=re.I)
    return re.sub(r"\s+", " ", seg).strip(" ,.-–/:")


def _cut_detail_tail_v2(text, admin_aliases):
    toks, norms = _word_spans_v2(text)
    if not toks:
        return ""
    raws = [unicodedata.normalize("NFC", w.lower()).strip(" ,.-–/:") for w, _, _ in toks]
    starts_with_named_street = raws[0] in {"đường", "duong", "phố", "pho"}
    # Kiểm tra xem có street keyword nào trong chuỗi không (để bảo vệ tên đường
    # dạng "Ngõ 43 Đường Ao Quan" — sau khi bỏ ngõ thì "Quan" là phần tên đường)
    street_kw_positions = {
        i for i, r in enumerate(raws)
        if r in {"đường", "duong", "phố", "pho"}
    }
    cut = None
    for i, n in enumerate(norms):
        phrase2_raw = " ".join(raws[i:i + 2])
        unit_info = _level4_unit_at_v2(norms, i)
        invalid_level4_word = (
            unit_info
            and (
                (unit_info[0] == "ban" and raws[i] != "bản")
                or (unit_info[0] == "soc" and raws[i] != "sóc")
                or (unit_info[0] == "lang" and raws[i] != "làng")
                or (unit_info[0] == "buon" and " ".join(norms[i:i + 3]) in {"buon me thuot", "buon ma thuot"})
            )
        )
        if unit_info and not invalid_level4_word:
            cut = toks[i][1]
            break
        if raws[i] in {"xã", "xa", "phường", "phuong", "quận", "quan", "huyện", "huyen", "tỉnh", "tinh", "tp", "tt", "tx", "p", "q", "h", "x", "t"}:
            # Chỉ cắt khi token tiếp theo rõ ràng là tên admin, KHÔNG cắt khi
            # đây là từ cuối segment tên đường (vd "Đường Ao Quan" — "Quan" là
            # phần tên, không phải tiền tố Quận).
            # Bảo vệ cả trường hợp "Ngõ 43 Đường Ao Quan" — "Quan" nằm sau "Đường"
            protected_by_street = bool(
                street_kw_positions and max(street_kw_positions) >= i - 3
            )
            nxt_norm = norms[i + 1] if i + 1 < len(norms) else ""
            if not nxt_norm:
                # Token cuối — chỉ cắt nếu không nằm sau street keyword và không phải tên đường
                if not starts_with_named_street and not protected_by_street:
                    cut = toks[i][1]
                    break
            else:
                nxt_is_admin = (
                    nxt_norm in admin_aliases
                    or nxt_norm.replace(" ", "") in admin_aliases
                    or any(
                        " ".join(norms[i + 1:i + 1 + L]) in admin_aliases
                        or " ".join(norms[i + 1:i + 1 + L]).replace(" ", "") in admin_aliases
                        for L in range(1, min(5, len(norms) - i))
                    )
                )
                if nxt_is_admin and not protected_by_street:
                    cut = toks[i][1]
                    break
        if phrase2_raw in {"thành phố", "thị xã", "thị trấn"}:
            cut = toks[i][1]
            break
        is_admin_alias = False
        for L in range(1, min(5, len(norms) - i) + 1):
            phrase = " ".join(norms[i:i + L])
            if phrase in admin_aliases or phrase.replace(" ", "") in admin_aliases:
                is_admin_alias = True
                break
        if is_admin_alias:
            if starts_with_named_street and i == 1:
                pass
            else:
                cut = toks[i][1]
                break
            if cut is not None:
                break
    text = text[:cut] if cut is not None else text
    text = re.sub(r"\b(?:cũ|cu|mới|moi|nay|thuộc|thuoc)\b.*$", " ", text, flags=re.I)
    return _clean_spaces_v2(text)


def _strip_alley_noise_v2(detail):
    detail = _clean_spaces_v2(detail)
    m = re.match(
        r"^\s*(?:hẻm|hem|ngõ|ngo|ngách|ngach|kiệt|kiet)\s*"
        r"\d+[A-Za-zĐđ]?(?:\s*[/\-]\s*\d+[A-Za-zĐđ]?)*\s*(?P<rest>.*)$",
        detail,
        flags=re.I,
    )
    if not m:
        return detail
    rest = m.group("rest").strip(" ,.-–/:")
    return rest


def _strip_leading_house_before_street_v2(detail):
    return re.sub(
        r"^\s*\d+[A-Za-zĐđ]?\s+(?=(?:đường|duong|phố|pho|"
        r"quốc\s*lộ|quoc\s*lo|tỉnh\s*lộ|tinh\s*lo|đt|dt|ql|tl|hl)\b)",
        "",
        detail,
        flags=re.I,
    ).strip(" ,.-–/:")


def _is_alley_only_v2(detail):
    raw = strip_diacritics(unicodedata.normalize("NFC", str(detail or "")).lower())
    house_token = r"(?:[a-z]{0,4}\d+[a-z0-9]*|\d+[a-z]+[0-9]*|[a-z]+\d+[a-z0-9]*)"
    raw = re.sub(
        r"^\s*(?:so\s*nha|snha|so|sn|nha)?\s*"
        + house_token
        + r"(?:\s*[/\-\.]\s*" + house_token + r")*"
        + r"(?:\s+[a-z])?\s*",
        "",
        raw,
        flags=re.I,
    )
    raw = re.sub(r"\s+", " ", raw).strip(" ,.-–/:")
    return bool(re.fullmatch(
        r"(?:ngo|ngach|hem|kiet)\s*\d+[a-z]?(?:\s*[/\-\.]\s*\d+[a-z]?)*",
        raw,
        flags=re.I,
    ))


def _looks_like_street_v2(detail):
    n = _norm_match_v2(detail)
    if not n:
        return False
    if _is_alley_only_v2(detail):
        return False
    if re.fullmatch(r"(ql|tl|hl|dt)\s*\d+[a-z]?", n):
        return True
    if re.fullmatch(r"(?:\d+[a-z]?|[a-z])", n):
        return False
    if n.startswith(("khu do thi", "khu cong nghiep", "khu che xuat", "khu dan cu")):
        return False
    if _level4_unit_at_v2(n.split(), 0):
        return False
    if n in {"cu", "moi"}:
        return False
    if any(n.startswith(p + " ") or n == p for p in (
        "xa", "phuong", "quan", "huyen", "tinh", "tp", "tt", "tx",
        "thanh pho", "thi xa", "thi tran",
    )):
        return False
    if _find_poi_keyword_v2(detail):
        return False
    words = n.split()
    return 1 <= len(words) <= 7


def _extract_streets_v2(segments, admin_aliases):
    streets = []
    seen = set()
    house_prefix = (
        r"(?:số\s*nhà|so\s*nha|số|so|sn|nhà|nha)?\s*"
        r"(?:" + HOUSE_TOKEN_V2 +
        r"(?:\s*[/\-\.]\s*" + HOUSE_TOKEN_V2 + r")*"
        r"(?:\s+[A-Za-zĐđ](?=\s|[,.\-–/]|$))?)"
    )
    patterns = [
        re.compile(r"^\s*" + house_prefix + r"\s+(?P<detail>.+)$", re.I),
        re.compile(r"\b(?:số\s*nhà|so\s*nha|số|so|sn|nhà|nha)\s*"
                   r"(?:\d+[A-Za-zĐđ]?(?:\s*[/\-]\s*[A-Za-zĐđ]?\d+[A-Za-zĐđ]?)*)"
                   r"\s+(?P<detail>.+)$", re.I),
    ]
    for idx, seg in enumerate(segments):
        if not seg or _is_note_segment_v2(seg) or _starts_with_admin_alias_v2(seg, admin_aliases):
            continue
        seg = _strip_leading_address_junk_v2(seg)
        if not seg:
            continue
        detail = ""
        if _is_house_only_v2(seg) and idx + 1 < len(segments):
            next_seg = _strip_leading_address_junk_v2(segments[idx + 1])
            if next_seg and not _starts_with_admin_alias_v2(next_seg, admin_aliases):
                detail = next_seg
        for pat in patterns:
            if not detail:
                m = pat.search(seg)
                if m:
                    detail = m.group("detail")
                    break
        street_match = _first_street_keyword_match_v2(seg)
        if not detail and street_match:
            found_poi = _find_poi_keyword_v2(seg)
            toks, _norms = _word_spans_v2(seg)
            if found_poi and toks and toks[found_poi[0]][1] < street_match[0]:
                continue
            detail = seg[street_match[0]:]
        if not detail:
            continue
        detail = _cut_detail_tail_v2(detail, admin_aliases)
        if _is_alley_only_v2(detail):
            continue
        detail = _strip_alley_noise_v2(detail)
        detail = _strip_leading_house_before_street_v2(detail)
        if not _looks_like_street_v2(detail):
            continue
        pretty = _pretty_piece_v2(detail)
        key = _norm_match_v2(pretty)
        if key and key not in seen:
            seen.add(key)
            streets.append(pretty)
    return streets


def _extract_unprefixed_level4_after_known_v2(segments, admin_aliases, existing_values=None):
    existing_values = existing_values or []
    seen = {_norm_match_v2(x) for x in existing_values if x}
    out = []
    for idx in range(1, len(segments)):
        prev = _strip_leading_address_junk_v2(segments[idx - 1])
        cur = _strip_leading_address_junk_v2(segments[idx])
        if not prev or not cur:
            continue
        if not _extract_level4s_v2(prev, admin_aliases):
            continue
        if not _has_admin_ahead_v2(segments, idx, admin_aliases):
            continue
        if _starts_with_admin_alias_v2(cur, admin_aliases):
            continue
        if _has_street_keyword_v2(cur) or _is_house_only_v2(cur) or _find_poi_keyword_v2(cur):
            continue
        candidate = _cut_detail_tail_v2(cur, admin_aliases)
        if not candidate or is_customer_name(candidate):
            continue
        n = _norm_match_v2(candidate)
        if not n or n in admin_aliases or n.replace(" ", "") in admin_aliases:
            continue
        if len(n.split()) > 4:
            continue
        key = _norm_match_v2(candidate)
        if key and key not in seen:
            seen.add(key)
            out.append({"value": _pretty_piece_v2(candidate), "type": "unprefixed", "pos": 100000 + idx})
    return out


def _is_house_only_v2(seg):
    n = _norm_match_v2(seg)
    if not n:
        return False
    house_token = r"(?:[a-z]{0,4}\d+[a-z0-9]*|\d+[a-z]+[0-9]*|[a-z]+\d+[a-z0-9]*)"
    return bool(re.fullmatch(
        r"(?:so nha|snha|so|sn|nha)?\s*"
        + house_token + r"(?:\s*" + house_token + r")*",
        n,
        flags=re.I,
    ))


def _has_admin_ahead_v2(segments, start_idx, admin_aliases):
    for seg in segments[start_idx + 1:]:
        clean = _strip_leading_address_junk_v2(seg)
        n = _norm_match_v2(clean)
        if not n:
            continue
        if _starts_with_admin_alias_v2(clean, admin_aliases):
            return True
        if n in admin_aliases or n.replace(" ", "") in admin_aliases:
            return True
    return False


def _is_local_detail_candidate_v2(seg, admin_aliases):
    clean = _strip_leading_address_junk_v2(seg)
    if not clean or _is_note_segment_v2(clean):
        return ""
    if _starts_with_admin_alias_v2(clean, admin_aliases):
        return ""
    if HOUSE_LIKE_RE_V2.match(clean):
        return ""
    if re.search(r"\b(?:số|so|sn|nhà|nha)\s*\d", clean, flags=re.I) and not _has_street_keyword_v2(clean):
        return ""
    if _is_house_only_v2(clean) or _has_street_keyword_v2(clean):
        return ""
    if _find_poi_keyword_v2(clean):
        return ""
    if _extract_level4s_v2(clean, admin_aliases):
        return ""
    if is_customer_name(clean):
        return ""

    candidate = _cut_detail_tail_v2(clean, admin_aliases)
    if not candidate or is_customer_name(candidate):
        return ""
    n = _norm_match_v2(candidate)
    if not n or n in admin_aliases or n.replace(" ", "") in admin_aliases:
        return ""
    if re.fullmatch(r"\d+[a-z]?", n):
        return ""
    if len(n.split()) > 5:
        return ""
    return _pretty_piece_v2(candidate)


def _extract_local_details_v2(segments, admin_aliases, existing_values=None):
    return []


def _detect_admin_conflict_v2(text, admin_aliases):
    toks, norms = _word_spans_v2(text)
    raws = [unicodedata.normalize("NFC", w.lower()).strip(" ,.-–/:") for w, _, _ in toks]
    i = 0
    while i < len(toks):
        pfx_len = 0
        if raws[i] in {"xã", "phường", "quận", "huyện", "tỉnh", "q", "p", "x"}:
            pfx_len = 1
        elif i + 1 < len(norms) and (raws[i], raws[i + 1]) in {
            ("thành", "phố"), ("thị", "xã"), ("thị", "trấn")
        }:
            pfx_len = 2
        if not pfx_len:
            i += 1
            continue
        j = i + pfx_len
        got = []
        while j < len(toks) and len(got) < 4:
            got.append(j)
            sep = text[toks[j][2]:toks[j + 1][1]] if j + 1 < len(toks) else ""
            if re.search(r"[,;.\-–—\n\r|()]", sep):
                break
            j += 1
        if got:
            matched_known = False
            for end in range(i + pfx_len, got[-1] + 1):
                full = " ".join(norms[i:end + 1])
                core = " ".join(norms[i + pfx_len:end + 1])
                if (
                    full in admin_aliases or full.replace(" ", "") in admin_aliases
                    or core in admin_aliases or core.replace(" ", "") in admin_aliases
                ):
                    matched_known = True
                    break
            if not matched_known:
                return True
            i = got[-1] + 1
            continue
        i += max(pfx_len, 1)
    return False


def _choose_level4_primary_v2(level4_items):
    if not level4_items:
        return "", ""
    if len(level4_items) == 1:
        return level4_items[0]["value"], ""
    main = [x for x in level4_items if x["type"] in LEVEL4_MAIN_TYPES_V2]
    primary = (main[0] if main else level4_items[0])["value"]
    sub = "; ".join(x["value"] for x in level4_items if x["value"] != primary)
    return primary, sub


def _score_result_v2(poi, level4, flags):
    score = 1.0
    if not poi:
        score -= 0.05
    if not level4:
        score -= 0.05
    penalty = {
        "MULTIPLE_POI_FOUND": 0.12,
        "MULTIPLE_LEVEL4_FOUND": 0.08,
        "RAW_ADMIN_CONFLICT_WITH_COLUMNS": 0.2,
        "ONLY_STREET_LEVEL_FOUND": 0.05,
        "ONLY_ADMIN_UNITS_FOUND": 0.08,
        "LOW_CONFIDENCE": 0.25,
        "AMBIGUOUS_HOUSE_TOKEN": 0.12,
    }
    for flag in flags:
        score -= penalty.get(flag, 0.03)
    return max(0.0, min(1.0, round(score, 3)))


def _ensure_street_prefix(s):
    # Không tự ý thêm "Đường" vào tên đường — cột Street chỉ cần tên thuần.
    # Hàm giữ nguyên để tương thích, không biến đổi giá trị.
    return s


def _ensure_level4_prefix(s, raw):
    return s


def parse_address_components(raw, admin_names=None, debug=False):
    raw_text = unicodedata.normalize("NFC", str(raw or ""))
    flags = []
    removed_parts = []
    if not raw_text.strip():
        flags.extend(["NO_POI_FOUND", "NO_LEVEL4_FOUND", "LOW_CONFIDENCE"])
        return {
            "raw_address": raw_text, "poi": "", "street": "", "level4": "",
            "level4_primary": "", "level4_sub": "", "removed_parts": [],
            "flags": flags, "confidence": _score_result_v2("", "", flags),
        }

    if extract_phone(raw_text) or re.search(r"(?<!\d)0\d{8,10}(?!\d)", raw_text):
        flags.append("RAW_CONTAINS_PHONE")

    text = _normalize_abbrev_v2(raw_text)
    text = re.sub(r"(?<!\d)0\d{8,10}(?!\d)", " ", text)
    text = re.sub(r"\+?\d[\d\.\-\s]{8,}\d", " ", text)
    text = _strip_intro_v2(text)
    text = _clean_spaces_v2(text)

    admin_aliases = _build_admin_aliases_v2(admin_names or [])
    admin_tail_aliases = _build_admin_tail_aliases_v2(admin_names or [])
    segments = _split_segments_v2(text)
    kept_segments = []
    for seg in segments:
        if _is_note_segment_v2(seg):
            removed_parts.append(_pretty_piece_v2(seg))
            flags.append("NOTE_AFTER_ADDRESS_REMOVED")
            continue
        kept_segments.append(seg)
    segments = kept_segments
    parse_text = ", ".join(segments)

    if _detect_admin_conflict_v2(parse_text, admin_aliases):
        flags.append("RAW_ADMIN_CONFLICT_WITH_COLUMNS")

    pois = _extract_pois_v2(segments, admin_aliases, admin_tail_aliases)
    streets = _extract_streets_v2(segments, admin_aliases)
    if streets:
        flags.append("STREET_FOUND_FROM_HOUSE_NUMBER")
    level4_segments = [
        seg for seg in segments
        if not _starts_with_admin_alias_v2(seg, admin_aliases)
    ]
    level4_items = _extract_level4s_v2(", ".join(level4_segments), admin_aliases)
    level4_items.extend(_extract_unprefixed_level4_after_known_v2(
        segments,
        admin_aliases,
        existing_values=[x["value"] for x in level4_items],
    ))
    level4_items.sort(key=lambda x: x["pos"])
    level4 = "; ".join(x["value"] for x in level4_items)
    level4_primary, level4_sub = _choose_level4_primary_v2(level4_items)
    local_details = _extract_local_details_v2(
        segments,
        admin_aliases,
        existing_values=pois + streets + [level4],
    )
    if local_details:
        flags.append("LOCAL_DETAIL_FOUND")
        streets.extend(local_details)

    if len(pois) > 1:
        flags.append("MULTIPLE_POI_FOUND")
    if len(level4_items) > 1:
        flags.append("MULTIPLE_LEVEL4_FOUND")
    if not pois:
        flags.append("NO_POI_FOUND")
    if not level4:
        flags.append("NO_LEVEL4_FOUND")

    street_parts = _street_removed_parts_v2(parse_text)
    removed_parts.extend(street_parts)
    if street_parts and not pois and not level4 and not streets:
        flags.append("ONLY_STREET_LEVEL_FOUND")

    non_note_segments = [
        seg for seg in segments
        if seg and not _starts_with_admin_alias_v2(seg, admin_aliases)
    ]
    if not pois and not level4 and not street_parts and not non_note_segments:
        flags.append("ONLY_ADMIN_UNITS_FOUND")

    # Tên người thường đứng đầu và không có keyword địa chỉ/POI.
    first_seg = segments[0] if segments else ""
    if first_seg and is_customer_name(first_seg):
        flags.append("RAW_CONTAINS_PERSON_NAME")

    flags = list(dict.fromkeys(flags))
    
    # Ensure street and level 4 prefixes are present
    streets = [_ensure_street_prefix(s) for s in streets if s]
    for item in level4_items:
        item["value"] = _ensure_level4_prefix(item["value"], raw_text)
    level4 = "; ".join(x["value"] for x in level4_items)
    level4_primary, level4_sub = _choose_level4_primary_v2(level4_items)

    poi = " | ".join(pois)
    confidence = _score_result_v2(poi, level4, flags)
    if confidence < 0.75 and "LOW_CONFIDENCE" not in flags:
        flags.append("LOW_CONFIDENCE")
        confidence = _score_result_v2(poi, level4, flags)

    result = {
        "raw_address": raw_text,
        "poi": poi,
        "street": " | ".join(streets),
        "level4": level4,
        "level4_primary": level4_primary,
        "level4_sub": level4_sub,
        "removed_parts": list(dict.fromkeys(removed_parts)),
        "flags": flags,
        "confidence": confidence,
    }
    if debug:
        print("[DEBUG v2]", result)
    return result


def extract_level4_and_poi(raw, admin_names, debug=False):
    parsed = parse_address_components(raw, admin_names, debug=debug)
    return parsed["poi"], parsed["level4"]


def refine_detail(raw, admin_names, strip_names=True):
    """Bóc chi tiết (trừ đơn vị hành chính) rồi áp luật lọc rác -> chi tiết sạch.
    strip_names mặc định True vì file có sẵn đơn vị hành chính (tín hiệu mạnh)."""
    return apply_number_rules(clean_detail(raw, admin_names), strip_names=strip_names)


def split_poi_and_detail(detail):
    if not detail or detail == "-":
        return "", ""
    parts = [p.strip() for p in detail.split(",") if p.strip()]
    pois = []
    streets = []
    
    strong_poi = {
        'ngan', 'truong', 'benh', 'vien', 'uy', 'ban', 'ty', 'cty', 'chua', 'tram', 
        'toa', 'building', 'ks', 'sieu', 'cho', 'ubnd', 'mam', 'non', 'tieu', 'hoc', 
        'trung', 'thpt', 'thcs', 'thuoc', 'quay', 'phong', 'kham', 'pk', 'sach', 
        'gara', 'garage', 'dai', 'ly', 'dl', 'hkd', 'y', 'te', 'buu', 'dien', 
        'mieu', 'congty', 'shop', 'kho', 'xuong', 'ben', 'cang'
    }
    
    for p in parts:
        words_raw = p.split()
        p_norm = [strip_diacritics(w.lower()).strip('.,-–') for w in words_raw]
        
        has_poi = False
        for w_n, w_r in zip(p_norm, words_raw):
            w_r_lower = w_r.lower().strip('.,-–')
            if w_n in strong_poi or w_r_lower == 'sở':
                has_poi = True
                break
        
        # Check phrases
        p_norm_str = ' ' + ' '.join(p_norm) + ' '
        phrases = [
            ' nha hang ', ' nha nghi ', ' nha sach ', ' nha thuoc ', ' nha van hoa ', 
            ' nha may ', ' nha ga ', ' nha tho ', ' nha chua ', ' nha tro ', ' nha khach ',
            ' khu cong nghiep ', ' kcn ', ' khu che xuat ', ' kcx ', ' khu dan cu ', 
            ' kdc ', ' khu do thi ', ' kdt ', ' cong vien ', ' ub nd ', ' phong kham ', 
            ' benh vien ', ' ngan hang '
        ]
        has_phrase = any(phrase in p_norm_str for phrase in phrases)
        
        has_street = any(w in {
            'duong', 'pho', 'ngo', 'ngach', 'hem', 'kiet', 'ap', 'thon', 'xom', 'to', 
            'khu', 'kp', 'tdp', 'doi', 'ban', 'buon', 'lang', 'ql', 'tl', 'hl', 'dt',
            'quoc', 'lo', 'tinh', 'lo', 'huong', 'lo', 'dai', 'lo', 'xa', 'lo', 'cao', 'toc'
        } for w in p_norm)
        
        is_poi = (has_poi or has_phrase) and not (has_street and p_norm and p_norm[0] in {'duong', 'pho', 'ngo', 'ngach', 'hem', 'kiet'})
        
        if is_poi:
            pois.append(p)
        else:
            streets.append(p)
            
    return ', '.join(pois), ', '.join(streets)


def build(in_path, out_path, addr_col=None):
    parser = Parser(DATA)
    wb = openpyxl.load_workbook(in_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])

    def find(*names):
        for i, h in enumerate(hdr):
            hn = strip_diacritics(str(h or "").lower())
            for nm in names:
                if nm in hn:
                    return i
        return None

    ci_addr = addr_col if addr_col is not None else find("dia chi", "dia chi goc")
    ci_w = find("phuong/xa", "phuong", "xa/phuong")
    ci_d = find("quan/huyen", "quan", "huyen")
    ci_p = find("tinh")
    out = openpyxl.Workbook()
    o = out.active
    o.title = "Địa chỉ sạch"
    o.append([
        "STT", "Địa chỉ gốc", "Điểm định vị (POI)", "Tên đường", "Cấp 4",
        "Cấp 4 chính", "Cấp 4 phụ", "removed_parts", "flags", "confidence",
        "Phường/Xã", "Quận/Huyện", "Tỉnh/TP",
        "ĐỊA CHỈ SẠCH",
    ])
    stats = {"full": 0, "detail": 0, "new": 0, "n": 0, "input_n": 0, "removed": 0}
    for idx, r in enumerate(rows[1:], 1):
        stats["input_n"] += 1
        raw = r[ci_addr] if ci_addr is not None else ""
        wname = r[ci_w] if ci_w is not None else None
        dname = r[ci_d] if ci_d is not None else None
        pname = r[ci_p] if ci_p is not None else None
        lab = lookup_labels(parser, wname, dname, pname)
        ward = lab.get("ward") or (str(wname).strip() if wname else None)
        dist = lab.get("district") or (str(dname).strip() if dname else None)
        prov = lab.get("province") or (str(pname).strip() if pname else None)
        ward_new = lab.get("ward_new")
        prov_new = lab.get("province_new")
        parsed = parse_address_components(
            raw,
            [wname, dname, pname, ward, dist, prov, ward_new, prov_new]
        )
        poi = parsed["poi"]
        street = parsed["street"]
        level4 = parsed["level4"]
        if not (poi or street or level4):
            stats["removed"] += 1
            continue
        parts = [p for p in [poi, street, level4, ward, dist, prov] if p]
        full = ", ".join(parts)
        o.append([
            idx, raw, poi, street, level4, parsed["level4_primary"], parsed["level4_sub"],
            "; ".join(parsed["removed_parts"]), "; ".join(parsed["flags"]),
            parsed["confidence"], ward, dist, prov, full,
        ])
        stats["n"] += 1
        if ward and dist and prov:
            stats["full"] += 1
        if poi or street or level4:
            stats["detail"] += 1
        if ward_new and prov_new:
            stats["new"] += 1
    out.save(out_path)
    return stats


if __name__ == "__main__":
    inp = sys.argv[1]
    outp = sys.argv[2]
    st = build(inp, outp)
    n = st["n"]
    denom = max(n, 1)
    print(f"Tong dong goc:             {st.get('input_n', n)}")
    print(f"Dong giu lai:              {n}")
    print(f"Loai vi thieu chi tiet:    {st.get('removed', 0)}")
    print(f"Du 3 cap (Xa+Huyen+Tinh):  {st['full']} = {100*st['full']/denom:.1f}%")
    print(f"Co POI/duong/cap 4:        {st['detail']} = {100*st['detail']/denom:.1f}%")
    print(f"Map duoc he 34 tinh moi:   {st['new']} = {100*st['new']/denom:.1f}%")
