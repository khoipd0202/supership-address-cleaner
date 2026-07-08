# -*- coding: utf-8 -*-
"""
build_reference.py — Dựng BỘ DATA ĐỊA CHỈ CHUẨN (reference dataset) từ dữ liệu đơn bẩn.

Mục tiêu: data nền để về sau chạy API địa chỉ (giống lớp dữ liệu của Vietmap).
  * KHÔNG giữ số nhà / hẻm / ngõ / ngách.
  * Giữ: Đường (chỉ tên) · Thôn/Xóm/Ấp/Khu phố · Xã/Phường · Huyện/Quận · Tỉnh/TP
    + ánh xạ hệ 34 tỉnh mới.
  * Đơn vị hành chính chuẩn hoá theo danh mục (vn_units_data.json).
  * LOẠI TRÙNG: mỗi điểm địa chỉ duy nhất 1 dòng, sắp xếp phân cấp.
"""
import os
import re
import sys
import glob
import unicodedata
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from parse_address import Parser, lookup_labels, strip_diacritics  # noqa: E402

DATA = os.path.join(HERE, "vn_units_data.json")


def _titlecase_vn(s):
    def fix(w):
        if not w:
            return w
        if any(ch.isdigit() for ch in w):
            return w.lower()
        return w[0].upper() + w[1:].lower()
    parts = re.split(r"(\s+|/)", s)
    return "".join(fix(p) if p.strip() and p != "/" else p for p in parts)


NOISE = re.compile(r"^\s*(none|null|nan|-|\.|,)?\s*$", re.I)


def clean_val(v):
    if v is None:
        return ""
    s = unicodedata.normalize("NFC", str(v)).strip()
    if NOISE.match(s):
        return ""
    # bỏ số nhà/hẻm/ngõ/ngách còn sót ở đầu tên đường
    s = re.sub(r"^(số\s*nhà|số|sn|hẻm|hem|ngõ|ngo|ngách|ngach|kiệt|kiet|lô|lo)\b[\s.:]*",
               "", s, flags=re.I)
    s = re.sub(r"^[\d/\-.\s]+", "", s)               # số dẫn đầu
    s = re.sub(r"(?<=\s)[̀-ͯ]+", "", s)               # dấu thanh mồ côi
    s = re.sub(r"\s+", " ", s).strip(" ,.-")
    if NOISE.match(s):
        return ""
    return _titlecase_vn(s)


def col_finder(hdr):
    norm = [strip_diacritics(str(h or "").lower()) for h in hdr]

    def find(keys, avoid=()):
        for i, h in enumerate(norm):
            if any(a in h for a in avoid):
                continue
            if all(k in h for k in keys):
                return i
        return None
    return find


def build(files, out_path):
    parser = Parser(DATA)
    seen = set()
    records = []
    n_in = 0
    for f in files:
        wb = openpyxl.load_workbook(f, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            continue
        find = col_finder(rows[0])
        ci_w = find(["phuong"], avoid=("tach", "moi"))
        ci_d = find(["quan"], avoid=("tach", "moi")) or find(["huyen"], avoid=("tach", "moi"))
        ci_p = find(["tinh"], avoid=("tach", "moi"))
        ci_st = find(["duong"])           # Tên đường (tách)
        ci_ha = find(["thon"]) or find(["xom"])   # Thôn/Xóm/Ấp (tách)
        for r in rows[1:]:
            n_in += 1
            wname = r[ci_w] if ci_w is not None else None
            dname = r[ci_d] if ci_d is not None else None
            pname = r[ci_p] if ci_p is not None else None
            lab = lookup_labels(parser, wname, dname, pname)
            prov = lab.get("province") or clean_admin(pname)
            dist = lab.get("district") or clean_admin(dname)
            ward = lab.get("ward") or clean_admin(wname)
            ward_new = lab.get("ward_new") or ""
            prov_new = lab.get("province_new") or ""
            street = clean_val(r[ci_st]) if ci_st is not None else ""
            hamlet = clean_val(r[ci_ha]) if ci_ha is not None else ""
            if not (prov or dist or ward):
                continue
            key = (prov, dist, ward, hamlet, street)
            if key in seen:
                continue
            seen.add(key)
            records.append((prov, dist, ward, hamlet, street, ward_new, prov_new))
    # sắp xếp phân cấp
    records.sort(key=lambda x: (x[0], x[1], x[2], x[3], x[4]))
    out = openpyxl.Workbook()
    o = out.active
    o.title = "Data địa chỉ chuẩn"
    o.append(["Tỉnh/Thành phố", "Huyện/Quận", "Xã/Phường",
              "Thôn/Xóm/Ấp/Khu phố", "Đường",
              "Xã/Phường MỚI (34 tỉnh)", "Tỉnh/TP MỚI (34 tỉnh)"])
    for rec in records:
        o.append(list(rec))
    out.save(out_path)
    return n_in, len(records)


def clean_admin(v):
    if v is None:
        return ""
    s = unicodedata.normalize("NFC", str(v)).strip()
    return "" if NOISE.match(s) else s


if __name__ == "__main__":
    files = sorted(glob.glob(os.path.join(HERE, "data", "*_parsed.xlsx")))
    outp = sys.argv[1] if len(sys.argv) > 1 else os.path.join(HERE, "DATA_DIA_CHI_CHUAN.xlsx")
    n_in, n_out = build(files, outp)
    print(f"So file nguon: {len(files)}")
    print(f"Dong dau vao: {n_in}")
    print(f"Dong sau loai trung: {n_out}")
