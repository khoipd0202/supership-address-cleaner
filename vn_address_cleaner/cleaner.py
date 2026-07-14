from __future__ import annotations

import io
import os
import re
import time
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from clean_address import DATA as DEFAULT_DATA_PATH
from clean_address import parse_address_components
from parse_address import Parser, base_name, lookup_labels, strip_diacritics

from .models import CleanResult, CleanStats, OUTPUT_HEADERS

# Flags khiến một dòng bị đưa vào sheet "Cần kiểm tra" (kèm lý do hiển thị)
REVIEW_SHEET_FLAGS = {
    "MISSING_ADMIN_COLUMN": "Thiếu/không khớp cột hành chính",
    "LLM_SKIPPED_QUOTA": "Dòng mơ hồ chưa qua LLM (hết quota)",
    "STRICT_DROPPED_DETAIL": "Rule loại chi tiết vì nghi vấn",
    "RULE_DETAIL_DROPPED_AFTER_LLM_REJECT": "LLM không xác nhận được chi tiết của rule",
    "LLM_NO_VALID_DETAIL": "LLM không tìm được chi tiết hợp lệ",
    "NEED_REVIEW_ADMIN_VERSION": "Địa chỉ có yếu tố cũ/mới, cần review",
    "RAW_HAS_OLD_AND_NEW_ADDRESS": "Raw chứa cả địa chỉ cũ và mới",
    "RAW_ADMIN_CONFLICT_WITH_COLUMNS": "Raw mâu thuẫn với cột hành chính",
    "MULTIPLE_ADDRESS_BLOCKS": "Raw chứa nhiều khối địa chỉ",
    "UNKNOWN_LONG_SEGMENT_BEFORE_ADMIN": "Cụm dài không nhận diện được",
    "LLM_MISSING_IN_RESPONSE": "LLM bỏ sót dòng trong phản hồi (đã retry, vẫn thiếu)",
    "LLM_BATCH_ERROR": "Lỗi khi gọi LLM, đang dùng kết quả rule",
    "FUZZY_ADMIN_TRIMMED": "Đã cắt admin gõ sai khỏi tên đường (đối chiếu địa chỉ gốc)",
    "ADMIN_FILLED_FROM_RAW": "Cột hành chính trống, đã bóc từ địa chỉ thô - cần xác nhận",
}

REVIEW_SHEET_HEADERS = (
    "Dòng gốc", "Địa chỉ gốc", "Lý do cần kiểm tra", "Flags",
    "POI", "Tên đường", "Cấp 4", "Phường/Xã", "Quận/Huyện", "Tỉnh/TP", "Confidence",
)

DROPPED_SHEET_HEADERS = ("Dòng gốc", "Địa chỉ gốc", "Lý do")


def dedupe_key(raw: Any, ward: Any, district: Any, province: Any) -> tuple[str, str, str, str]:
    """Khóa lọc trùng: địa chỉ thô + 3 cột hành chính (bỏ dấu, chuẩn hóa khoảng trắng)."""
    def norm(value: Any) -> str:
        text = strip_diacritics(str(value or "").lower())
        return " ".join(re.sub(r"[^a-z0-9]+", " ", text).split())

    return (norm(raw), norm(ward), norm(district), norm(province))


_ADMIN_SORT_PREFIXES = (
    "thanh pho", "thi tran", "thi xa", "phuong", "quan", "huyen",
    "tinh", "xa", "tp", "tt", "tx",
)


def _natural_sort_token(value: Any, *, strip_admin_prefix: bool = False) -> tuple[Any, ...]:
    """Khóa sắp xếp không dấu, số tự nhiên; giá trị trống luôn nằm cuối."""
    text = strip_diacritics(str(value or "").lower())
    text = " ".join(re.sub(r"[^a-z0-9]+", " ", text).split())
    if not text:
        return (1, ())
    if strip_admin_prefix:
        for prefix in _ADMIN_SORT_PREFIXES:
            if text.startswith(prefix + " "):
                text = text[len(prefix):].strip()
                break
    natural = tuple(
        (0, int(part)) if part.isdigit() else (1, part)
        for part in re.split(r"(\d+)", text)
        if part
    )
    return (0, natural)


def admin_sort_key(province: Any, district: Any, ward: Any) -> tuple[Any, ...]:
    """Sắp theo Tỉnh/TP → Quận/Huyện → Phường/Xã, bỏ qua tiền tố hành chính."""
    return (
        _natural_sort_token(province, strip_admin_prefix=True),
        _natural_sort_token(district, strip_admin_prefix=True),
        _natural_sort_token(ward, strip_admin_prefix=True),
    )


def sort_worksheet_by_admin(sheet) -> bool:
    """Sắp các dòng dữ liệu của worksheet theo 3 cột hành chính nếu có."""
    if sheet.max_row <= 2:
        return False

    headers = [
        " ".join(re.sub(
            r"[^a-z0-9]+", " ", strip_diacritics(str(cell.value or "").lower())
        ).split())
        for cell in sheet[1]
    ]

    def find_header(candidates: set[str]) -> int | None:
        return next((idx for idx, header in enumerate(headers) if header in candidates), None)

    province_idx = find_header({"tinh tp", "tinh thanh pho", "tinh"})
    district_idx = find_header({"quan huyen", "quan", "huyen"})
    ward_idx = find_header({"phuong xa", "xa phuong", "phuong", "xa"})
    if province_idx is None:
        return False

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    rows.sort(key=lambda row: admin_sort_key(
        row[province_idx] if province_idx < len(row) else "",
        row[district_idx] if district_idx is not None and district_idx < len(row) else "",
        row[ward_idx] if ward_idx is not None and ward_idx < len(row) else "",
    ))
    for row_idx, values in enumerate(rows, 2):
        for col_idx, value in enumerate(values, 1):
            sheet.cell(row=row_idx, column=col_idx).value = value
    return True


def append_stats_sheet(workbook, stats: CleanStats) -> None:
    sheet = workbook.create_sheet("Thống kê")
    sheet.append(["Chỉ số", "Giá trị"])
    unique_n = stats.input_n - stats.duplicates
    rows = [
        ("Tổng dòng đầu vào", stats.input_n),
        ("Dòng trùng lặp đã bỏ", stats.duplicates),
        ("Dòng đưa vào xử lý (sau lọc trùng)", unique_n),
        ("Dòng bị loại (không bóc được thông tin)", stats.removed),
        ("Dòng xuất ra (đã tách thành phần)", stats.output_n),
        ("Dòng đủ cấp hành chính", stats.full_admin),
        ("Dòng địa chỉ 2 cấp mới (sau sáp nhập)", stats.mapped_new),
        ("Dòng cần kiểm tra tay", stats.review_n),
    ]
    for row in rows:
        sheet.append(list(row))
    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 12


class ColumnNotFoundError(ValueError):
    pass


def _flex_admin_key(value: Any) -> str:
    """Chuẩn hóa tên đơn vị để so khớp linh hoạt: bỏ dấu, bỏ dấu câu
    (Văn Miếu-Quốc Tử Giám == Văn Miếu - Quốc Tử Giám), bỏ tiền tố loại."""
    text = strip_diacritics(str(value or "").lower())
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    for prefix in ("thanh pho", "thi tran", "thi xa", "phuong", "tinh", "xa", "tp", "tt", "tx"):
        if text == prefix:
            return ""
        if text.startswith(prefix + " "):
            text = text[len(prefix):].strip()
            break
    return text.replace(" ", "")


def lookup_new_admin(parser, ward: Any, province: Any) -> dict[str, str] | None:
    """Đối chiếu (Phường/Xã, Tỉnh/TP) với danh mục 2 cấp mới (sau sáp nhập 2025).

    Trả về tên chuẩn nếu cặp này tồn tại trong hệ mới, ngược lại None.
    """
    if not ward or not province:
        return None
    prov_key = _flex_admin_key(province)
    if not prov_key:
        return None
    for name, pid in parser.new_prov_by_name.items():
        if _flex_admin_key(name) != prov_key:
            continue
        ward_key = _flex_admin_key(ward)
        for unit in parser.new_wards_by_prov.get(pid, []):
            if _flex_admin_key(unit.name) == ward_key:
                return {"ward": unit.name, "province": name}
        return None
    return None


def resolve_admin(parser, raw: Any, ward: Any, district: Any, province: Any):
    """Chuẩn hóa 3 cột hành chính, có 2 fallback khi người dùng nhập thiếu:

    1. Địa chỉ 2 cấp mới (không có Quận/Huyện) -> khớp danh mục mới.
    2. Cột trống -> bóc bổ sung từ chính địa chỉ thô (KHÔNG ghi đè cột có sẵn).

    Trả về (labels, ward, district, province, extra_flags).
    """
    labels = lookup_labels(parser, ward, district, province)
    clean_ward = labels.get("ward") or (str(ward).strip() if ward else "")
    clean_district = labels.get("district") or (str(district).strip() if district else "")
    clean_province = labels.get("province") or (str(province).strip() if province else "")
    flags: list[str] = []

    # (1) Địa chỉ 2 cấp mới
    if not clean_district:
        new_admin = lookup_new_admin(parser, clean_ward or ward, clean_province or province)
        if new_admin:
            clean_ward = new_admin["ward"]
            if not clean_province:
                clean_province = new_admin["province"]
            flags.append("ADMIN_NEW_FORMAT")

    # (2) Cột thiếu -> bóc từ địa chỉ thô
    if "ADMIN_NEW_FORMAT" not in flags and (
        not clean_ward or not clean_district or not clean_province
    ):
        try:
            parsed_raw = parser.parse(str(raw or "")) or {}
        except Exception:
            parsed_raw = {}
        filled = False
        if not clean_ward and parsed_raw.get("ward"):
            clean_ward = str(parsed_raw["ward"])
            filled = True
        if not clean_district and parsed_raw.get("district"):
            clean_district = str(parsed_raw["district"])
            filled = True
        if not clean_province and parsed_raw.get("province"):
            clean_province = str(parsed_raw["province"])
            filled = True
        if filled:
            flags.append("ADMIN_FILLED_FROM_RAW")
            if not clean_district:
                new_admin = lookup_new_admin(parser, clean_ward, clean_province)
                if new_admin:
                    clean_ward = new_admin["ward"]
                    flags.append("ADMIN_NEW_FORMAT")

    if not clean_ward or not clean_province or (
        not clean_district and "ADMIN_NEW_FORMAT" not in flags
    ):
        flags.append("MISSING_ADMIN_COLUMN")
    return labels, clean_ward, clean_district, clean_province, flags


def _default_data_path() -> str:
    packaged = Path(__file__).resolve().parent / "data" / "vn_units_data.json"
    if packaged.exists():
        return str(packaged)
    return str(DEFAULT_DATA_PATH)


def _norm_header(value: Any) -> str:
    text = strip_diacritics(str(value or "").lower())
    return " ".join(text.replace("/", " ").replace("-", " ").split())


def _find_column(headers: list[Any], *needles: str) -> int | None:
    normalized = [_norm_header(h) for h in headers]
    for idx, header in enumerate(normalized):
        for needle in needles:
            n = _norm_header(needle)
            if n and (n == header or n in header):
                return idx
    return None


def detect_columns(headers: list[Any]) -> dict[str, int]:
    columns = {
        "address": _find_column(headers, "Địa chỉ", "Địa chỉ chi tiết", "Địa chỉ gốc"),
        "ward": _find_column(headers, "Phường/Xã", "Xã/Phường", "Phường", "Xã"),
        "district": _find_column(headers, "Quận/Huyện", "Huyện/Quận", "Quận", "Huyện"),
        "province": _find_column(headers, "Tỉnh/Thành Phố", "Tỉnh/TP", "Tỉnh"),
    }
    missing = [name for name, idx in columns.items() if idx is None]
    if missing:
        available = ", ".join(str(h or "") for h in headers)
        raise ColumnNotFoundError(f"Missing required columns: {', '.join(missing)}. Headers: {available}")
    return {k: int(v) for k, v in columns.items() if v is not None}


def _component_compare_key(value: Any) -> str:
    text = strip_diacritics(str(value or "").lower())
    text = re.sub(r"\b(?:duong|pho|hem|ngo|ngach|kiet)\b", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def _merge_validated_llm_result(rule_result: CleanResult, validated: dict[str, Any]) -> CleanResult:
    values = {
        "poi": rule_result.poi,
        "street": rule_result.street,
        "level4": rule_result.level4,
    }
    changed = False
    dropped_after_reject = False
    validated_flags = list(validated.get("flags") or [])
    for field in ("poi", "street", "level4"):
        value = validated.get(field) or ""
        if not value:
            # LLM không xác nhận được field này. Nếu validator đã bác giá trị
            # LLM đề xuất (LLM_<FIELD>_*_REJECTED) thì giá trị rule vốn mơ hồ
            # cũng không được tin: bỏ luôn thay vì xuất bừa (vd "Đường Liên").
            reject_prefix = f"LLM_{field.upper()}_"
            was_rejected = any(
                flag.startswith(reject_prefix) and flag.endswith("_REJECTED")
                for flag in validated_flags
            )
            if values[field] and was_rejected:
                values[field] = ""
                changed = True
                dropped_after_reject = True
            continue
        current = values[field]
        if field in {"poi", "street"} and current:
            current_key = _component_compare_key(current)
            proposed_key = _component_compare_key(value)
            if (
                len(current_key.split()) >= 2
                and current_key != proposed_key
                and current_key in proposed_key
            ):
                # Không cho LLM nối lại phần vị trí/admin mà rule đã cắt đúng
                # ("Công ty Hương Sen" -> "... sau bến xe...", "Lê Lai" ->
                # "Lê Lai Hợp Thành").
                continue
        if current and _component_compare_key(value) == _component_compare_key(current):
            continue
        if value != current:
            changed = True
        values[field] = value

    flags = list(rule_result.flags)
    if dropped_after_reject and "RULE_DETAIL_DROPPED_AFTER_LLM_REJECT" not in flags:
        flags.append("RULE_DETAIL_DROPPED_AFTER_LLM_REJECT")
    for flag in validated.get("flags") or []:
        if flag not in flags:
            flags.append(flag)
    if changed and "LLM_CHANGED_RULE_DETAIL" not in flags:
        flags.append("LLM_CHANGED_RULE_DETAIL")

    return CleanResult(
        poi=values["poi"],
        street=values["street"],
        level4=values["level4"],
        ward=validated.get("ward") or rule_result.ward,
        district=validated.get("district") or rule_result.district,
        province=validated.get("province") or rule_result.province,
        confidence=float(validated.get("confidence") or rule_result.confidence or 0.0),
        flags=tuple(flags),
    )


class AddressCleaner:
    """Clean SuperShip-style Excel files into a compact six-column workbook."""

    def __init__(
        self,
        data_path: str | os.PathLike[str] | None = None,
        use_cerebras: bool = False,
        cerebras_api_key: str | None = None,
        cerebras_model: str | None = None,
        queue_all: bool = False,
    ):
        self.data_path = str(data_path or _default_data_path())
        self.parser = Parser(self.data_path)
        self.use_cerebras = bool(use_cerebras or cerebras_api_key)
        # Queue mode: xử lý TOÀN BỘ dòng mơ hồ tuần tự, không cap, tự giãn nhịp
        self.queue_all = bool(queue_all or os.environ.get("CEREBRAS_QUEUE_ALL"))
        self.cerebras_api_key = (
            cerebras_api_key
            or os.environ.get("CEREBRAS_API_KEY")
            or os.environ.get("CEREBRAS_API_TOKEN")
        )
        self.cerebras_model = cerebras_model or os.environ.get("CEREBRAS_MODEL") or "gpt-oss-120b"

    def _lookup_new_admin(self, ward: Any, province: Any) -> dict[str, str] | None:
        return lookup_new_admin(self.parser, ward, province)

    def clean(
        self,
        raw_address: Any,
        ward: Any = None,
        district: Any = None,
        province: Any = None,
    ) -> CleanResult:
        labels, clean_ward, clean_district, clean_province, extra_flags = resolve_admin(
            self.parser, raw_address, ward, district, province
        )
        ward_new = labels.get("ward_new")
        province_new = labels.get("province_new")

        parsed = parse_address_components(
            raw_address,
            [ward, district, province, clean_ward, clean_district, clean_province, ward_new, province_new],
        )
        flags = list(parsed.get("flags") or ())
        for flag in extra_flags:
            if flag not in flags:
                flags.append(flag)
        return CleanResult(
            poi=parsed.get("poi") or "",
            street=parsed.get("street") or "",
            level4=parsed.get("level4") or "",
            ward=clean_ward or "",
            district=clean_district or "",
            province=clean_province or "",
            confidence=float(parsed.get("confidence") or 0.0),
            flags=tuple(flags),
        )

    def clean_excel(
        self,
        input_path: str | os.PathLike[str],
        output_path: str | os.PathLike[str],
        *,
        include_empty_rows: bool = False,
        split_components: bool = True,
        sheet_name: str | None = None,
    ) -> CleanStats:
        workbook = openpyxl.load_workbook(input_path, data_only=True)
        source = workbook[sheet_name] if sheet_name else workbook.active
        rows = list(source.iter_rows(values_only=True))
        output, stats = self.clean_rows_to_workbook(
            rows,
            include_empty_rows=include_empty_rows,
            split_components=split_components,
        )
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        output.save(output_path)
        return stats

    def clean_bytes(
        self,
        workbook_bytes: bytes,
        *,
        include_empty_rows: bool = False,
        split_components: bool = True,
        sheet_name: str | None = None,
    ) -> tuple[bytes, CleanStats]:
        workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True)
        source = workbook[sheet_name] if sheet_name else workbook.active
        rows = list(source.iter_rows(values_only=True))
        output, stats = self.clean_rows_to_workbook(
            rows,
            include_empty_rows=include_empty_rows,
            split_components=split_components,
        )
        buffer = io.BytesIO()
        output.save(buffer)
        return buffer.getvalue(), stats

    def clean_rows_to_workbook(
        self,
        rows: list[tuple[Any, ...]],
        *,
        include_empty_rows: bool = False,
        split_components: bool = True,
    ) -> tuple[openpyxl.Workbook, CleanStats]:
        output = openpyxl.Workbook()
        sheet = output.active
        sheet.title = "Địa chỉ sạch"
        sheet.append(list(OUTPUT_HEADERS))
        self._style_output(sheet)

        stats = CleanStats()
        if not rows:
            return output, stats

        headers = list(rows[0])
        columns = detect_columns(headers)

        results = {}
        raws: dict[int, str] = {}
        llm_candidates = []
        seen_keys: dict[tuple[str, str, str, str], int] = {}
        duplicate_of: dict[int, int] = {}

        for idx, row in enumerate(rows[1:], 1):
            stats.input_n += 1
            raw = row[columns["address"]] if columns["address"] < len(row) else ""
            raw_text = "" if raw is None else str(raw)
            raws[idx] = raw_text
            ward = row[columns["ward"]] if columns["ward"] < len(row) else ""
            district = row[columns["district"]] if columns["district"] < len(row) else ""
            province = row[columns["province"]] if columns["province"] < len(row) else ""

            # Lọc trùng TRƯỚC khi parse: cùng địa chỉ thô + 3 cột hành chính
            key = dedupe_key(raw_text, ward, district, province)
            if any(key) and key in seen_keys:
                duplicate_of[idx] = seen_keys[key]
                stats.duplicates += 1
                continue
            if any(key):
                seen_keys[key] = idx

            res = self.clean(raw, ward, district, province)
            results[idx] = res

            if self.use_cerebras:
                from .llm import should_send_to_llm
                rule_res_dict = {
                    "poi": res.poi,
                    "street": res.street,
                    "level4": res.level4,
                    "ward": res.ward,
                    "district": res.district,
                    "province": res.province,
                    "confidence": res.confidence,
                    "flags": list(res.flags),
                    "raw_address": raw_text,
                }
                if should_send_to_llm(rule_res_dict):
                    llm_candidates.append({
                        "row_id": idx,
                        "raw_address": raw_text,
                        "ward": res.ward or "",
                        "district": res.district or "",
                        "province": res.province or "",
                        "source_columns": {
                            "ward": ward or "",
                            "district": district or "",
                            "province": province or "",
                        },
                        "rule_candidates": rule_res_dict,
                        "flags": list(res.flags),
                        "confidence": res.confidence,
                    })

        if self.use_cerebras and llm_candidates:
            from .llm import CerebrasRateLimitError, dedupe_llm_rows, parse_rows_with_cerebras, validate_llm_result

            unique_candidates, dedup_map = dedupe_llm_rows(llm_candidates)

            def _flag_llm_rows(items, flag):
                from dataclasses import replace as _dc_replace
                for item in items:
                    _, item_dedup = dedupe_llm_rows([item])
                    item_key = next(iter(item_dedup))
                    for rid in dedup_map.get(item_key, []):
                        res = results[rid]
                        if flag not in res.flags:
                            results[rid] = _dc_replace(
                                res, flags=tuple(list(res.flags) + [flag])
                            )

            def _flag_skipped_llm(items):
                _flag_llm_rows(items, "LLM_SKIPPED_QUOTA")

            def _apply_llm_response(batch, response):
                """Merge kết quả LLM vào results; trả về các dòng bị model bỏ sót."""
                source_by_row_id = {item["row_id"]: item for item in batch}
                returned_ids = set()
                for result in response.get("results", []):
                    source_item = source_by_row_id.get(result.get("row_id"))
                    if not source_item:
                        continue
                    returned_ids.add(result.get("row_id"))
                    validated = validate_llm_result(result, source_item)
                    _, source_dedup = dedupe_llm_rows([source_item])
                    key = next(iter(source_dedup))
                    for original_row_id in dedup_map.get(key, []):
                        results[original_row_id] = _merge_validated_llm_result(
                            results[original_row_id],
                            validated,
                        )
                return [item for item in batch if item["row_id"] not in returned_ids]

            batch_size = max(1, min(int(os.environ.get("CEREBRAS_BATCH_SIZE", "5")), 20))
            if self.queue_all:
                # QUEUE MODE: không cap, xử lý tuần tự toàn bộ với pacing.
                # Cerebras free tier ~30 request/phút -> mặc định 2.2s/request.
                pacing = max(0.0, float(os.environ.get("CEREBRAS_PACING_SEC", "2.2")))
                total_batches = -(-len(unique_candidates) // batch_size)
                est_min = total_batches * (pacing + 1.5) / 60
                max_min = total_batches * (pacing + 60 + 225) / 60
                print(
                    f"QUEUE MODE: {len(unique_candidates)} dòng mơ hồ, "
                    f"{total_batches} batch x {batch_size} dòng, pacing {pacing}s/request."
                )
                print(
                    f"Ước tính thời gian LLM: ~{est_min:.0f} phút (bình thường), "
                    f"tối đa ~{max_min:.0f} phút (trường hợp xấu nhất: timeout + backoff mọi batch)."
                )
            else:
                pacing = 0.0
                max_rows = max(1, min(int(os.environ.get("CEREBRAS_MAX_ROWS_PER_RUN", "150")), 200))
                if len(unique_candidates) > max_rows:
                    skipped = len(unique_candidates) - max_rows
                    print(f"Cerebras free-tier cap: chỉ gửi {max_rows} dòng mơ hồ, bỏ qua {skipped} dòng còn lại (dùng --queue-all để xử lý hết).")
                    _flag_skipped_llm(unique_candidates[max_rows:])
                    unique_candidates = unique_candidates[:max_rows]
                total_batches = -(-len(unique_candidates) // batch_size)

            def _call_with_backoff(batch):
                """Gọi API; queue mode gặp 429 thì chờ backoff (15/30/60/120s) rồi thử lại."""
                attempts = 0
                while True:
                    try:
                        return parse_rows_with_cerebras(
                            batch,
                            model=self.cerebras_model,
                            api_key=self.cerebras_api_key,
                        )
                    except CerebrasRateLimitError:
                        if not self.queue_all:
                            raise
                        attempts += 1
                        if attempts > 4:
                            raise
                        wait = min(120, 15 * (2 ** (attempts - 1)))
                        print(f"Rate limit, hàng chờ tạm nghỉ {wait}s rồi tiếp tục...")
                        time.sleep(wait)

            retry_items: list[dict[str, Any]] = []
            rate_limited = False
            done_batches = 0
            for i in range(0, len(unique_candidates), batch_size):
                batch = unique_candidates[i : i + batch_size]
                started = time.time()
                try:
                    response = _call_with_backoff(batch)
                    missing = _apply_llm_response(batch, response)
                    if missing:
                        retry_items.extend(missing)
                except CerebrasRateLimitError as e:
                    remaining = len(unique_candidates) - i
                    print(f"Cerebras API chạm rate/quota limit, dừng gọi API cho {remaining} dòng còn lại: {e}")
                    _flag_skipped_llm(unique_candidates[i:])
                    rate_limited = True
                    break
                except Exception as e:
                    print(f"Error calling Cerebras API: {e}. Sẽ thử lại các dòng trong batch này.")
                    retry_items.extend(batch)
                done_batches += 1
                if self.queue_all:
                    elapsed = time.time() - started
                    if pacing > elapsed:
                        time.sleep(pacing - elapsed)
                    if done_batches % 20 == 0:
                        eta_min = (total_batches - done_batches) * (pacing + 1.5) / 60
                        print(f"Queue: {done_batches}/{total_batches} batch xong, còn ~{eta_min:.0f} phút.")

            # Thử lại MỘT lượt các dòng bị model bỏ sót/batch lỗi, batch nhỏ hơn
            if retry_items and not rate_limited:
                print(f"Thử lại {len(retry_items)} dòng LLM bị bỏ sót/lỗi...")
                retry_batch_size = max(1, min(3, batch_size))
                for i in range(0, len(retry_items), retry_batch_size):
                    batch = retry_items[i : i + retry_batch_size]
                    started = time.time()
                    try:
                        response = _call_with_backoff(batch)
                        still_missing = _apply_llm_response(batch, response)
                        if still_missing:
                            _flag_llm_rows(still_missing, "LLM_MISSING_IN_RESPONSE")
                    except CerebrasRateLimitError as e:
                        print(f"Rate limit khi retry, dừng: {e}")
                        _flag_skipped_llm(retry_items[i:])
                        break
                    except Exception as e:
                        print(f"Retry thất bại: {e}")
                        _flag_llm_rows(batch, "LLM_BATCH_ERROR")
                    if self.queue_all:
                        elapsed = time.time() - started
                        if pacing > elapsed:
                            time.sleep(pacing - elapsed)
            elif retry_items:
                _flag_skipped_llm(retry_items)

        clean_output_rows: list[list[Any]] = []
        review_rows: list[list[Any]] = []
        dropped_rows: list[list[Any]] = []
        for idx in range(1, len(rows)):
            raw_text = raws.get(idx, "")
            if idx in duplicate_of:
                dropped_rows.append([
                    idx + 1, raw_text,
                    f"Trùng lặp với dòng {duplicate_of[idx] + 1} (cùng địa chỉ và 3 cột hành chính)",
                ])
                continue
            result = results[idx]

            reasons = [
                label for flag, label in REVIEW_SHEET_FLAGS.items()
                if flag in result.flags
            ]
            if reasons:
                review_rows.append([
                    idx + 1,  # số dòng Excel gốc (1 = header)
                    raw_text,
                    "; ".join(reasons),
                    ", ".join(result.flags),
                    result.poi, result.street, result.level4,
                    result.ward, result.district, result.province,
                    result.confidence,
                ])

            if "ADMIN_NEW_FORMAT" in result.flags:
                stats.mapped_new += 1

            if not include_empty_rows and not result.has_detail:
                stats.removed += 1
                drop_reason = "Không bóc được POI/tên đường/cấp 4 từ địa chỉ thô"
                if not raw_text.strip():
                    drop_reason = "Địa chỉ thô để trống"
                elif "LLM_SKIPPED_QUOTA" in result.flags:
                    drop_reason += "; dòng mơ hồ chưa qua LLM (hết quota)"
                elif not self.use_cerebras:
                    from .llm import SEND_LLM_FLAGS
                    if set(result.flags) & SEND_LLM_FLAGS:
                        drop_reason += "; dòng mơ hồ cần LLM (Cerebras chưa bật)"
                dropped_rows.append([idx + 1, raw_text, drop_reason])
                continue

            output_rows = result.as_component_rows() if split_components else [result.as_output_row()]
            for output_row in output_rows:
                clean_output_rows.append(output_row)
            stats.output_n += len(output_rows)
            # Địa chỉ 2 cấp mới: không có Quận/Huyện vẫn là đủ cấp hành chính
            has_full_admin = result.ward and result.province and (
                result.district or "ADMIN_NEW_FORMAT" in result.flags
            )
            if has_full_admin:
                stats.full_admin += len(output_rows)

        clean_output_rows.sort(key=lambda row: admin_sort_key(row[5], row[4], row[3]))
        for row in clean_output_rows:
            sheet.append(row)

        stats.review_n = len(review_rows)
        if review_rows:
            review_sheet = output.create_sheet("Cần kiểm tra")
            review_sheet.append(list(REVIEW_SHEET_HEADERS))
            self._style_output(review_sheet)
            review_rows.sort(key=lambda row: admin_sort_key(row[9], row[8], row[7]))
            for row in review_rows:
                review_sheet.append(row)
            widths = [9, 46, 34, 34, 24, 20, 20, 20, 20, 20, 11]
            for w_idx, width in enumerate(widths, 1):
                review_sheet.column_dimensions[openpyxl.utils.get_column_letter(w_idx)].width = width
            for row in review_sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

        if dropped_rows:
            dropped_sheet = output.create_sheet("Dòng bị loại")
            dropped_sheet.append(list(DROPPED_SHEET_HEADERS))
            self._style_output(dropped_sheet)
            for row in dropped_rows:
                dropped_sheet.append(row)
            for w_idx, width in enumerate([9, 60, 50], 1):
                dropped_sheet.column_dimensions[openpyxl.utils.get_column_letter(w_idx)].width = width
            for row in dropped_sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

        append_stats_sheet(output, stats)
        self._autosize_output(sheet)
        return output, stats

    @staticmethod
    def _style_output(sheet) -> None:
        sheet.freeze_panes = "A2"
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def _autosize_output(sheet) -> None:
        widths = [32, 26, 26, 24, 26, 26]
        for idx, width in enumerate(widths, 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)


def clean_excel(
    input_path: str | os.PathLike[str],
    output_path: str | os.PathLike[str],
    *,
    include_empty_rows: bool = False,
    split_components: bool = True,
    sheet_name: str | None = None,
    data_path: str | os.PathLike[str] | None = None,
    use_cerebras: bool = False,
    cerebras_api_key: str | None = None,
    cerebras_model: str | None = None,
) -> CleanStats:
    return AddressCleaner(
        data_path=data_path,
        use_cerebras=use_cerebras,
        cerebras_api_key=cerebras_api_key,
        cerebras_model=cerebras_model,
    ).clean_excel(
        input_path,
        output_path,
        include_empty_rows=include_empty_rows,
        split_components=split_components,
        sheet_name=sheet_name,
    )


def clean_workbook_bytes(
    workbook_bytes: bytes,
    *,
    include_empty_rows: bool = False,
    split_components: bool = True,
    sheet_name: str | None = None,
    data_path: str | os.PathLike[str] | None = None,
    use_cerebras: bool = False,
    cerebras_api_key: str | None = None,
    cerebras_model: str | None = None,
) -> tuple[bytes, CleanStats]:
    return AddressCleaner(
        data_path=data_path,
        use_cerebras=use_cerebras,
        cerebras_api_key=cerebras_api_key,
        cerebras_model=cerebras_model,
    ).clean_bytes(
        workbook_bytes,
        include_empty_rows=include_empty_rows,
        split_components=split_components,
        sheet_name=sheet_name,
    )
